# <--- hecho por claude code: ANA Network Manager — vistas Fase 1 (superadmin por ahora)
import json

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied, ValidationError
from django.db.models import Count, Q
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse

from django.db.models import ProtectedError

from .models import (VLAN, IPAddress, Device, NetworkInterface,
                     Campus, Edificio, Ubicacion, Gabinete, RedAuditLog)
from .forms import (VLANForm, IPForm, CampusForm, EdificioForm, UbicacionForm,
                    GabineteForm, DeviceForm, InterfaceForm)
from . import services


def _gate(request, escribir=False):
    """<--- hecho por claude code: acceso por PERMISOS (grupo `red`), no solo superuser.
    - superuser: siempre.
    - lectura: permiso `red.ver_red` (grupo red-lectura / soporte).
    - escritura (escribir=True): además `red.administrar_red` (responsable de red).
    Se valida en backend: ocultar botones nunca sustituye este chequeo."""
    u = request.user
    if u.is_superuser:
        return
    if not u.has_perm('red.ver_red'):
        raise PermissionDenied('Acceso restringido: no tienes permiso para ver el módulo de Red.')
    if escribir and not u.has_perm('red.administrar_red'):
        raise PermissionDenied('Acceso de solo lectura: no tienes permiso para modificar la red.')


def puede_administrar_red(user):
    """Helper para templates/vistas: True si puede escribir en el módulo."""
    return user.is_superuser or user.has_perm('red.administrar_red')


def _ip_cliente(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    return (xff.split(',')[0].strip() if xff else request.META.get('REMOTE_ADDR'))


def _audit(request, modulo, registro, accion, antes=None, despues=None, motivo=''):
    RedAuditLog.objects.create(
        usuario=request.user if request.user.is_authenticated else None,
        ip_usuario=_ip_cliente(request), modulo=modulo, registro=str(registro)[:160],
        accion=accion, valores_antes=antes, valores_despues=despues, motivo=motivo[:255])


def _snapshot_rangos(v):
    return {c: str(getattr(v, c) or '') for c in [
        'subred', 'gateway', 'ip_inicial_asignable', 'ip_final_asignable',
        'ip_inicio_reservado', 'ip_fin_reservado', 'ip_inicio_dhcp', 'ip_fin_dhcp',
        'limite_dispositivos', 'porcentaje_alerta']}


# ═══════════════════════ Dashboard ═══════════════════════
@login_required
def dashboard(request):
    _gate(request)
    vlans = list(VLAN.objects.all())
    caps = [(v, services.capacidad_vlan(v)) for v in vlans]
    alertas = [(v, c) for v, c in caps if c['estado'] != 'normal']
    dispositivos_tipo = list(Device.objects.values('tipo').annotate(n=Count('id')).order_by('-n'))
    ctx = {
        'nav_home_url': '/',
        # <--- hecho por claude code: navegación del módulo desde el dashboard
        'modulos': [
            {'t': 'Panel avanzado', 'i': 'ti-chart-histogram', 'c': '#2fb344', 'url': reverse('red:panel')},
            {'t': 'VLANs',       'i': 'ti-router',    'c': '#206bc4', 'url': reverse('red:vlan_list')},
            {'t': 'Ubicaciones', 'i': 'ti-building',  'c': '#f76707', 'url': reverse('red:ubicaciones')},
            {'t': 'Dispositivos','i': 'ti-devices',   'c': '#ae3ec9', 'url': reverse('red:device_list')},
            {'t': 'Switches',    'i': 'ti-switch-3',  'c': '#4263eb', 'url': reverse('red:switch_list')},
            {'t': 'Racks',       'i': 'ti-server-2',  'c': '#495057', 'url': reverse('red:racks_list')},
            {'t': 'Enlaces',     'i': 'ti-link',      'c': '#d6336c', 'url': reverse('red:link_list')},
            {'t': 'Mapa campus', 'i': 'ti-map-2',     'c': '#12b886', 'url': reverse('red:planos_list')},
            {'t': 'Topología',   'i': 'ti-topology-star-3', 'c': '#e8590c', 'url': reverse('red:topologia')},
            {'t': 'Pruebas de red', 'i': 'ti-plug-connected', 'c': '#2fb344', 'url': reverse('red:pruebas')},
            {'t': 'Migración VLAN', 'i': 'ti-arrows-exchange', 'c': '#f76707', 'url': reverse('red:migracion')},
            {'t': 'Buscar',      'i': 'ti-search',    'c': '#4263eb', 'url': reverse('red:buscar')},
            {'t': 'Exportar Excel', 'i': 'ti-file-spreadsheet', 'c': '#12b886', 'url': reverse('red:export_excel')},
            {'t': 'Auditoría',   'i': 'ti-history',   'c': '#7048e8', 'url': reverse('red:auditoria')},
        ],
        'total_vlans': len(vlans),
        'vlans_produccion': sum(1 for v in vlans if v.estado == 'produccion'),
        'total_dispositivos': Device.objects.count(),
        'ips_asignadas': IPAddress.objects.filter(estado='asignada').count(),
        'ips_libres': IPAddress.objects.filter(estado='libre').count(),
        'ips_conflicto': IPAddress.objects.filter(estado='conflicto').count(),
        'caps': caps,
        'alertas': alertas,
        # datos para gráficas (Chart.js)
        'chart_vlan_labels': json.dumps([f'VLAN {v.vlan_id}' for v, _ in caps]),
        'chart_vlan_ocup':   json.dumps([c['pct_ocupacion'] for _, c in caps]),
        'chart_disp_labels': json.dumps([dict(Device.TIPOS).get(d['tipo'], d['tipo']) for d in dispositivos_tipo]),
        'chart_disp_data':   json.dumps([d['n'] for d in dispositivos_tipo]),
    }
    return render(request, 'red/dashboard.html', ctx)


# ═══════════════════════ FASE 4 · Panel avanzado (analítica) ═══════════════════════
@login_required
def panel(request):
    """<--- hecho por claude code: Fase 4 — dashboard avanzado (KPIs + gráficas Chart.js)."""
    _gate(request)
    from django.db.models import Q as _Qp

    # ── KPIs ──
    ip_estados = dict(IPAddress.objects.values_list('estado').annotate(n=Count('id')))
    total_ips = sum(ip_estados.values())
    en_uso = ip_estados.get('asignada', 0) + ip_estados.get('dhcp', 0)
    pct_uso = round(en_uso / total_ips * 100, 1) if total_ips else 0

    # ── Dispositivos por tipo ──
    d_tipo = list(Device.objects.values('tipo').annotate(n=Count('id')).order_by('-n'))
    disp_labels = [dict(Device.TIPOS).get(x['tipo'], x['tipo']) for x in d_tipo]
    disp_data = [x['n'] for x in d_tipo]

    # ── Dispositivos por edificio ──
    d_edif = list(Device.objects.values('ubicacion__edificio__nombre').annotate(n=Count('id')).order_by('-n'))
    edif_labels = [(x['ubicacion__edificio__nombre'] or 'Sin ubicación') for x in d_edif]
    edif_data = [x['n'] for x in d_edif]

    # ── IPs usadas vs libres por VLAN ──
    ipv = list(IPAddress.objects.filter(vlan__isnull=False)
               .values('vlan__vlan_id')
               .annotate(total=Count('id'), libre=Count('id', filter=_Qp(estado='libre')))
               .order_by('vlan__vlan_id'))
    vlan_labels = [f"VLAN {x['vlan__vlan_id']}" for x in ipv]
    vlan_usadas = [x['total'] - x['libre'] for x in ipv]
    vlan_libres = [x['libre'] for x in ipv]

    # ── Estado de capacidad de VLANs ──
    cap = dict(VLAN.objects.values_list('estado_capacidad').annotate(n=Count('id')))
    cap_orden = ['normal', 'advertencia', 'critico', 'agotado']
    cap_labels = ['Normal', 'Advertencia', 'Crítico', 'Agotado']
    cap_data = [cap.get(k, 0) for k in cap_orden]

    # ── Enlaces por tipo ──
    lk = dict(NetworkLink.objects.values_list('tipo').annotate(n=Count('id')))
    enl_map = {'rj45': 'RJ45 (cobre)', 'fibra': 'Fibra', 'radio': 'Radio', 'virtual': 'Virtual'}
    enl_labels = [enl_map.get(k, k) for k in lk.keys()]
    enl_data = list(lk.values())

    # ── Ocupación de puertos por switch ──
    sw_labels, sw_ocup, sw_libre = [], [], []
    for s in Switch.objects.all():
        total_p = s.puertos.count() or (s.cantidad_puertos or 0)
        ocup = s.puertos.filter(dispositivo_conectado__isnull=False).count()
        sw_labels.append(s.nombre[:22])
        sw_ocup.append(ocup)
        sw_libre.append(max(0, total_p - ocup))

    graf = {
        'disp': {'labels': disp_labels, 'data': disp_data},
        'edif': {'labels': edif_labels, 'data': edif_data},
        'vlan': {'labels': vlan_labels, 'usadas': vlan_usadas, 'libres': vlan_libres},
        'cap':  {'labels': cap_labels, 'data': cap_data},
        'enl':  {'labels': enl_labels, 'data': enl_data},
        'sw':   {'labels': sw_labels, 'ocup': sw_ocup, 'libre': sw_libre},
    }
    ctx = {
        'nav_home_url': '/',
        'k_vlans': VLAN.objects.count(),
        'k_dispositivos': Device.objects.count(),
        'k_switches': Switch.objects.count(),
        'k_enlaces': NetworkLink.objects.count(),
        'k_planos': Plano.objects.count(),
        'k_ips_total': total_ips,
        'k_ips_uso': en_uso,
        'k_ips_libres': ip_estados.get('libre', 0),
        'k_ips_reservadas': ip_estados.get('reservada', 0),
        'k_pct_uso': pct_uso,
        'graf_json': json.dumps(graf),
    }
    return render(request, 'red/panel.html', ctx)


# ═══════════════════════ VLAN ═══════════════════════
@login_required
def vlan_list(request):
    _gate(request)
    filas = [(v, services.capacidad_vlan(v)) for v in VLAN.objects.all()]
    return render(request, 'red/vlan_list.html', {'filas': filas, 'nav_home_url': '/'})


@login_required
def vlan_detail(request, pk):
    _gate(request)
    v = get_object_or_404(VLAN, pk=pk)
    cap = services.recalcular_capacidad(v)
    solapes = services.subredes_solapadas(v.subred, exclude_pk=v.pk)
    return render(request, 'red/vlan_detail.html', {
        'v': v, 'cap': cap, 'solapes': solapes, 'nav_home_url': '/',
        'chart_labels': json.dumps(['Usadas', 'Libres']),
        'chart_data': json.dumps([cap['usadas'], cap['libres']]),
    })


@login_required
def vlan_form(request, pk=None):
    _gate(request, escribir=True)
    v = get_object_or_404(VLAN, pk=pk) if pk else None
    antes = _snapshot_rangos(v) if v else None
    if request.method == 'POST':
        form = VLANForm(request.POST, instance=v)
        if form.is_valid():
            obj = form.save(commit=False)
            if not obj.pk:
                obj.creado_por = request.user
            obj.modificado_por = request.user
            obj.save()
            form.save_m2m()
            services.generar_ips_vlan(obj)          # crea/actualiza IPs del rango
            services.recalcular_capacidad(obj)
            # Punto 12: auditar cambios de rango/límite
            despues = _snapshot_rangos(obj)
            if antes != despues:
                _audit(request, 'VLAN', obj, 'editar_rango' if v else 'crear',
                       antes, despues, request.POST.get('motivo', ''))
            messages.success(request, f'VLAN {obj.vlan_id} guardada.')
            return redirect('red:vlan_detail', pk=obj.pk)
    else:
        form = VLANForm(instance=v)
    solapes = services.subredes_solapadas(v.subred, exclude_pk=v.pk) if v else []
    return render(request, 'red/vlan_form.html', {'form': form, 'v': v, 'solapes': solapes, 'nav_home_url': '/'})


# ═══════════════════════ IPAM ═══════════════════════
_COLOR_ESTADO = {
    'libre': ('#2fb344', 'Libre'), 'asignada': ('#206bc4', 'Asignada'),
    'reservada': ('#f59f00', 'Reservada'), 'dhcp': ('#ae3ec9', 'DHCP'),
    'conflicto': ('#d63939', 'Conflicto'), 'pendiente': ('#f76707', 'Pendiente'),
    'no_utilizable': ('#adb5bd', 'No utilizable'),
}


@login_required
def ipam_grid(request, pk):
    _gate(request)
    v = get_object_or_404(VLAN, pk=pk)
    q = (request.GET.get('q') or '').strip()
    ips = IPAddress.objects.filter(vlan=v).select_related('dispositivo').order_by('direccion_int')
    if q:
        ips = ips.filter(Q(direccion__icontains=q) | Q(mac__icontains=q) | Q(hostname__icontains=q))
    celdas = [{'obj': ip, 'color': _COLOR_ESTADO.get(ip.estado, ('#ccc', ip.estado))[0]} for ip in ips]
    cap = services.recalcular_capacidad(v)
    return render(request, 'red/ipam_grid.html', {
        'v': v, 'celdas': celdas, 'cap': cap, 'q': q,
        'leyenda': _COLOR_ESTADO, 'nav_home_url': '/'})


@login_required
def ip_form(request, vlan_pk, pk=None):
    _gate(request, escribir=True)
    v = get_object_or_404(VLAN, pk=vlan_pk)
    ip = get_object_or_404(IPAddress, pk=pk, vlan=v) if pk else None
    if request.method == 'POST':
        form = IPForm(request.POST, instance=ip, vlan=v)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.vlan = v
            if not obj.pk:
                obj.creado_por = request.user
            obj.modificado_por = request.user
            obj.save()
            services.recalcular_capacidad(v)
            _audit(request, 'IP', obj, 'editar' if ip else 'crear')
            messages.success(request, f'IP {obj.direccion} guardada.')
            return redirect('red:ipam_grid', pk=v.pk)
    else:
        form = IPForm(instance=ip, vlan=v)
    return render(request, 'red/ip_form.html', {'form': form, 'v': v, 'ip': ip, 'nav_home_url': '/'})


@login_required
def ip_liberar(request, pk):
    _gate(request, escribir=True)
    ip = get_object_or_404(IPAddress, pk=pk)
    if request.method == 'POST':
        services.liberar_ip(ip, usuario=request.user)
        _audit(request, 'IP', ip, 'liberar', motivo=request.POST.get('motivo', ''))
        messages.success(request, f'IP {ip.direccion} liberada (histórico conservado).')
    return redirect('red:ipam_grid', pk=ip.vlan_id)


# ═══════════════════════ Ubicaciones (Campus/Edificio/Ubicación/Gabinete) ═══════════════════════
def _form_generico(request, FormClass, instance, titulo, redirect_name, modulo=''):
    if request.method == 'POST':
        form = FormClass(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            obj = form.save(commit=False)
            if not obj.pk and hasattr(obj, 'creado_por'):
                obj.creado_por = request.user
            if hasattr(obj, 'modificado_por'):
                obj.modificado_por = request.user
            obj.save()
            form.save_m2m()
            if modulo:
                _audit(request, modulo, obj, 'editar' if instance else 'crear')
            messages.success(request, 'Guardado correctamente.')
            return redirect(redirect_name)
    else:
        form = FormClass(instance=instance)
    return render(request, 'red/generico_form.html', {'form': form, 'titulo': titulo, 'nav_home_url': '/'})


def _eliminar_generico(request, obj, redirect_name, etiqueta):
    """Elimina solo si no tiene dependencias (regla del spec)."""
    try:
        obj.delete()
        messages.success(request, f'{etiqueta} eliminado.')
    except ProtectedError:
        messages.error(request, f'No se puede eliminar: {etiqueta} tiene registros dependientes.')
    return redirect(redirect_name)


@login_required
def ubicaciones(request):
    _gate(request)
    campus = Campus.objects.prefetch_related('edificios__ubicaciones__gabinetes').all()
    return render(request, 'red/ubicaciones.html', {'campus': campus, 'nav_home_url': '/'})


@login_required
def campus_form(request, pk=None):
    _gate(request, escribir=True)
    obj = get_object_or_404(Campus, pk=pk) if pk else None
    return _form_generico(request, CampusForm, obj, 'Campus', 'red:ubicaciones', 'Campus')


@login_required
def edificio_form(request, pk=None):
    _gate(request, escribir=True)
    obj = get_object_or_404(Edificio, pk=pk) if pk else None
    return _form_generico(request, EdificioForm, obj, 'Edificio', 'red:ubicaciones', 'Edificio')


@login_required
def ubicacion_form(request, pk=None):
    _gate(request, escribir=True)
    obj = get_object_or_404(Ubicacion, pk=pk) if pk else None
    return _form_generico(request, UbicacionForm, obj, 'Ubicación', 'red:ubicaciones', 'Ubicacion')


@login_required
def gabinete_form(request, pk=None):
    _gate(request, escribir=True)
    obj = get_object_or_404(Gabinete, pk=pk) if pk else None
    return _form_generico(request, GabineteForm, obj, 'Gabinete', 'red:ubicaciones', 'Gabinete')


@login_required
def ubicacion_borrar(request, tipo, pk):
    _gate(request, escribir=True)
    modelos = {'campus': (Campus, 'Campus'), 'edificio': (Edificio, 'Edificio'),
               'ubicacion': (Ubicacion, 'Ubicación'), 'gabinete': (Gabinete, 'Gabinete')}
    if tipo not in modelos or request.method != 'POST':
        return redirect('red:ubicaciones')
    Model, etq = modelos[tipo]
    obj = get_object_or_404(Model, pk=pk)
    return _eliminar_generico(request, obj, 'red:ubicaciones', etq)


# ═══════════════════════ Dispositivos + Interfaces ═══════════════════════
@login_required
def device_list(request):
    _gate(request)
    q = (request.GET.get('q') or '').strip()
    devs = Device.objects.select_related('ubicacion', 'vlan').all()
    if q:
        devs = devs.filter(Q(nombre__icontains=q) | Q(hostname__icontains=q) | Q(ip__icontains=q) |
                           Q(mac_principal__icontains=q) | Q(numero_serie__icontains=q))
    return render(request, 'red/device_list.html', {'devs': devs, 'q': q, 'nav_home_url': '/'})


@login_required
def device_form(request, pk=None):
    _gate(request, escribir=True)
    obj = get_object_or_404(Device, pk=pk) if pk else None
    if request.method == 'POST':
        form = DeviceForm(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            d = form.save(commit=False)
            if not d.pk:
                d.creado_por = request.user
            d.modificado_por = request.user
            d.save()
            _audit(request, 'Dispositivo', d, 'editar' if obj else 'crear')
            messages.success(request, 'Dispositivo guardado.')
            return redirect('red:device_detail', pk=d.pk)
    else:
        form = DeviceForm(instance=obj)
    return render(request, 'red/device_form.html', {'form': form, 'obj': obj, 'nav_home_url': '/'})


@login_required
def device_detail(request, pk):
    _gate(request)
    d = get_object_or_404(Device.objects.select_related('ubicacion', 'gabinete', 'vlan'), pk=pk)
    return render(request, 'red/device_detail.html', {
        'd': d, 'interfaces': d.interfaces.select_related('vlan').all(), 'nav_home_url': '/'})


@login_required
def device_borrar(request, pk):
    _gate(request, escribir=True)
    d = get_object_or_404(Device, pk=pk)
    if request.method == 'POST':
        return _eliminar_generico(request, d, 'red:device_list', 'Dispositivo')
    return redirect('red:device_detail', pk=pk)


@login_required
def interface_form(request, device_pk, pk=None):
    _gate(request, escribir=True)
    d = get_object_or_404(Device, pk=device_pk)
    obj = get_object_or_404(NetworkInterface, pk=pk, dispositivo=d) if pk else None
    if request.method == 'POST':
        form = InterfaceForm(request.POST, instance=obj)
        if form.is_valid():
            i = form.save(commit=False)
            i.dispositivo = d
            if not i.pk:
                i.creado_por = request.user
            i.modificado_por = request.user
            i.save()
            messages.success(request, 'Interfaz guardada.')
            return redirect('red:device_detail', pk=d.pk)
    else:
        form = InterfaceForm(instance=obj)
    return render(request, 'red/generico_form.html',
                  {'form': form, 'titulo': f'Interfaz · {d.nombre}',
                   'volver_url': reverse('red:device_detail', args=[d.pk]), 'nav_home_url': '/'})


@login_required
def interface_borrar(request, pk):
    _gate(request, escribir=True)
    i = get_object_or_404(NetworkInterface, pk=pk)
    dev = i.dispositivo_id
    if request.method == 'POST':
        i.delete()
        messages.success(request, 'Interfaz eliminada.')
    return redirect('red:device_detail', pk=dev)


# ═══════════════════════ Auditoría ═══════════════════════
@login_required
def auditoria(request):
    _gate(request)
    logs = RedAuditLog.objects.select_related('usuario').all()[:300]
    return render(request, 'red/auditoria.html', {'logs': logs, 'nav_home_url': '/'})


# ═══════════════════════ FASE 2 · Switches / Puertos / Enlaces ═══════════════════════
from .models import Switch, SwitchPort, NetworkLink  # noqa: E402
from .forms import SwitchForm, SwitchPortForm, NetworkLinkForm  # noqa: E402


@login_required
def switch_list(request):
    _gate(request)
    switches = Switch.objects.select_related('ubicacion').all()
    return render(request, 'red/switch_list.html', {'switches': switches, 'nav_home_url': '/'})


@login_required
def switch_detail(request, pk):
    _gate(request)
    sw = get_object_or_404(Switch.objects.select_related('ubicacion', 'gabinete'), pk=pk)
    puertos = sw.puertos.select_related('vlan_access', 'vlan_nativa', 'dispositivo_conectado').all()
    mapa = [{'p': p, 'color': services.color_puerto(p)} for p in puertos]
    return render(request, 'red/switch_detail.html', {
        'sw': sw, 'mapa': mapa, 'alertas': services.alertas_switch(sw), 'nav_home_url': '/'})


@login_required
def switch_form(request, pk=None):
    _gate(request, escribir=True)
    sw = get_object_or_404(Switch, pk=pk) if pk else None
    if request.method == 'POST':
        form = SwitchForm(request.POST, instance=sw)
        if form.is_valid():
            o = form.save(commit=False)
            if not o.pk:
                o.creado_por = request.user
            o.modificado_por = request.user
            o.save()
            _audit(request, 'Switch', o, 'editar' if sw else 'crear')
            messages.success(request, 'Switch guardado.')
            return redirect('red:switch_detail', pk=o.pk)
    else:
        form = SwitchForm(instance=sw)
    return render(request, 'red/switch_form.html', {'form': form, 'sw': sw, 'nav_home_url': '/'})


@login_required
def switch_borrar(request, pk):
    _gate(request, escribir=True)
    sw = get_object_or_404(Switch, pk=pk)
    if request.method == 'POST':
        return _eliminar_generico(request, sw, 'red:switch_list', 'Switch')
    return redirect('red:switch_detail', pk=pk)


@login_required
def port_form(request, switch_pk, pk=None):
    _gate(request, escribir=True)
    sw = get_object_or_404(Switch, pk=switch_pk)
    p = get_object_or_404(SwitchPort, pk=pk, switch=sw) if pk else None
    if request.method == 'POST':
        form = SwitchPortForm(request.POST, instance=p)
        if form.is_valid():
            o = form.save(commit=False)
            o.switch = sw
            if not o.pk:
                o.creado_por = request.user
            o.modificado_por = request.user
            o.save()
            _audit(request, 'Puerto', o, 'editar' if p else 'crear')
            messages.success(request, f'Puerto {o.nombre or o.numero} guardado.')
            return redirect('red:switch_detail', pk=sw.pk)
    else:
        form = SwitchPortForm(instance=p)
    return render(request, 'red/generico_form.html',
                  {'form': form, 'titulo': f'Puerto · {sw.nombre}',
                   'volver_url': reverse('red:switch_detail', args=[sw.pk]), 'nav_home_url': '/'})


@login_required
def link_list(request):
    _gate(request)
    enlaces = NetworkLink.objects.select_related('dispositivo_origen', 'dispositivo_destino').all()
    en_ciclo = services.detectar_ciclos()
    return render(request, 'red/link_list.html', {'enlaces': enlaces, 'en_ciclo': en_ciclo, 'nav_home_url': '/'})


@login_required
def link_form(request, pk=None):
    _gate(request, escribir=True)
    l = get_object_or_404(NetworkLink, pk=pk) if pk else None
    avisos = []
    if request.method == 'POST':
        form = NetworkLinkForm(request.POST, instance=l)
        if form.is_valid():
            o = form.save(commit=False)
            if not o.pk:
                o.creado_por = request.user
            o.modificado_por = request.user
            o.save()
            _, avisos = services.validar_link(o)
            for a in avisos:
                messages.warning(request, a)
            _audit(request, 'Enlace', o, 'editar' if l else 'crear')
            messages.success(request, 'Enlace guardado.')
            return redirect('red:link_list')
    else:
        form = NetworkLinkForm(instance=l)
    return render(request, 'red/link_form.html', {'form': form, 'l': l, 'nav_home_url': '/'})


@login_required
def link_borrar(request, pk):
    _gate(request, escribir=True)
    l = get_object_or_404(NetworkLink, pk=pk)
    if request.method == 'POST':
        l.delete()
        messages.success(request, 'Enlace eliminado.')
    return redirect('red:link_list')


# ═══════════════════════ FASE 3 · Mapa de campus ═══════════════════════
from django.http import JsonResponse  # noqa: E402
from django.views.decorators.http import require_POST  # noqa: E402
from .models import Plano, Marcador, LineaPlano, RackItem, Gabinete as _Gab  # noqa: E402
from .forms import PlanoForm  # noqa: E402


@login_required
def planos_list(request):
    _gate(request)
    planos = Plano.objects.select_related('edificio').all()
    return render(request, 'red/planos_list.html', {'planos': planos, 'nav_home_url': '/'})


@login_required
def plano_form(request, pk=None):
    _gate(request, escribir=True)
    p = get_object_or_404(Plano, pk=pk) if pk else None
    if request.method == 'POST':
        form = PlanoForm(request.POST, request.FILES, instance=p)
        if form.is_valid():
            o = form.save(commit=False)
            if not o.pk:
                o.creado_por = request.user
            o.modificado_por = request.user
            o.save()
            _audit(request, 'Plano', o, 'editar' if p else 'crear')
            messages.success(request, 'Plano guardado.')
            return redirect('red:plano_view', pk=o.pk)
    else:
        form = PlanoForm(instance=p)
    return render(request, 'red/generico_form.html',
                  {'form': form, 'titulo': 'Plano', 'volver_url': reverse('red:planos_list'), 'nav_home_url': '/'})


@login_required
def plano_view(request, pk):
    _gate(request)
    p = get_object_or_404(Plano, pk=pk)
    marcadores = list(p.marcadores.all().values(
        'id', 'etiqueta', 'tipo', 'x', 'y', 'color', 'forma', 'tamano', 'rotacion',
        'gabinete_id', 'dispositivo_id', 'switch_id', 'observaciones'))
    lineas = list(p.lineas.all().values('id', 'origen_id', 'destino_id', 'color', 'etiqueta', 'estilo', 'mx', 'my',
                                        'equipo_origen_id', 'equipo_destino_id'))
    # <--- hecho por claude code: ítems de rack por gabinete (para que una línea apunte a un equipo)
    racks = {}
    for it in RackItem.objects.all().values('id', 'nombre', 'gabinete_id'):
        racks.setdefault(it['gabinete_id'], []).append({'id': it['id'], 'nombre': it['nombre']})
    import json as _json
    ctx = {
        'p': p,
        'marcadores_json': _json.dumps(marcadores),
        'lineas_json': _json.dumps(lineas),
        'racks_json': _json.dumps(racks),
        'gabinetes': _Gab.objects.select_related('ubicacion').all(),
        'dispositivos': Device.objects.all(),
        'switches': Switch.objects.all(),
        'nav_home_url': '/',
    }
    return render(request, 'red/plano_view.html', ctx)


def _marcador_dict(m):
    ref = m.gabinete_id or m.dispositivo_id or m.switch_id or ''
    return {'id': m.id, 'etiqueta': m.etiqueta, 'tipo': m.tipo, 'x': m.x, 'y': m.y, 'color': m.color,
            'forma': m.forma, 'tamano': m.tamano, 'rotacion': m.rotacion, 'ref_pk': ref,
            'gabinete_id': m.gabinete_id, 'dispositivo_id': m.dispositivo_id, 'switch_id': m.switch_id}


@login_required
@require_POST
def marcador_add(request, plano_pk):
    _gate(request, escribir=True)
    p = get_object_or_404(Plano, pk=plano_pk)
    d = request.POST
    m = Marcador(plano=p, etiqueta=(d.get('etiqueta') or 'Marcador'), tipo=(d.get('tipo') or 'nota'),
                 x=float(d.get('x') or 50), y=float(d.get('y') or 50),
                 color=(d.get('color') or '#206bc4'), forma=(d.get('forma') or 'pin'),
                 tamano=int(d.get('tamano') or 30), rotacion=int(d.get('rotacion') or 0),
                 observaciones=d.get('observaciones', ''),
                 creado_por=request.user, modificado_por=request.user)
    ref = d.get('ref_pk')
    if ref:
        if m.tipo == 'gabinete':
            m.gabinete_id = ref
        elif m.tipo == 'dispositivo':
            m.dispositivo_id = ref
        elif m.tipo == 'switch':
            m.switch_id = ref
    m.save()
    _audit(request, 'Marcador', m, 'crear')
    return JsonResponse({'ok': True, 'm': _marcador_dict(m)})


@login_required
@require_POST
def marcador_move(request, pk):
    _gate(request, escribir=True)
    m = get_object_or_404(Marcador, pk=pk)
    m.x = float(request.POST.get('x', m.x))
    m.y = float(request.POST.get('y', m.y))
    m.modificado_por = request.user
    m.save(update_fields=['x', 'y', 'modificado_por', 'fecha_modificado'])
    return JsonResponse({'ok': True})


@login_required
@require_POST
def marcador_delete(request, pk):
    _gate(request, escribir=True)
    m = get_object_or_404(Marcador, pk=pk)
    m.delete()
    return JsonResponse({'ok': True})


@login_required
@require_POST
def marcador_edit(request, pk):
    """Edita etiqueta/tipo/color/vínculo de un marcador existente."""
    _gate(request, escribir=True)
    m = get_object_or_404(Marcador, pk=pk)
    d = request.POST
    if d.get('etiqueta'):
        m.etiqueta = d.get('etiqueta')
    m.tipo = d.get('tipo') or m.tipo
    m.color = d.get('color') or m.color
    m.forma = d.get('forma') or m.forma
    if d.get('tamano'):
        m.tamano = int(d.get('tamano'))
    if d.get('rotacion') is not None and d.get('rotacion') != '':
        m.rotacion = int(d.get('rotacion'))
    m.observaciones = d.get('observaciones', m.observaciones)
    # re-vincular según el tipo
    m.gabinete_id = m.dispositivo_id = m.switch_id = None
    ref = d.get('ref_pk')
    if ref:
        if m.tipo == 'gabinete':
            m.gabinete_id = ref
        elif m.tipo == 'dispositivo':
            m.dispositivo_id = ref
        elif m.tipo == 'switch':
            m.switch_id = ref
    m.modificado_por = request.user
    m.save()
    _audit(request, 'Marcador', m, 'editar')
    return JsonResponse({'ok': True, 'm': _marcador_dict(m)})


@login_required
def marcador_link(request, pk):
    """Devuelve la URL a abrir cuando se hace clic en un marcador enlazado."""
    _gate(request)
    m = get_object_or_404(Marcador, pk=pk)
    url = None
    if m.gabinete_id:
        url = reverse('red:rack_view', args=[m.gabinete_id])
    elif m.dispositivo_id:
        url = reverse('red:device_detail', args=[m.dispositivo_id])
    elif m.switch_id:
        url = reverse('red:switch_detail', args=[m.switch_id])
    return JsonResponse({'ok': True, 'url': url})


# <--- hecho por claude code: líneas/cables entre marcadores sobre el plano
# <--- hecho por claude code: edición masiva de tamaño/giro sobre varios marcadores
@login_required
@require_POST
def marcadores_bulk(request):
    _gate(request, escribir=True)
    d = request.POST
    ids = [i for i in d.get('ids', '').split(',') if i.strip().isdigit()]
    if not ids:
        return JsonResponse({'ok': False, 'error': 'Sin selección.'}, status=400)
    updates = {}
    if d.get('tamano') not in (None, ''):
        updates['tamano'] = max(8, min(120, int(d.get('tamano'))))
    if d.get('rotacion') not in (None, ''):
        updates['rotacion'] = int(d.get('rotacion')) % 360
    if not updates:
        return JsonResponse({'ok': False, 'error': 'Nada que aplicar.'}, status=400)
    updates['modificado_por'] = request.user
    n = Marcador.objects.filter(pk__in=ids).update(**updates)
    return JsonResponse({'ok': True, 'n': n, 'aplicado': {k: v for k, v in updates.items() if k != 'modificado_por'}})


@login_required
@require_POST
def linea_add(request, plano_pk):
    _gate(request, escribir=True)
    p = get_object_or_404(Plano, pk=plano_pk)
    d = request.POST
    o = get_object_or_404(Marcador, pk=d.get('origen'), plano=p)
    de = get_object_or_404(Marcador, pk=d.get('destino'), plano=p)
    if o.id == de.id:
        return JsonResponse({'ok': False, 'error': 'La línea no puede unir un marcador consigo mismo.'}, status=400)
    ln = LineaPlano.objects.create(
        plano=p, origen=o, destino=de, color=(d.get('color') or '#495057'),
        etiqueta=(d.get('etiqueta') or ''), estilo=(d.get('estilo') or 'recta'),
        creado_por=request.user, modificado_por=request.user)
    _audit(request, 'LineaPlano', ln, 'crear')
    return JsonResponse({'ok': True, 'l': _linea_dict(ln)})


def _linea_dict(ln):
    return {'id': ln.id, 'origen_id': ln.origen_id, 'destino_id': ln.destino_id,
            'color': ln.color, 'etiqueta': ln.etiqueta, 'estilo': ln.estilo, 'mx': ln.mx, 'my': ln.my,
            'equipo_origen': ln.equipo_origen_id or '', 'equipo_destino': ln.equipo_destino_id or ''}


@login_required
@require_POST
def linea_edit(request, pk):
    _gate(request, escribir=True)
    ln = get_object_or_404(LineaPlano, pk=pk)
    d = request.POST
    if 'etiqueta' in d:
        ln.etiqueta = d.get('etiqueta') or ''
    if d.get('color'):
        ln.color = d.get('color')
    if d.get('estilo') in ('recta', 'orto'):
        ln.estilo = d.get('estilo')
    if d.get('mx') not in (None, ''):
        ln.mx = float(d.get('mx'))
    if d.get('my') not in (None, ''):
        ln.my = float(d.get('my'))
    if 'equipo_origen' in d:
        v = d.get('equipo_origen')
        ln.equipo_origen_id = int(v) if v and v.isdigit() and RackItem.objects.filter(pk=v).exists() else None
    if 'equipo_destino' in d:
        v = d.get('equipo_destino')
        ln.equipo_destino_id = int(v) if v and v.isdigit() and RackItem.objects.filter(pk=v).exists() else None
    ln.modificado_por = request.user
    ln.save()
    _audit(request, 'LineaPlano', ln, 'editar')
    return JsonResponse({'ok': True, 'l': _linea_dict(ln)})


@login_required
@require_POST
def linea_delete(request, pk):
    _gate(request, escribir=True)
    ln = get_object_or_404(LineaPlano, pk=pk)
    ln.delete()
    return JsonResponse({'ok': True})


# ═══════════════════════ Rack / Gabinete (elevación tipo draw.io) ═══════════════════════
def _rackitem_dict(it):
    ref = ('s:%s' % it.switch_id) if it.switch_id else (('d:%s' % it.device_id) if it.device_id else '')
    return {'id': it.id, 'nombre': it.nombre, 'tipo': it.tipo, 'cara': it.cara, 'u_pos': it.u_pos, 'u_alto': it.u_alto,
            'color': it.color, 'ref': ref, 'device_id': it.device_id, 'switch_id': it.switch_id,
            'observaciones': it.observaciones}


@login_required
def racks_list(request):
    """<--- hecho por claude code: lista de gabinetes con acceso directo a su rack."""
    _gate(request)
    gabs = _Gab.objects.select_related('ubicacion__edificio').annotate(n_items=Count('rack_items')).order_by('nombre')
    return render(request, 'red/racks_list.html', {'gabinetes': gabs, 'nav_home_url': '/'})


@login_required
def rack_view(request, gabinete_pk):
    _gate(request)
    g = get_object_or_404(_Gab, pk=gabinete_pk)
    items = list(g.rack_items.all().values(
        'id', 'nombre', 'tipo', 'cara', 'u_pos', 'u_alto', 'color', 'device_id', 'switch_id', 'observaciones'))
    unidades = g.unidades_rack or 42
    import json as _json
    ctx = {
        'g': g, 'unidades': unidades,
        'items_json': _json.dumps(items),
        'dispositivos': Device.objects.all(),
        'switches': Switch.objects.all(),
        'nav_home_url': '/',
    }
    return render(request, 'red/rack_view.html', ctx)


def _rack_set_links(it, d):
    it.device_id = it.switch_id = None
    ref, kind = d.get('ref_pk'), d.get('ref_kind')
    if ref and kind == 'switch':
        it.switch_id = ref
    elif ref and kind == 'device':
        it.device_id = ref


@login_required
@require_POST
def rack_item_add(request, gabinete_pk):
    _gate(request, escribir=True)
    g = get_object_or_404(_Gab, pk=gabinete_pk)
    d = request.POST
    it = RackItem(gabinete=g, nombre=(d.get('nombre') or 'Equipo'), tipo=(d.get('tipo') or 'switch'),
                  cara=(d.get('cara') if d.get('cara') in ('frente', 'atras') else 'frente'),
                  u_pos=int(d.get('u_pos') or 1), u_alto=max(1, int(d.get('u_alto') or 1)),
                  color=(d.get('color') or '#495057'), observaciones=d.get('observaciones', ''),
                  creado_por=request.user, modificado_por=request.user)
    _rack_set_links(it, d)
    it.save()
    _audit(request, 'RackItem', it, 'crear')
    return JsonResponse({'ok': True, 'it': _rackitem_dict(it)})


@login_required
@require_POST
def rack_item_edit(request, pk):
    _gate(request, escribir=True)
    it = get_object_or_404(RackItem, pk=pk)
    d = request.POST
    if d.get('nombre'):
        it.nombre = d.get('nombre')
    it.tipo = d.get('tipo') or it.tipo
    if d.get('cara') in ('frente', 'atras'):
        it.cara = d.get('cara')
    if d.get('u_pos'):
        it.u_pos = int(d.get('u_pos'))
    if d.get('u_alto'):
        it.u_alto = max(1, int(d.get('u_alto')))
    if d.get('color'):
        it.color = d.get('color')
    it.observaciones = d.get('observaciones', it.observaciones)
    _rack_set_links(it, d)
    it.modificado_por = request.user
    it.save()
    _audit(request, 'RackItem', it, 'editar')
    return JsonResponse({'ok': True, 'it': _rackitem_dict(it)})


@login_required
@require_POST
def rack_item_move(request, pk):
    _gate(request, escribir=True)
    it = get_object_or_404(RackItem, pk=pk)
    it.u_pos = max(1, int(request.POST.get('u_pos', it.u_pos)))
    it.modificado_por = request.user
    it.save(update_fields=['u_pos', 'modificado_por', 'fecha_modificado'])
    return JsonResponse({'ok': True})


@login_required
@require_POST
def rack_item_delete(request, pk):
    _gate(request, escribir=True)
    it = get_object_or_404(RackItem, pk=pk)
    it.delete()
    return JsonResponse({'ok': True})


# ═══════════════════════ FASE 3 · Editor de topología (Cytoscape) ═══════════════════════
_TIPO_COLOR = {
    'servidor': '#d63939', 'router': '#4263eb', 'switch_adm': '#206bc4', 'switch_noadm': '#748ffc',
    'access_point': '#0ca678', 'mikrotik': '#f76707', 'camara': '#ae3ec9', 'nvr': '#862e9c',
    'computadora': '#2fb344', 'laptop': '#37b24d', 'telefono_ip': '#f59f00', 'impresora': '#e8590c',
    'televisor': '#1098ad', 'reloj': '#66a80f', 'ups': '#495057', 'pbx': '#e64980', 'otro': '#adb5bd',
}


_FORMA_COLOR = {
    'router': '#4263eb', 'switch': '#206bc4', 'servidor': '#d63939', 'ap': '#0ca678',
    'firewall': '#e8590c', 'camara': '#ae3ec9', 'pc': '#2fb344', 'impresora': '#e64980',
    'gabinete': '#495057', 'nube': '#1098ad', 'telefono': '#f59f00', 'reloj': '#66a80f',
    'acceso': '#862e9c', 'pin': '#206bc4',
}


@login_required
def topologia(request):
    _gate(request)
    import json as _json

    # <--- hecho por claude code: modo "plano" → dibuja los marcadores y líneas de un plano
    plano_pk = request.GET.get('plano') or ''
    if plano_pk:
        pl = get_object_or_404(Plano, pk=plano_pk)
        marks = list(pl.marcadores.all())
        nodes = []
        for m in marks:
            nodes.append({'data': {
                'id': f'm{m.id}', 'label': m.etiqueta, 'ip': '', 'tipo': m.get_forma_display(),
                'modelo': '', 'ubic': str(pl.nombre), 'color': (m.color or _FORMA_COLOR.get(m.forma, '#206bc4')),
                'ciclo': False},
                'position': {'x': round(m.x * 12, 1), 'y': round(m.y * 8, 1)}})
        edges = []
        for ln in pl.lineas.select_related('origen', 'destino').all():
            edges.append({'data': {
                'id': f'lp{ln.id}', 'source': f'm{ln.origen_id}', 'target': f'm{ln.destino_id}',
                'tipo': 'cobre', 'modo': 'acceso', 'label': (ln.etiqueta or ''),
                'nativa': '', 'permitidas': ''}})
        con = set()
        for ln in pl.lineas.all():
            con.add(ln.origen_id); con.add(ln.destino_id)
        sueltos = [m.etiqueta for m in marks if m.id not in con]
        mm = ['graph LR'] + [f'  m{m.id}["{m.etiqueta}"]' for m in marks]
        for ln in pl.lineas.all():
            mm.append(f'  m{ln.origen_id} --- m{ln.destino_id}')
        return render(request, 'red/topologia.html', {
            'nodes_json': _json.dumps(nodes), 'edges_json': _json.dumps(edges),
            'sueltos': sueltos, 'hay_ciclo': False, 'mermaid': '\n'.join(mm),
            'edificios': Edificio.objects.all(), 'vlans': VLAN.objects.all(),
            'tipos': Device.TIPOS, 'f_edif': '', 'f_vlan': '', 'f_tipo': '', 'colorby': 'tipo',
            'planos': Plano.objects.all(), 'plano_sel': pl,
            'nav_home_url': '/'})

    f_edif = request.GET.get('edificio') or ''
    f_vlan = request.GET.get('vlan') or ''
    f_tipo = request.GET.get('tipo') or ''
    colorby = request.GET.get('colorby') or 'tipo'

    devs = Device.objects.select_related('ubicacion__edificio', 'vlan').all()
    if f_edif:
        devs = devs.filter(ubicacion__edificio_id=f_edif)
    if f_vlan:
        devs = devs.filter(vlan_id=f_vlan)
    if f_tipo:
        devs = devs.filter(tipo=f_tipo)
    devs = list(devs)
    ids = {d.id for d in devs}

    en_ciclo = services.detectar_ciclos()
    nodes = []
    for d in devs:
        color = (d.vlan.color if (colorby == 'vlan' and d.vlan) else _TIPO_COLOR.get(d.tipo, '#adb5bd'))
        nodes.append({'data': {
            'id': f'd{d.id}', 'label': d.nombre, 'ip': d.ip or '', 'tipo': d.get_tipo_display(),
            'modelo': d.modelo or '', 'ubic': str(d.ubicacion or ''), 'color': color,
            'ciclo': d.id in en_ciclo},
            'position': ({'x': d.topo_x, 'y': d.topo_y} if d.topo_x is not None and d.topo_y is not None else None)})

    edges = []
    con = set()
    for l in NetworkLink.objects.select_related('vlan_nativa').all():
        if l.dispositivo_origen_id in ids and l.dispositivo_destino_id in ids:
            con.add(l.dispositivo_origen_id); con.add(l.dispositivo_destino_id)
            edges.append({'data': {
                'id': f'l{l.id}', 'source': f'd{l.dispositivo_origen_id}', 'target': f'd{l.dispositivo_destino_id}',
                'tipo': l.tipo, 'modo': l.modo, 'label': l.get_tipo_display(),
                'nativa': (l.vlan_nativa.vlan_id if l.vlan_nativa else ''), 'permitidas': l.vlans_permitidas}})
    sueltos = [d.nombre for d in devs if d.id not in con]

    # Mermaid (solo lectura)
    mm = ['graph LR']
    for d in devs:
        mm.append(f'  d{d.id}["{d.nombre}"]')
    for l in NetworkLink.objects.all():
        if l.dispositivo_origen_id in ids and l.dispositivo_destino_id in ids:
            mm.append(f'  d{l.dispositivo_origen_id} --- d{l.dispositivo_destino_id}')

    import json as _json
    return render(request, 'red/topologia.html', {
        'nodes_json': _json.dumps(nodes), 'edges_json': _json.dumps(edges),
        'sueltos': sueltos, 'hay_ciclo': bool(en_ciclo), 'mermaid': '\n'.join(mm),
        'edificios': Edificio.objects.all(), 'vlans': VLAN.objects.all(),
        'tipos': Device.TIPOS, 'f_edif': f_edif, 'f_vlan': f_vlan, 'f_tipo': f_tipo, 'colorby': colorby,
        'planos': Plano.objects.all(), 'plano_sel': None,
        'nav_home_url': '/'})


@login_required
@require_POST
def nodo_pos(request, pk):
    _gate(request, escribir=True)
    d = get_object_or_404(Device, pk=pk)
    d.topo_x = float(request.POST.get('x', 0))
    d.topo_y = float(request.POST.get('y', 0))
    Device.objects.filter(pk=d.pk).update(topo_x=d.topo_x, topo_y=d.topo_y)
    return JsonResponse({'ok': True})


# ═══════════════════════ FASE 4 · Pruebas de conectividad ═══════════════════════
from django.utils import timezone as _tz  # noqa: E402


@login_required
def pruebas(request):
    _gate(request)
    # <--- hecho por claude code: GenericIPAddressField convierte ''→NULL; filtrar solo por isnull
    devs = Device.objects.exclude(ip__isnull=True).order_by('nombre')
    sws = Switch.objects.exclude(ip_admin__isnull=True).order_by('nombre')
    return render(request, 'red/pruebas.html', {'dispositivos': devs, 'switches': sws, 'nav_home_url': '/'})


def _aplicar_ping(obj, ip, manager):
    ok, lat = services.ping_host(ip)
    estado = 'ok' if ok else 'caido'
    now = _tz.now()
    manager.filter(pk=obj.pk).update(conectividad=estado, latencia_ms=lat, probado_en=now)
    return {'ok': True, 'estado': estado, 'latencia': lat, 'probado': now.strftime('%d/%m %H:%M')}


@login_required
@require_POST
def device_ping(request, pk):
    _gate(request, escribir=True)
    d = get_object_or_404(Device, pk=pk)
    return JsonResponse(_aplicar_ping(d, d.ip, Device.objects))


@login_required
@require_POST
def switch_ping(request, pk):
    _gate(request, escribir=True)
    s = get_object_or_404(Switch, pk=pk)
    return JsonResponse(_aplicar_ping(s, s.ip_admin, Switch.objects))


# ═══════════════════════ FASE 4 · Migración/planificación de VLAN (preview) ═══════════════════════
@login_required
def migracion(request):
    """Análisis de impacto de mover dispositivos de una VLAN a otra (solo lectura).
    La EJECUCIÓN real vive en `migracion_ejecutar` (POST, permiso aprobar_migracion)."""
    _gate(request)
    origen_id = request.GET.get('origen') or ''
    destino_id = request.GET.get('destino') or ''
    contexto = {'vlans': VLAN.objects.all(), 'origen_id': origen_id, 'destino_id': destino_id, 'nav_home_url': '/',
                # <--- hecho por claude code: solo quien puede aprobar migraciones ve el botón de ejecutar
                'puede_ejecutar': request.user.is_superuser or request.user.has_perm('red.aprobar_migracion')}
    if origen_id and destino_id and origen_id != destino_id:
        origen = get_object_or_404(VLAN, pk=origen_id)
        destino = get_object_or_404(VLAN, pk=destino_id)
        devs = list(Device.objects.filter(vlan=origen).order_by('nombre'))
        cap = services.capacidad_vlan(destino)
        contexto.update({
            'origen': origen, 'destino': destino, 'devs': devs,
            'n_devs': len(devs), 'cap_destino': cap,
            'suficiente': cap['libres'] >= len(devs),
        })
    return render(request, 'red/migracion.html', contexto)


@login_required
@require_POST
def migracion_ejecutar(request):
    """<--- hecho por claude code: EJECUCIÓN real de la migración de VLAN.
    Requisitos: permiso `red.aprobar_migracion` (o superuser), motivo obligatorio y
    confirmación explícita. Se procesa EQUIPO POR EQUIPO en transacción: si uno falla se
    revierte solo ese y se continúa; nada queda a medias. Todo va a RedAuditLog (antes/después)."""
    _gate(request, escribir=True)
    u = request.user
    if not (u.is_superuser or u.has_perm('red.aprobar_migracion')):
        raise PermissionDenied('Se requiere el permiso "Aprobar migraciones".')
    origen = get_object_or_404(VLAN, pk=request.POST.get('origen'))
    destino = get_object_or_404(VLAN, pk=request.POST.get('destino'))
    motivo = (request.POST.get('motivo') or '').strip()
    if origen.pk == destino.pk:
        messages.error(request, 'Origen y destino deben ser VLAN distintas.')
        return redirect(f"{reverse('red:migracion')}?origen={origen.pk}&destino={destino.pk}")
    if len(motivo) < 5:
        messages.error(request, 'Indica un motivo (mínimo 5 caracteres) para dejar rastro en la auditoría.')
        return redirect(f"{reverse('red:migracion')}?origen={origen.pk}&destino={destino.pk}")
    if request.POST.get('confirmar') != 'SI':
        messages.error(request, 'Debes marcar la casilla de confirmación.')
        return redirect(f"{reverse('red:migracion')}?origen={origen.pk}&destino={destino.pk}")

    # Solo los equipos marcados (permite migrar un subconjunto); si no se marca ninguno, todos los de la VLAN
    ids = request.POST.getlist('dispositivos')
    devs = Device.objects.filter(vlan=origen)
    if ids:
        devs = devs.filter(pk__in=ids)
    resultado = services.migrar_dispositivos_vlan(list(devs.order_by('nombre')), destino, usuario=u, motivo=motivo)

    for r in resultado['ok']:
        _audit(request, 'Migracion VLAN', r['dispositivo'], 'migrar',
               antes={'vlan': origen.vlan_id, 'ip': r['ip_antes']},
               despues={'vlan': destino.vlan_id, 'ip': r['ip_despues']}, motivo=motivo)
    for r in resultado['error']:
        _audit(request, 'Migracion VLAN', r['dispositivo'], 'migrar_error',
               antes={'vlan': origen.vlan_id}, despues={'error': r['error']}, motivo=motivo)

    n_ok, n_err = len(resultado['ok']), len(resultado['error'])
    if n_ok:
        messages.success(request, f'Migración completada: {n_ok} equipo(s) movidos a VLAN {destino.vlan_id} con IP nueva.')
    if n_err:
        messages.warning(request, f'{n_err} equipo(s) NO se movieron (quedaron en VLAN {origen.vlan_id}): ' +
                         '; '.join(f"{r['dispositivo']}: {r['error']}" for r in resultado['error'][:5]))
    return redirect(f"{reverse('red:migracion')}?origen={origen.pk}&destino={destino.pk}")


# ═══════════════════════ FASE 5 · Buscador global ═══════════════════════
def _buscar(q):
    q = (q or '').strip()
    res = {'ips': [], 'dispositivos': [], 'switches': [], 'vlans': []}
    if len(q) < 2:
        return res
    for ip in IPAddress.objects.filter(Q(direccion__icontains=q) | Q(mac__icontains=q) | Q(hostname__icontains=q)).select_related('vlan')[:25]:
        res['ips'].append({'texto': ip.direccion, 'sub': f"{ip.get_estado_display()} · VLAN {ip.vlan.vlan_id if ip.vlan else '—'}",
                           'url': reverse('red:ipam_grid', args=[ip.vlan_id]) if ip.vlan_id else ''})
    for d in Device.objects.filter(Q(nombre__icontains=q) | Q(ip__icontains=q) | Q(mac_principal__icontains=q) | Q(hostname__icontains=q))[:25]:
        res['dispositivos'].append({'texto': d.nombre, 'sub': f"{d.get_tipo_display()} · {d.ip or 's/IP'}",
                                    'url': reverse('red:device_detail', args=[d.id])})
    for s in Switch.objects.filter(Q(nombre__icontains=q) | Q(ip_admin__icontains=q) | Q(mac_admin__icontains=q))[:25]:
        res['switches'].append({'texto': s.nombre, 'sub': f"{s.ip_admin or 's/IP'}",
                                'url': reverse('red:switch_detail', args=[s.id])})
    for v in VLAN.objects.filter(Q(nombre__icontains=q) | Q(vlan_id__icontains=q))[:25]:
        res['vlans'].append({'texto': f"VLAN {v.vlan_id} · {v.nombre}", 'sub': v.subred or '',
                             'url': reverse('red:vlan_detail', args=[v.id])})
    return res


@login_required
def buscar(request):
    _gate(request)
    q = request.GET.get('q', '')
    res = _buscar(q)
    total = sum(len(v) for v in res.values())
    res_grupos = [
        ('Direcciones IP', res['ips'], 'ti-network'),
        ('Dispositivos', res['dispositivos'], 'ti-device-desktop'),
        ('Switches', res['switches'], 'ti-switch-3'),
        ('VLANs', res['vlans'], 'ti-router'),
    ]
    return render(request, 'red/buscar.html', {'q': q, 'res': res, 'res_grupos': res_grupos, 'total': total, 'nav_home_url': '/'})


@login_required
def api_buscar(request):
    _gate(request)
    return JsonResponse(_buscar(request.GET.get('q', '')))


# ═══════════════════════ FASE 5 · Exportar a Excel ═══════════════════════
@login_required
def export_excel(request):
    _gate(request)
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='206BC4')

    def hoja(ws, cols, filas):
        ws.append(cols)
        for c in ws[1]:
            c.font = hdr_font; c.fill = hdr_fill
        for f in filas:
            ws.append(f)
        for i, col in enumerate(cols, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(12, len(col) + 4)

    ws = wb.active; ws.title = 'Dispositivos'
    hoja(ws, ['Nombre', 'Tipo', 'IP', 'MAC', 'VLAN', 'Ubicación', 'Estado', 'Conectividad'],
         [[d.nombre, d.get_tipo_display(), d.ip or '', d.mac_principal or '',
           (d.vlan.vlan_id if d.vlan_id else ''), str(d.ubicacion or ''), d.get_estado_display(),
           d.get_conectividad_display()] for d in Device.objects.select_related('vlan', 'ubicacion').all()])

    hoja(wb.create_sheet('Switches'), ['Nombre', 'Modelo', 'IP admin', 'Puertos', 'Ubicación', 'Estado'],
         [[s.nombre, s.modelo or '', s.ip_admin or '', s.cantidad_puertos, str(s.ubicacion or ''), s.get_estado_display()]
          for s in Switch.objects.select_related('ubicacion').all()])

    hoja(wb.create_sheet('VLANs'), ['VLAN', 'Nombre', 'Subred', 'Estado', 'IP disponibles'],
         [[v.vlan_id, v.nombre, v.subred or '', v.get_estado_display(),
           (v.cantidad_ip_disponibles if v.cantidad_ip_disponibles is not None else '')] for v in VLAN.objects.all()])

    hoja(wb.create_sheet('IPs asignadas'), ['IP', 'Estado', 'VLAN', 'Dispositivo', 'MAC', 'Hostname'],
         [[ip.direccion, ip.get_estado_display(), (ip.vlan.vlan_id if ip.vlan_id else ''),
           (ip.dispositivo.nombre if ip.dispositivo_id else ''), ip.mac or '', ip.hostname or '']
          for ip in IPAddress.objects.select_related('vlan', 'dispositivo').exclude(estado='libre')[:5000]])

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename=ana_red_inventario.xlsx'
    wb.save(resp)
    return resp
