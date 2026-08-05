import json
from datetime import date
from functools import wraps

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.contrib import messages
from django.conf import settings
from django.contrib.auth.decorators import login_required  # <--- hecho por claude code
from django.db import connections
from django.db.models import Max
from django.views.decorators.http import require_POST

from .models import (
    EjecucionCurso, InformeContable,
    TALLER_CHOICES, TALLER_LABEL, EGRESO_GRUPOS,
    CursoNota, ModuloNota, NotaIntento, HorasMetaMes, HorasParticipanteMes,
    Participante, MESES_FORMACION, MES_LABEL, NOTA_APROBADO, NOTA_APROBADO_PRACTICO,
    JORNADA_LABEL, InstructorCurso
)


def cfp_required(view):
    @wraps(view)
    def _w(request, *args, **kwargs):
        u = request.user
        if not u.is_authenticated:
            return redirect(f"{settings.LOGIN_URL}?next={request.path}")
        # <--- hecho por claude code: 'contabilidad_cfp' entra solo a la parte contable,
        # sin acceso a Notas CFP (eso lo controla cfp_notas_required).
        if not _puede_contabilidad(u):
            return HttpResponseForbidden('No tienes permiso para el módulo CFP.')
        return view(request, *args, **kwargs)
    return _w


def _anio(request):
    try:
        return int(request.GET.get('anio') or request.POST.get('anio') or date.today().year)
    except (TypeError, ValueError):
        return date.today().year


def _anios_disponibles(anio_actual):
    ys = set(EjecucionCurso.objects.values_list('anio', flat=True))
    ys.add(anio_actual)
    ys.add(date.today().year)
    return sorted(ys, reverse=True)


# ─── DASHBOARD: tarjetas por taller ───────────────────────────────────────────
@cfp_required
def dashboard(request):
    anio = _anio(request)
    qs = list(EjecucionCurso.objects.filter(anio=anio))
    talleres = []
    for key, label in TALLER_CHOICES:
        ejs = [e for e in qs if e.taller == key]
        talleres.append({
            'key': key, 'label': label, 'n': len(ejs),
            'monto': round(sum(e.monto_cfp for e in ejs), 2),
        })
    return render(request, 'cfp/dashboard.html', {
        'anio': anio, 'anios': _anios_disponibles(anio),
        'talleres': talleres,
        'total': round(sum(e.monto_cfp for e in qs), 2),
    })


# ─── SUBDASHBOARD por taller (3 tabs) ─────────────────────────────────────────
@cfp_required
def taller(request, taller):
    if taller not in TALLER_LABEL:
        return redirect('cfp:dashboard')
    anio = _anio(request)
    ejecuciones = list(EjecucionCurso.objects.filter(anio=anio, taller=taller)
                       .select_related('informe'))

    tot = {
        'horas': sum(e.horas for e in ejecuciones),
        'part': sum(e.part_pago for e in ejecuciones),
        'contrato': round(sum(e.monto_contrato for e in ejecuciones), 2),
        'perdida': round(sum(e.perdida for e in ejecuciones), 2),
        'anticipo': round(sum(e.anticipo for e in ejecuciones), 2),
        'cancelacion': round(sum(e.cancelacion for e in ejecuciones), 2),
        'costo_real': round(sum(e.costo_real for e in ejecuciones), 2),
        'cfp': round(sum(e.monto_cfp for e in ejecuciones), 2),
    }

    # Tab 2: distribución agrupada por nombre de curso
    grupos = {}
    for e in ejecuciones:
        g = grupos.setdefault(e.nombre_curso, {'curso': e.nombre_curso, 'monto': 0.0})
        g['monto'] += e.monto_cfp
    dist = []
    for g in grupos.values():
        m = round(g['monto'], 2)
        dist.append({'curso': g['curso'], 'monto': m,
                     'seguro': round(m * 0.6, 2), 'instr': round(m * 0.2, 2), 'admin': round(m * 0.2, 2)})
    dist.sort(key=lambda x: x['curso'])
    dist_tot = {
        'monto': round(sum(d['monto'] for d in dist), 2),
        'seguro': round(sum(d['seguro'] for d in dist), 2),
        'instr': round(sum(d['instr'] for d in dist), 2),
        'admin': round(sum(d['admin'] for d in dist), 2),
    }

    return render(request, 'cfp/taller.html', {
        'taller_key': taller, 'taller_label': TALLER_LABEL[taller],
        'anio': anio, 'anios': _anios_disponibles(anio),
        'ejecuciones': ejecuciones, 'tot': tot,
        'dist': dist, 'dist_tot': dist_tot,
        'es_mecanica': taller == 'mecanica',
    })


def _to_int(v, d=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return d


def _to_dec(v, d=0):
    # <--- hecho por claude code: los campos de gasto se escriben con separador de
    # miles ("5,460.00"); sin quitar las comas float() reventaría y guardaría 0.
    if isinstance(v, str):
        v = v.replace(',', '').replace(' ', '').strip()
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return d


@cfp_required
def ejecucion_guardar(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    pk = request.POST.get('pk')
    taller = request.POST.get('taller')
    if taller not in TALLER_LABEL:
        return JsonResponse({'ok': False, 'error': 'Taller inválido'}, status=400)
    obj = get_object_or_404(EjecucionCurso, pk=pk) if pk else EjecucionCurso()
    obj.anio         = _to_int(request.POST.get('anio'), date.today().year)
    obj.taller       = taller
    obj.taller_anio  = _to_int(request.POST.get('taller_anio')) or None
    # <--- hecho por claude code: si se deja vacío se genera CFP-ANA-001-26 solo;
    # si se escribe algo, manda lo escrito (los códigos viejos no se tocan).
    obj.no_ejecucion = (request.POST.get('no_ejecucion') or '').strip()
    if not obj.no_ejecucion:
        from .models import generar_no_ejecucion
        obj.no_ejecucion = generar_no_ejecucion(obj.anio)
    obj.no_curso     = (request.POST.get('no_curso') or '').strip()
    obj.no_contrato  = (request.POST.get('no_contrato') or '').strip()
    obj.nombre_curso = (request.POST.get('nombre_curso') or '').strip()
    obj.horas        = _to_int(request.POST.get('horas'))
    obj.part_inicial = _to_int(request.POST.get('part_inicial'))
    obj.part_pago    = _to_int(request.POST.get('part_pago'))
    obj.costo_hora   = _to_dec(request.POST.get('costo_hora'))
    obj.horario      = (request.POST.get('horario') or '').strip()
    if not obj.nombre_curso:
        return JsonResponse({'ok': False, 'error': 'El nombre del curso es obligatorio'}, status=400)
    obj.save()
    return JsonResponse({'ok': True})


@cfp_required
def ejecucion_detalle(request, pk):
    e = get_object_or_404(EjecucionCurso, pk=pk)
    return JsonResponse({
        'pk': e.id, 'anio': e.anio, 'taller': e.taller, 'taller_anio': e.taller_anio or '',
        'no_ejecucion': e.no_ejecucion, 'no_curso': e.no_curso, 'no_contrato': e.no_contrato,
        'nombre_curso': e.nombre_curso, 'horas': e.horas,
        'part_inicial': e.part_inicial, 'part_pago': e.part_pago, 'costo_hora': float(e.costo_hora),
        'horario': e.horario,
    })


@cfp_required
def ejecucion_eliminar(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    get_object_or_404(EjecucionCurso, pk=pk).delete()
    return JsonResponse({'ok': True})


# ─── Tab 3: Informe contable (form + PDF) ─────────────────────────────────────
def _datos_generales():
    """Registro único (compartido por todos los cursos) de datos generales."""
    from .models import CfpDatosGenerales
    dg, _ = CfpDatosGenerales.objects.get_or_create(id=1)
    return dg


def _lugar_fecha(inf):
    """<--- hecho por claude code: "Lugar; 05 de junio del año 2026" para los PDF."""
    _MESES = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio',
              'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    lugar = inf.lugar or inf.lugar_fecha
    fstr = ''
    if inf.fecha_lugar:
        f = inf.fecha_lugar
        fstr = f"{f.day:02d} de {_MESES[f.month - 1]} del año {f.year}"
    return '; '.join(p for p in (lugar, fstr) if p)


def _informe_ctx(ej, inf):
    # Sugerir la división del 20% (Instructores y Director) en los 3 gastos de
    # personal: cada campo vacío muestra su tercio (para que se vea y cuadre).
    instr20 = ej.dist_instr
    if instr20:
        tercio = round(instr20 / 3, 2)
        defaults = {
            'personal_instructor': tercio,
            'personal_encargado':  tercio,
            'personal_apoyo':      round(instr20 - 2 * tercio, 2),
        }
        eg = dict(inf.egresos or {})
        for k, v in defaults.items():
            if not eg.get(k):           # vacío / ausente → sugerir división
                eg[k] = v
        inf.egresos = eg  # solo en memoria (no se guarda salvo POST)

    # <--- hecho por claude code: cada grupo lleva su % de la distribución del curso
    # (20/60/20) y el monto al que debería cuadrar, para que el contador lo vea al llenar.
    from .models import EGRESO_GRUPO_PCT
    montos = {'personal': ej.dist_instr, 'materia': ej.dist_seguro, 'administracion': ej.dist_admin}

    from .models import _CLAVES_CARGO
    grupos = []
    for g, lbl, items in EGRESO_GRUPOS:
        # <--- hecho por claude code: los cargos de Administración son de solo lectura,
        # su monto sale del reparto de la planilla.
        filas = [{'key': f'{g}_{k}', 'label': il,
                  'valor': inf.egresos.get(f'{g}_{k}', ''),
                  'calculado': g == 'administracion' and k in _CLAVES_CARGO}
                 for k, il in items]
        nota = ''
        if g in EGRESO_GRUPO_PCT:
            pct, desc = EGRESO_GRUPO_PCT[g]
            nota = f'↳ {pct:.0%} {desc} — L {montos.get(g, 0):,.2f}'
            if g == 'personal':
                nota += ' dividido en 3'
        grupos.append({'key': g, 'label': lbl, 'filas': filas, 'subtotal': inf.grupo_subtotal(g),
                       'nota': nota})

    # Datos generales COMPARTIDOS: sobrescriben los del informe (en memoria, para mostrar/PDF)
    dg = _datos_generales()
    inf.localidad   = dg.localidad
    inf.direccion   = dg.direccion
    inf.instructor  = dg.instructor
    inf.horario     = dg.horario
    inf.lugar       = dg.lugar
    inf.fecha_lugar = dg.fecha_lugar
    # <--- hecho por claude code: Regional y Centro-Prog-Proy son del centro, no del curso
    inf.regional        = dg.regional
    inf.centro_programa = dg.centro_programa

    # Opciones de los selects agregables (catálogo + valor guardado actual)
    from .models import CfpOpcion
    actuales = {'localidad': inf.localidad, 'direccion': inf.direccion,
                'instructor': inf.instructor, 'horario': inf.horario or ej.horario,
                'lugar': inf.lugar}
    opciones = {}
    for campo, cur in actuales.items():
        vals = set(CfpOpcion.objects.filter(campo=campo).values_list('valor', flat=True))
        if cur:
            vals.add(cur)
        opciones[campo] = sorted(vals)

    lugar_fecha_str = _lugar_fecha(inf)

    return {'ej': ej, 'inf': inf, 'grupos': grupos, 'opciones': opciones,
            'lugar_fecha_str': lugar_fecha_str,
            'taller_label': TALLER_LABEL.get(ej.taller, ej.taller)}


@cfp_required
def informe_form(request, pk):
    ej = get_object_or_404(EjecucionCurso, pk=pk)
    inf, _ = InformeContable.objects.get_or_create(ejecucion=ej)
    # <--- hecho por claude code: autoguardado — el mismo POST, pero devolviendo JSON
    es_ajax = request.GET.get('ajax') == '1'
    if request.method == 'POST':
        from .models import CfpOpcion
        def _opcion(campo):
            """Resuelve un select agregable: si es '__nueva__' toma el texto y lo cataloga."""
            val = (request.POST.get(campo) or '').strip()
            if val == '__nueva__':
                val = (request.POST.get(campo + '_nueva') or '').strip()
            # <--- hecho por claude code: al autoguardar NO se cataloga mientras se teclea
            # (si no entrarían "Aldea de La V", "Aldea de La Ve"...). El JS manda
            # catalogar=1 al SALIR del campo, que es cuando el texto ya está completo.
            if val and (not es_ajax or request.POST.get('catalogar') == '1'):
                CfpOpcion.objects.get_or_create(campo=campo, valor=val)
            return val
        # Datos generales COMPARTIDOS por todos los cursos (registro único)
        dg = _datos_generales()
        dg.localidad   = _opcion('localidad')
        dg.direccion   = _opcion('direccion')
        dg.instructor  = _opcion('instructor')
        dg.horario     = _opcion('horario')
        dg.lugar       = _opcion('lugar')
        dg.fecha_lugar = request.POST.get('fecha_lugar') or None
        dg.save()

        inf.fecha_inicio    = request.POST.get('fecha_inicio') or None
        inf.fecha_fin       = request.POST.get('fecha_fin') or None
        inf.convenio        = (request.POST.get('convenio') or '').strip()
        # <--- hecho por claude code: regional y centro_programa ya NO se leen del POST;
        # son del centro (CfpDatosGenerales) y se muestran de solo lectura.
        inf.regional        = dg.regional
        inf.centro_programa = dg.centro_programa
        inf.egresados       = _to_int(request.POST.get('egresados'))
        # <--- hecho por claude code: los cargos de Administración NO se escriben aquí;
        # los calcula la planilla. Se ignora lo que venga del POST y se re-sincroniza.
        from .models import _CLAVES_CARGO
        calculadas = {f'administracion_{k}' for k in _CLAVES_CARGO}
        egresos = {}
        for g, _lbl, items in EGRESO_GRUPOS:
            for k, _il in items:
                clave = f'{g}_{k}'
                if clave in calculadas:
                    continue
                val = _to_dec(request.POST.get(clave))
                if val:
                    egresos[clave] = val
        inf.egresos = egresos
        inf.save()
        inf.sincronizar_cargos()   # el seguro/teléfono cambió → cambia el reparto
        if es_ajax:
            # Devuelve los cargos recalculados para refrescarlos sin recargar la página.
            return JsonResponse({
                'ok': True,
                'cargos': {f'administracion_{k}': inf.egresos.get(f'administracion_{k}', 0) or 0
                           for k in _CLAVES_CARGO},
                'a_repartir': inf.admin_a_repartir,
            })
        messages.success(request, 'Informe guardado.')
        return redirect(request.path)
    return render(request, 'cfp/informe_form.html', _informe_ctx(ej, inf))


def _planilla_ctx(ej, inf):
    """<--- hecho por claude code: contexto común de la planilla (pantalla y PDF)."""
    from .models import CARGOS_ADMIN, ADMIN_CAPTURADOS
    capturados = [{'label': lbl,
                   'valor': inf.egresos.get(f'administracion_{k}', 0) or 0}
                  for k, lbl in ADMIN_CAPTURADOS]
    reparto = inf.reparto_planilla()
    return {
        'ej': ej, 'inf': inf,
        'cargos':      CARGOS_ADMIN,
        'capturados':  capturados,
        'reparto':     [{'p': p, 'monto': m} for p, m in reparto],
        'total_admin': ej.dist_admin,
        'capturado':   inf.admin_capturado,
        'a_repartir':  inf.admin_a_repartir,
        'n_personas':  len(reparto),
        'lugar_fecha_str': _lugar_fecha(inf),
        'taller_label': TALLER_LABEL.get(ej.taller, ej.taller),
    }


@cfp_required
def planilla_form(request, pk):
    """Planilla de gastos administrativos: quién cobra el resto del 20%."""
    from .models import GastoAdministrativo
    ej = get_object_or_404(EjecucionCurso, pk=pk)
    inf, _ = InformeContable.objects.get_or_create(ejecucion=ej)

    if request.method == 'POST':
        # Se reescribe la planilla completa con lo que venga del formulario.
        nombres = request.POST.getlist('nombre')
        dnis    = request.POST.getlist('dni')
        cargos  = request.POST.getlist('cargo')
        ej.planilla.all().delete()
        nuevas = []
        for i, nombre in enumerate(nombres):
            nombre = (nombre or '').strip()
            if not nombre:                      # una fila sin nombre no es nadie
                continue
            nuevas.append(GastoAdministrativo(
                ejecucion=ej, nombre=nombre,
                dni=(dnis[i] if i < len(dnis) else '').strip(),
                cargo=(cargos[i] if i < len(cargos) else '').strip(),
                orden=i))
        GastoAdministrativo.objects.bulk_create(nuevas)
        inf.sincronizar_cargos()                # el informe refleja la planilla
        # <--- hecho por claude code: autoguardado — mismo POST, respuesta JSON
        if request.GET.get('ajax') == '1':
            return JsonResponse({'ok': True,
                                 'a_repartir': inf.admin_a_repartir,
                                 'personas': len(nuevas)})
        messages.success(request, 'Planilla guardada.')
        return redirect(request.path)

    ctx = _planilla_ctx(ej, inf)
    # Sugerencias de personas ya usadas en otros cursos (evita reescribir el DNI)
    ctx['conocidos'] = list(GastoAdministrativo.objects
                            .exclude(nombre='')
                            .values('nombre', 'dni', 'cargo').distinct()[:200])
    return render(request, 'cfp/planilla_form.html', ctx)


@cfp_required
def planilla_pdf(request, pk):
    ej = get_object_or_404(EjecucionCurso, pk=pk)
    inf, _ = InformeContable.objects.get_or_create(ejecucion=ej)
    from weasyprint import HTML, CSS
    from django.contrib.staticfiles import finders

    ctx = _planilla_ctx(ej, inf)
    logo = finders.find('cfp/img/logo_ana.png')
    ctx['logo_url'] = f'file://{logo}' if logo else ''
    html  = render(request, 'cfp/planilla_pdf.html', ctx).content.decode('utf-8')
    hoja  = finders.find('cfp/css/planilla_pdf.css')
    pdf = HTML(string=html).write_pdf(stylesheets=[CSS(filename=hoja)] if hoja else [])
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="planilla_{ej.no_ejecucion or ej.id}.pdf"'
    return resp


@cfp_required
def informe_pdf(request, pk):
    ej = get_object_or_404(EjecucionCurso, pk=pk)
    inf, _ = InformeContable.objects.get_or_create(ejecucion=ej)
    from weasyprint import HTML, CSS
    from django.contrib.staticfiles import finders

    # <--- hecho por claude code: el CSS y el logo se pasan por RUTA DE DISCO.
    # Con HTML(string=...) WeasyPrint no tiene base_url, así que un <link> a
    # /static/... no se resolvía y el PDF salía SIN estilos (y en A4, no legal).
    ctx = _informe_ctx(ej, inf)
    logo = finders.find('cfp/img/logo_ana.png')
    ctx['logo_url'] = f'file://{logo}' if logo else ''

    html  = render(request, 'cfp/informe_pdf.html', ctx).content.decode('utf-8')
    hoja  = finders.find('cfp/css/informe_pdf.css')
    hojas = [CSS(filename=hoja)] if hoja else []
    pdf = HTML(string=html).write_pdf(stylesheets=hojas)
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="informe_{ej.no_ejecucion or ej.id}.pdf"'
    return resp


# ══════════════════════════════════════════════════════════════════════════════
#  PROGRAMA 2 — SISTEMA DE NOTAS CFP
# ══════════════════════════════════════════════════════════════════════════════

# Cursos candidatos (de SQL Server). Mismo IN(...) del query del usuario.
CURSOS_CFP = [
    'Lubricador Engrasador', 'Asistente de Mecánico Automotríz', 'Ajustador de Motores',
    'Afinador de Motores', 'Mecánico de Sistema de Transmisión', 'Electricista Automotriz',
    'Mecánico de Precisión',
    'Panadero', 'Repostero', 'Asistente Técnico en Computación', 'Estilista de Belleza',
]
# Etiqueta de visualización (el query SQL sigue usando el nombre real de tc.Descripcion)
CURSO_LABEL = {
    'Lubricador Engrasador':              'Mec. Automotriz 1° · Lubricador Engrasador',
    'Asistente de Mecánico Automotríz':   'Mec. Automotriz 1° · Asistente de Mecánica',
    'Ajustador de Motores':               'Mec. Automotriz 1° · Mecánico Ajustador',
    'Afinador de Motores':                'Mec. Automotriz 2° · Afinador de Motores',
    'Mecánico de Sistema de Transmisión': 'Mec. Automotriz 2° · Mecánico de Sist. Transmisión',
    'Electricista Automotriz':            'Mec. Automotriz 2° · Electricista Automotriz',
    'Panadero':                           'Panadería / Repostería · Panadero',
    'Repostero':                          'Panadería / Repostería · Repostería',
}

# Cursos con dos jornadas (Matutina/Vespertina): roster manual por grupo.
SPLIT_CURSOS = {'Asistente Técnico en Computación', 'Panadero', 'Repostero', 'Estilista de Belleza'}
JORNADAS = [('matutina', 'Matutina'), ('vespertina', 'Vespertina')]
# Cursos que comparten el roster de otro (mismos alumnos, distinto plan de módulos).
# Repostería usa el roster (dividido) de Panadero; los de Mec. Automotriz 1° usan el de Lubricador.
ROSTER_ALIAS = {
    'Repostero': 'Panadero',
    'Asistente de Mecánico Automotríz': 'Lubricador Engrasador',
    'Ajustador de Motores': 'Lubricador Engrasador',
    'Mecánico de Sistema de Transmisión': 'Afinador de Motores',
    'Electricista Automotriz': 'Afinador de Motores',
}


def _label(curso):
    return CURSO_LABEL.get(curso, curso)


MAX_MODULOS = 25


def _puede_contabilidad(u):
    """<--- hecho por claude code: quién entra a Contabilidad (Programa 1)."""
    return u.is_superuser or u.groups.filter(
        name__in=['director_cfp', 'contabilidad_cfp']).exists()


def _cursos_instructor(u):
    """<--- hecho por claude code: cursos asignados a un instructor.
    Devuelve None si ve TODOS (superusuario o director); un set de nombres si es
    instructor con asignaciones."""
    if u.is_superuser or u.groups.filter(name='director_cfp').exists():
        return None
    return set(InstructorCurso.objects.filter(instructor=u).values_list('curso', flat=True))


def _puede_notas(u):
    """Quién entra a Notas (Programa 2)."""
    return u.is_superuser or u.groups.filter(
        name__in=['director_cfp', 'instructores']).exists()


def _es_director(u):
    """Director del CFP. OJO: no confundir con 'puede entrar a Contabilidad'
    (eso lo decide cfp_required, que además acepta 'contabilidad_cfp')."""
    return u.is_superuser or u.groups.filter(name='director_cfp').exists()


def _n_participantes(anio, curso, conteo):
    """Participantes del curso: alias a curso en vivo, roster guardado (divididos) o conteo SQL."""
    base = ROSTER_ALIAS.get(curso)
    if base and base not in SPLIT_CURSOS:        # alias a un curso en vivo (ej. Lubricador)
        return conteo.get(base, 0)
    if curso in SPLIT_CURSOS:
        return Participante.objects.filter(curso__anio=anio, curso__curso=curso).count()
    return conteo.get(curso, 0)


def cfp_notas_required(view):
    """Acceso al Programa 2: superuser, director_cfp o instructores."""
    @wraps(view)
    def _w(request, *args, **kwargs):
        u = request.user
        if not u.is_authenticated:
            return redirect(f"{settings.LOGIN_URL}?next={request.path}")
        if not _puede_notas(u):
            return HttpResponseForbidden('No tienes permiso para el módulo de Notas CFP.')
        return view(request, *args, **kwargs)
    return _w


def _conteo_cursos(anio):
    """{curso: n_participantes} desde SQL Server para el año dado."""
    sql = """
      SELECT tc.Descripcion AS Curso, COUNT(*) AS N
      FROM tblEdcIngrPrevisto ip
        INNER JOIN tblPrsDtosGen d ON d.PersonaID = ip.PersonaID
        INNER JOIN tblEdcTipoCrso tc ON tc.TipoCrsoID = ip.TipoCursoID
      WHERE ip.Año = %s AND tc.Descripcion IN (%s)
      GROUP BY tc.Descripcion ORDER BY tc.Descripcion
    """ % ('%s', ','.join(['%s'] * len(CURSOS_CFP)))
    try:
        with connections['padres_sqlserver'].cursor() as c:
            c.execute(sql, [anio] + CURSOS_CFP)
            return {row[0]: row[1] for row in c.fetchall()}
    except Exception as e:
        print('[cfp notas] ERROR conteo SQL Server:', e)
        return {}


def _participantes_curso(anio, curso):
    """Lista de participantes de un curso desde SQL Server: [{persona_id, nombre, identidad}]."""
    sql = """
      SELECT
        ip.PersonaID,
        LTRIM(RTRIM(
          ISNULL(d.Nombre1,'')   + ' ' + ISNULL(d.Nombre2,'') + ' ' +
          ISNULL(d.Apellido1,'') + ' ' + ISNULL(d.Apellido2,''))) AS Nombre,
        d.NumeroID AS Identidad,
        d.Sexo
      FROM tblEdcIngrPrevisto ip
        INNER JOIN tblPrsDtosGen d ON d.PersonaID = ip.PersonaID
        INNER JOIN tblEdcTipoCrso tc ON tc.TipoCrsoID = ip.TipoCursoID
      WHERE ip.Año = %s AND tc.Descripcion = %s
      ORDER BY d.Sexo, d.Nombre1, d.Nombre2
    """
    try:
        with connections['padres_sqlserver'].cursor() as c:
            c.execute(sql, [anio, curso])
            return [{'persona_id': int(r[0]),
                     'nombre': ' '.join((r[1] or '').split()),
                     'identidad': (r[2] or '').strip(),
                     'sexo': (r[3] or '').strip()} for r in c.fetchall()]
    except Exception as e:
        print(f'[cfp notas] ERROR participantes SQL Server ({curso}):', e)
        return []


def _color_intento(intentos, umbral=NOTA_APROBADO):
    """Clase de color del intento con que aprobó: verde/azul/rojo (1º/2º/3º) o negro."""
    clases = ['nota-verde', 'nota-azul', 'nota-rojo']
    for i, v in enumerate(intentos):
        if v is not None and float(v) >= umbral:
            return clases[i]
    return 'nota-negro'


@login_required
def programas(request):
    """Landing de CFP. <--- hecho por claude code: antes exigía permiso de NOTAS,
    así que quien solo tiene Contabilidad recibía 403 al entrar al módulo. Ahora
    basta con tener acceso a alguna parte, y cada tarjeta aparece según su permiso."""
    u = request.user
    if not (_puede_contabilidad(u) or _puede_notas(u)):
        return HttpResponseForbidden('No tienes permiso para el módulo CFP.')
    return render(request, 'cfp/programas.html', {
        'puede_contabilidad': _puede_contabilidad(u),
        'puede_notas':        _puede_notas(u),
        'es_director':        _es_director(u),
    })


@cfp_notas_required
def notas_resumen(request):
    """Panel global (solo director): todos los cursos y jornadas con su resumen."""
    if not _es_director(request.user):
        return HttpResponseForbidden('Solo el director CFP puede ver el panel global.')
    anio = _anio(request)
    conteo = _conteo_cursos(anio)
    filas = []
    tot = {'part': 0, 'capt': 0}
    for curso in CURSOS_CFP:
        if not _n_participantes(curso=curso, anio=anio, conteo=conteo):
            continue
        grupos = JORNADAS if curso in SPLIT_CURSOS else [('unica', '')]
        for jkey, jlabel in grupos:
            cn = CursoNota.objects.filter(anio=anio, curso=curso, jornada=jkey).first()
            modulos = list(cn.modulos.all()) if cn else []
            parts = _roster(cn, anio, curso) if cn else (
                [] if curso in SPLIT_CURSOS else _participantes_curso(anio, curso))
            notas = ({(n.modulo_id, n.persona_id): n
                      for n in NotaIntento.objects.filter(modulo__curso=cn)} if cn else {})
            finals, capt = [], 0
            for p in parts:
                pid = p['persona_id']
                s, has = 0.0, False
                for m in modulos:
                    n = notas.get((m.id, pid))
                    if n:
                        has = True
                        s += n.resultado_modulo
                if has:
                    capt += 1
                finals.append(round(s / len(modulos), 2) if modulos else 0)
            prom = round(sum(finals) / len(finals), 2) if finals else 0
            filas.append({
                'curso': curso, 'label': _label(curso), 'jornada': jkey, 'jlabel': jlabel,
                'es_split': curso in SPLIT_CURSOS,
                'n_part': len(parts), 'n_mod': len(modulos), 'capt': capt, 'prom': prom,
            })
            tot['part'] += len(parts)
            tot['capt'] += capt
    anios = sorted({anio, date.today().year, 2026}, reverse=True)
    return render(request, 'cfp/notas_resumen.html', {
        'anio': anio, 'anios': anios, 'filas': filas, 'tot': tot,
    })


@cfp_notas_required
def notas_cursos(request):
    """Lista de cursos del Programa 2 con su conteo de participantes (de SQL Server)."""
    anio = _anio(request)
    conteo = _conteo_cursos(anio)
    permitidos = _cursos_instructor(request.user)   # None = todos
    cursos = []
    for nombre in CURSOS_CFP:
        if permitidos is not None and nombre not in permitidos:
            continue
        n = _n_participantes(anio, nombre, conteo)
        if n:
            cn = CursoNota.objects.filter(anio=anio, curso=nombre).first()
            cursos.append({'curso': nombre, 'label': _label(nombre), 'n': n, 'configurado': bool(cn),
                           'n_modulos': cn.modulos.count() if cn else 0})
    anios = sorted({anio, date.today().year, 2026}, reverse=True)
    return render(request, 'cfp/notas_cursos.html', {
        'anio': anio, 'anios': anios, 'cursos': cursos,
        'total_cursos': len(cursos), 'es_director': _es_director(request.user),
    })


def _curso_obj(anio, curso, jornada='unica'):
    obj, _ = CursoNota.objects.get_or_create(anio=anio, curso=curso, jornada=jornada)
    return obj


def _jornada_de(request, curso):
    """Resuelve la jornada activa: 'unica' o matutina/vespertina (cursos divididos)."""
    if curso not in SPLIT_CURSOS:
        return 'unica'
    j = request.GET.get('jornada')
    return j if j in dict(JORNADAS) else 'matutina'


def _roster(cn, anio, curso):
    """Participantes del grupo: alias a curso en vivo, roster guardado (divididos) o SQL Server."""
    base = ROSTER_ALIAS.get(curso)
    if base and base not in SPLIT_CURSOS:        # mismos alumnos del curso base (ej. Lubricador)
        return _participantes_curso(anio, base)
    if curso in SPLIT_CURSOS:
        return [{'persona_id': p.persona_id, 'nombre': p.nombre, 'identidad': p.identidad}
                for p in cn.participantes.all()]
    return _participantes_curso(anio, curso)


@cfp_notas_required
def notas_curso(request, anio, curso):
    # <--- hecho por claude code: un instructor solo entra a SUS cursos
    _permitidos = _cursos_instructor(request.user)
    if _permitidos is not None and curso not in _permitidos:
        return HttpResponseForbidden('Este curso no está asignado a tu usuario.')
    """Vista del curso con los 4 tabs (Progreso · Compilación · Módulos · Horas)."""
    if curso not in CURSOS_CFP:
        return redirect('cfp:notas_cursos')
    jornada = _jornada_de(request, curso)
    cn = _curso_obj(anio, curso, jornada)
    modulos = list(cn.modulos.all())
    participantes = _roster(cn, anio, curso)

    # Mapa de notas guardadas: (modulo_id, persona_id) -> NotaIntento
    notas = {(n.modulo_id, n.persona_id): n
             for n in NotaIntento.objects.filter(modulo__curso=cn)}
    # Fallback de nombre/identidad si SQL Server no responde
    if not participantes and notas:
        vistos = {}
        for n in notas.values():
            vistos.setdefault(n.persona_id, {'persona_id': n.persona_id,
                                             'nombre': n.nombre, 'identidad': n.identidad})
        participantes = sorted(vistos.values(), key=lambda x: x['nombre'])

    # Horas (Tab 4)
    meta = {h.mes: float(h.horas) for h in cn.horas_meta.all()}
    horas_part = {(h.persona_id, h.mes): float(h.horas) for h in cn.horas_part.all()}
    meses = [{'mes': m, 'label': MES_LABEL[m], 'meta': meta.get(m, 0)} for m in MESES_FORMACION]
    total_meta = round(sum(meta.get(m, 0) for m in MESES_FORMACION), 2)

    suma_puntaje = sum(float(m.puntaje) for m in modulos) or 0

    filas = []
    for i, p in enumerate(participantes, 1):
        pid = p['persona_id']
        celdas, suma_result = [], 0.0
        for m in modulos:
            n = notas.get((m.id, pid))
            t = [n.t1, n.t2, n.t3] if n else [None, None, None]
            pr = [n.p1, n.p2, n.p3] if n else [None, None, None]
            compT = n.teorico_compilado if n else 0.0
            compP = n.practico_compilado if n else 0.0
            resultado = round((compT + compP) / 2, 2)
            suma_result += resultado
            celdas.append({
                'modulo_id': m.id,
                't': [('' if v is None else v) for v in t],
                'p': [('' if v is None else v) for v in pr],
                't_color': _color_intento(t),
                'p_color': _color_intento(pr, NOTA_APROBADO_PRACTICO),
                'compT': compT, 'compP': compP, 'resultado': resultado,
            })
        nota_final = round(suma_result / len(modulos), 2) if modulos else 0
        # Horas del participante
        h_part = [{'mes': m, 'horas': horas_part.get((pid, m), 0)} for m in MESES_FORMACION]
        total_h = round(sum(horas_part.get((pid, m), 0) for m in MESES_FORMACION), 2)
        pct = round(total_h / total_meta * 100, 1) if total_meta else 0
        filas.append({
            'no': i, 'persona_id': pid, 'nombre': p['nombre'], 'identidad': p['identidad'],
            'celdas': celdas, 'nota_final': nota_final,
            'horas': h_part, 'total_horas': total_h, 'pct': pct,
        })

    return render(request, 'cfp/notas_curso.html', {
        'cn': cn, 'anio': anio, 'curso': curso, 'curso_label': _label(curso),
        'modulos': modulos, 'n_modulos': len(modulos), 'max_modulos': MAX_MODULOS,
        'filas': filas, 'n_part': len(participantes),
        'suma_puntaje': suma_puntaje,
        'meses': meses, 'total_meta': total_meta,
        'es_director': _es_director(request.user),
        'es_split': curso in SPLIT_CURSOS, 'jornada': jornada,
        'jornada_label': JORNADA_LABEL.get(jornada, ''), 'jornadas': JORNADAS,
    })


# ── Guardados AJAX (JSON) ─────────────────────────────────────────────────────
def _body(request):
    try:
        return json.loads(request.body or '{}')
    except (ValueError, TypeError):
        return {}


def _dec(v):
    """A Decimal-compatible float o None si vacío."""
    if v in (None, '', 'null'):
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _cursos_jornadas(cn):
    """CursoNota que comparten módulos: ambas jornadas si el curso es dividido; si no, solo cn."""
    if cn.curso in SPLIT_CURSOS:
        return [_curso_obj(cn.anio, cn.curso, j) for j, _ in JORNADAS]
    return [cn]


@cfp_notas_required
@require_POST
def modulo_guardar(request):
    """Crea o edita un módulo. En cursos divididos se espeja en ambas jornadas."""
    d = _body(request)
    cn = get_object_or_404(CursoNota, pk=d.get('curso_pk'))
    targets = _cursos_jornadas(cn)

    # Prefijo del curso (se aplica a todas las jornadas)
    prefijo = (d.get('prefijo') or '').strip()
    if prefijo:
        for t in targets:
            if t.codigo != prefijo:
                t.codigo = prefijo
                t.save(update_fields=['codigo'])

    codigo_in = (d.get('codigo') or '').strip()
    puntaje   = _dec(d.get('puntaje')) or 100
    fini      = d.get('fecha_inicio') or None
    ffin      = d.get('fecha_fin') or None

    pk = d.get('pk')
    if pk:
        base = get_object_or_404(ModuloNota, pk=pk)
        numero = base.numero
    else:
        mx = max((t.modulos.aggregate(x=Max('numero'))['x'] or 0) for t in targets)
        if mx >= MAX_MODULOS:
            return JsonResponse({'ok': False, 'error': f'Máximo {MAX_MODULOS} módulos.'}, status=400)
        numero = mx + 1

    for t in targets:
        m, _ = ModuloNota.objects.get_or_create(curso=t, numero=numero)
        m.codigo       = codigo_in or f"{t.prefijo}-{numero:02d}"
        m.puntaje      = puntaje
        m.fecha_inicio = fini
        m.fecha_fin    = ffin
        m.save()
    return JsonResponse({'ok': True})


@cfp_notas_required
@require_POST
def modulo_eliminar(request):
    """Elimina el módulo en todas las jornadas que lo comparten."""
    d = _body(request)
    m = get_object_or_404(ModuloNota, pk=d.get('pk'))
    targets = _cursos_jornadas(m.curso)
    ModuloNota.objects.filter(curso__in=targets, numero=m.numero).delete()
    return JsonResponse({'ok': True})


@cfp_notas_required
@require_POST
def notas_guardar(request):
    """Guarda la cuadrícula de notas por intento (Tab 1)."""
    d = _body(request)
    cn = get_object_or_404(CursoNota, pk=d.get('curso_pk'))
    mod_ids = set(cn.modulos.values_list('id', flat=True))
    n = 0
    for row in d.get('notas', []):
        mid = row.get('modulo'); pid = row.get('persona')
        if mid not in mod_ids or not pid:
            continue
        vals = {k: _dec(row.get(k)) for k in ('t1', 't2', 't3', 'p1', 'p2', 'p3')}
        if all(v is None for v in vals.values()):
            NotaIntento.objects.filter(modulo_id=mid, persona_id=pid).delete()
            continue
        obj, _ = NotaIntento.objects.get_or_create(modulo_id=mid, persona_id=pid)
        obj.nombre = (row.get('nombre') or obj.nombre)[:200]
        obj.identidad = (row.get('identidad') or obj.identidad)[:40]
        for k, v in vals.items():
            setattr(obj, k, v)
        obj.save()
        n += 1
    return JsonResponse({'ok': True, 'guardados': n})


@cfp_notas_required
@require_POST
def horas_guardar(request):
    """Guarda horas-meta (solo director) y horas reales por participante (Tab 4)."""
    d = _body(request)
    cn = get_object_or_404(CursoNota, pk=d.get('curso_pk'))
    if _es_director(request.user):
        for mes, val in (d.get('meta') or {}).items():
            try:
                mes = int(mes)
            except (TypeError, ValueError):
                continue
            if mes in MESES_FORMACION:
                HorasMetaMes.objects.update_or_create(
                    curso=cn, mes=mes, defaults={'horas': _dec(val) or 0})
    for row in d.get('part', []):
        pid = row.get('persona'); mes = row.get('mes')
        try:
            mes = int(mes)
        except (TypeError, ValueError):
            continue
        if not pid or mes not in MESES_FORMACION:
            continue
        HorasParticipanteMes.objects.update_or_create(
            curso=cn, persona_id=pid, mes=mes, defaults={'horas': _dec(row.get('horas')) or 0})
    return JsonResponse({'ok': True})


@cfp_notas_required
def notas_excel(request, anio, curso):
    """Exporta a Excel las 4 vistas (Progreso, Compilación, Módulos, Horas)."""
    if curso not in CURSOS_CFP:
        return redirect('cfp:notas_cursos')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    jornada = _jornada_de(request, curso)
    cn = _curso_obj(anio, curso, jornada)
    modulos = list(cn.modulos.all())
    participantes = _roster(cn, anio, curso)
    notas = {(n.modulo_id, n.persona_id): n
             for n in NotaIntento.objects.filter(modulo__curso=cn)}
    if not participantes and notas:
        vistos = {}
        for n in notas.values():
            vistos.setdefault(n.persona_id, {'persona_id': n.persona_id,
                                             'nombre': n.nombre, 'identidad': n.identidad})
        participantes = sorted(vistos.values(), key=lambda x: x['nombre'])
    meta = {h.mes: float(h.horas) for h in cn.horas_meta.all()}
    horas_part = {(h.persona_id, h.mes): float(h.horas) for h in cn.horas_part.all()}
    total_meta = round(sum(meta.get(m, 0) for m in MESES_FORMACION), 2)

    wb = Workbook()
    hdr = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='2FB344')
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def cab(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = hdr; c.fill = fill; c.alignment = ctr

    # Hoja 1 — Progreso (intentos)
    ws1 = wb.active; ws1.title = 'Progreso'
    cols = ['No', 'Participante', 'Identidad']
    for m in modulos:
        cols += [f'{m.codigo_auto} T1', 'T2', 'T3', 'P1', 'P2', 'P3']
    cab(ws1, cols)
    # Hoja 2 — Compilación
    ws2 = wb.create_sheet('Compilación')
    c2 = ['No', 'Participante']
    for m in modulos:
        c2 += [f'{m.codigo_auto} T', 'P']
    cab(ws2, c2)
    # Hoja 3 — Módulos (resultado) + Nota Final
    ws3 = wb.create_sheet('Módulos')
    c3 = ['No', 'Participante'] + [m.codigo_auto for m in modulos] + ['Nota Final']
    cab(ws3, c3)
    # Hoja 4 — Horas
    ws4 = wb.create_sheet('Horas')
    c4 = ['No', 'Participante'] + [MES_LABEL[m] for m in MESES_FORMACION] + ['Total', '% Asist.']
    cab(ws4, c4)
    ws4.append(['', 'META (jornada)'] + [meta.get(m, 0) for m in MESES_FORMACION] + [total_meta, ''])

    for i, p in enumerate(participantes, 1):
        pid = p['persona_id']
        r1 = [i, p['nombre'], p['identidad']]
        r2 = [i, p['nombre']]
        r3 = [i, p['nombre']]
        suma_result = 0.0
        for m in modulos:
            n = notas.get((m.id, pid))
            r1 += [('' if not n or getattr(n, k) is None else float(getattr(n, k)))
                   for k in ('t1', 't2', 't3', 'p1', 'p2', 'p3')]
            compT = n.teorico_compilado if n else 0.0
            compP = n.practico_compilado if n else 0.0
            r2 += [compT, compP]
            res = round((compT + compP) / 2, 2)
            suma_result += res
            r3.append(res)
        r3.append(round(suma_result / len(modulos), 2) if modulos else 0)
        ws1.append(r1); ws2.append(r2); ws3.append(r3)
        th = round(sum(horas_part.get((pid, m), 0) for m in MESES_FORMACION), 2)
        pct = round(th / total_meta * 100, 1) if total_meta else 0
        ws4.append([i, p['nombre']] + [horas_part.get((pid, m), 0) for m in MESES_FORMACION] + [th, pct])

    # ── Formato uniforme: Arial 11, centrado (h+v) y todos los bordes ──
    arial = Font(name='Arial', size=11)
    borde = Border(*(4 * [Side(style='thin', color='000000')]))
    for ws in wb.worksheets:
        for fila in ws.iter_rows():
            for c in fila:
                c.border = borde
                c.alignment = ctr
                if c.row != 1:           # encabezado conserva su fuente blanca/negrita
                    c.font = arial
        # Ancho de columnas: Participante más ancho, el resto fijo
        for col in range(1, ws.max_column + 1):
            letra = get_column_letter(col)
            encab = (ws.cell(row=1, column=col).value or '')
            ws.column_dimensions[letra].width = 26 if encab == 'Participante' else 12
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'C2'

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="notas_{curso}_{anio}.xlsx"'
    wb.save(resp)
    return resp
