# ==============================================================
# INVENTARIO – VISTAS COMPLETAS Y COMPATIBLES CON AJAX + MODALES
# ==============================================================

import datetime
import io
import qrcode
from PIL import Image

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.db.models import F, Value, CharField
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import cm

from .models import (
    InventoryItem,
    Computadora, Televisor, Impresora, Router, DataShow, Monitor,
    ModeloMonitor,
    ModeloTelevisor, GradoTelevisor, AreaTelevisor,
    ModeloComputadora, SerieComputadora, AsignadoAComputadora,
    AreaComputadora, GradoComputadora, SubtipoGradoComputadora, EdificioComputadora,
)

from .forms import (
    CategoryUpdateForm,
    ComputadoraForm,
    TelevisorForm,
    ImpresoraForm,
    RouterForm,
    DataShowForm,
    ComputadoraFilterForm,
    MonitorForm
)

# ==============================================================
# QR PARA FICHAS
# ==============================================================

def _get_asset_id(tipo, pk):
    tipo = tipo.lower()
    mapa = {
        'computadora': Computadora,
        'televisor':   Televisor,
        'impresora':   Impresora,
        'router':      Router,
        'datashow':    DataShow,
        'monitor':     Monitor,
    }
    modelo = mapa.get(tipo)
    if modelo:
        obj = modelo.objects.filter(pk=pk).first()
        if obj:
            return obj.asset_id
    return f"{tipo.upper()}-{pk}"


def descargar_qr(request, tipo, pk):
    from PIL import ImageDraw, ImageFont

    path    = reverse("inventario:download_model_pdf", args=[tipo.lower(), pk])
    pdf_url = f"https://servicios.ana-hn.org:437{path}"
    asset_id = _get_asset_id(tipo, pk)

    qr_img = qrcode.make(pdf_url).convert("RGB")
    qr_w, qr_h = qr_img.size

    # Barra inferior con el asset_id
    bar_h    = 48
    total_h  = qr_h + bar_h
    canvas   = Image.new("RGB", (qr_w, total_h), "white")
    canvas.paste(qr_img, (0, 0))

    draw = ImageDraw.Draw(canvas)
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 22)
    except Exception:
        font = ImageFont.load_default()

    bbox = draw.textbbox((0, 0), asset_id, font=font)
    text_w = bbox[2] - bbox[0]
    x = (qr_w - text_w) // 2
    y = qr_h + (bar_h - (bbox[3] - bbox[1])) // 2
    draw.text((x, y), asset_id, fill="black", font=font)

    buffer = io.BytesIO()
    canvas.save(buffer, format="JPEG", quality=90)
    buffer.seek(0)

    return HttpResponse(
        buffer.read(),
        content_type="image/jpeg",
        headers={"Content-Disposition": f'attachment; filename="qr_{asset_id}.jpg"'}
    )

# ==============================================================
# DASHBOARD
# ==============================================================

@login_required
def dashboard(request):
    year = datetime.datetime.now().year
    return render(request, "inventario/dashboard.html", {"year": year})

# ==============================================================
# POR CATEGORÍA
# ==============================================================

@login_required
def inventario_por_categoria(request):
    year = datetime.datetime.now().year

    if request.method == "POST":
        form = CategoryUpdateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoría actualizada con éxito.")
            return redirect("inventario:inventario_por_categoria")
        messages.error(request, "Error al actualizar la categoría.")
    else:
        form = CategoryUpdateForm()

    from django.db.models import Q as _Q
    mapping = [
        (Computadora, "Computadora", "modelo",  "computadora"),
        (Televisor,   "Televisor",   "modelo",  "televisor"),
        (Impresora,   "Impresora",   "nombre",  "impresora"),
        (Router,      "Router",      "modelo",  "router"),
        (DataShow,    "DataShow",    "serie",   "datashow"),
        (Monitor,     "Monitor",     "modelo",  "monitor"),
    ]

    grupos = []
    sin_cat = []
    for Model, label, field, tipo_key in mapping:
        qs = list(Model.objects.annotate(
            tipo=Value(label, output_field=CharField()),
            tipo_key=Value(tipo_key, output_field=CharField()),
            descripcion=F(field),
            categoria=F("category"),
        ).values("tipo", "tipo_key", "id", "descripcion", "categoria"))
        categorized = [i for i in qs if i["categoria"]]
        uncategorized = [i for i in qs if not i["categoria"]]
        if categorized:
            grupos.append({"tipo": label, "tipo_key": tipo_key, "items": categorized})
        sin_cat.extend(uncategorized)

    return render(request, "inventario/inventario_por_categoria.html", {
        "grupos":   grupos,
        "sin_cat":  sin_cat,
        "form":     form,
        "year":     year,
        "category_choices": InventoryItem.CATEGORY_CHOICES,
    })

# ==============================================================
# INVENTARIOS (CREATE + LIST)
# ==============================================================

_PREFIJOS_COMP = {
    'estandar': 'IDANACOMP',
    'lab_bl':   'IDANALABBL',
    'lab_col':  'IDANALABCOL',
    'informatica': 'IDCFPLAB',
}

def _siguiente_asset_id_computadora(prefijo_key='estandar'):
    prefix = _PREFIJOS_COMP.get(prefijo_key, 'IDANACOMP')
    ultimo = Computadora.objects.filter(
        asset_id__startswith=prefix
    ).order_by('-asset_id').values_list('asset_id', flat=True).first()
    if ultimo:
        try:
            n = int(ultimo[len(prefix):])
        except ValueError:
            n = 0
    else:
        n = 0
    return f"{prefix}{n+1:03d}"


@login_required
def inventario_computadoras(request):
    if request.method == "POST":
        form = ComputadoraForm(request.POST)
        if form.is_valid():
            comp = form.save(commit=False)
            prefijo_key = request.POST.get('id_prefix', 'estandar')
            if prefijo_key not in _PREFIJOS_COMP:
                prefijo_key = 'estandar'
            comp.asset_id = _siguiente_asset_id_computadora(prefijo_key)
            ip_post = request.POST.get('ip', '').strip()
            comp.ip = ip_post if ip_post and ip_post != '0.0.0.0' else '0.0.0.0'
            serie_post = request.POST.get('serie', '').strip()
            comp.serie = serie_post if serie_post and serie_post != '—' else ''
            if prefijo_key == 'estandar':
                comp.edificio = request.POST.get('edificio', '').strip() or None
            else:
                comp.edificio = None
            grado_subtipo = request.POST.get('grado_subtipo', '').strip()
            if comp.grado == 'Otros' and grado_subtipo:
                comp.grado = f'Otros/{grado_subtipo}'
            comp.save()
            return redirect("inventario:inventario_computadoras")
    else:
        form = ComputadoraForm()

    total_comp = Computadora.objects.count()
    prefijos_tabs = [
        {"key": "IDANACOMP",   "label": "General",   "count": Computadora.objects.filter(asset_id__startswith="IDANACOMP").count()},
        {"key": "IDANALABBL",  "label": "Lab BL",    "count": Computadora.objects.filter(asset_id__startswith="IDANALABBL").count()},
        {"key": "IDANALABCOL", "label": "Lab COL",   "count": Computadora.objects.filter(asset_id__startswith="IDANALABCOL").count()},
        {"key": "IDCFPLAB",    "label": "Lab CFP",   "count": Computadora.objects.filter(asset_id__startswith="IDCFPLAB").count()},
    ]

    return render(request, "inventario/inventario_computadoras.html", {
        "form":              form,
        "next_id":           _siguiente_asset_id_computadora('estandar'),
        "next_id_bl":        _siguiente_asset_id_computadora('lab_bl'),
        "next_id_col":       _siguiente_asset_id_computadora('lab_col'),
        "next_id_inf":       _siguiente_asset_id_computadora('informatica'),
        "computadoras":      Computadora.objects.order_by("-id"),
        "edificios":         EdificioComputadora.objects.all(),
        "subtipos_otros":    SubtipoGradoComputadora.objects.all(),
        "total_comp":        total_comp,
        "prefijos_tabs":     prefijos_tabs,
    })


@login_required
def computadoras_list(request):
    year = datetime.datetime.now().year
    form = ComputadoraFilterForm(request.GET or None)
    qs = Computadora.objects.order_by("-fecha_instalado")

    if form.is_valid():
        cd = form.cleaned_data
        for field, val in cd.items():
            if val:
                qs = qs.filter(**{f"{field}__icontains": val})

    return render(request, "inventario/filtro_computadoras.html", {
        "form": form,
        "computadoras": qs,
        "year": year
    })


def _siguiente_asset_id_televisor():
    prefix = 'IDANATV'
    ultimo = Televisor.objects.filter(
        asset_id__startswith=prefix
    ).order_by('-asset_id').values_list('asset_id', flat=True).first()
    if ultimo:
        try:
            n = int(ultimo[len(prefix):])
        except ValueError:
            n = 0
    else:
        n = 0
    return f"{prefix}{n+1:03d}"


@login_required
def inventario_televisores(request):
    if request.method == "POST":
        form = TelevisorForm(request.POST)
        if form.is_valid():
            tv = form.save(commit=False)
            tv.asset_id = _siguiente_asset_id_televisor()
            tv.serie    = 'Pendiente'
            tv.ip       = '0.0.0.0'
            tv.save()
            return redirect("inventario:inventario_televisores")
    else:
        form = TelevisorForm()

    return render(request, "inventario/inventario_televisores.html", {
        "form":       form,
        "next_id":    _siguiente_asset_id_televisor(),
        "televisores": Televisor.objects.order_by("-id"),
    })


# <--- hecho por claude code: Asset ID autoincremental de impresoras (ANAIMP-001), mirror de _siguiente_asset_id_televisor
def _siguiente_asset_id_impresora():
    prefix = 'ANAIMP-'
    ultimo = Impresora.objects.filter(
        asset_id__startswith=prefix
    ).order_by('-asset_id').values_list('asset_id', flat=True).first()
    if ultimo:
        try:
            n = int(ultimo[len(prefix):])
        except ValueError:
            n = 0
    else:
        n = 0
    return f"{prefix}{n+1:03d}"


@login_required
def inventario_impresoras(request):
    year = datetime.datetime.now().year

    if request.method == "POST":
        form = ImpresoraForm(request.POST)
        if form.is_valid():
            imp = form.save(commit=False)
            imp.asset_id = _siguiente_asset_id_impresora()  # <--- hecho por claude code: ID auto en el servidor
            imp.save()
            return redirect("inventario:inventario_impresoras")
    else:
        form = ImpresoraForm()

    return render(request, "inventario/inventario_impresoras.html", {
        "form": form,
        "year": year,
        "next_id": _siguiente_asset_id_impresora(),  # <--- hecho por claude code: mostrar el próximo ID
        "impresoras": Impresora.objects.order_by("-id")
    })


@login_required
def inventario_routers(request):
    year = datetime.datetime.now().year

    if request.method == "POST":
        form = RouterForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("inventario:inventario_routers")
    else:
        form = RouterForm()

    return render(request, "inventario/inventario_routers.html", {
        "form": form,
        "year": year,
        "routers": Router.objects.order_by("-id")
    })


@login_required
def inventario_datashows(request):
    year = datetime.datetime.now().year

    if request.method == "POST":
        form = DataShowForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("inventario:inventario_datashows")
    else:
        form = DataShowForm()

    return render(request, "inventario/inventario_datashows.html", {
        "form": form,
        "year": year,
        "datashows": DataShow.objects.order_by("-id")
    })


@login_required
def inventario_monitores(request):
    import re as _re
    year = datetime.datetime.now().year

    if request.method == "POST":
        form = MonitorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("inventario:inventario_monitores")
    else:
        form = MonitorForm()

    # Prefijos de ID y siguiente número disponible
    prefijos = [
        ("ANAMONI",    "General"),
        ("LABBLMONI",  "Lab BL"),
        ("LABCOLMONI", "Lab COL"),
        ("CFPINFO",    "Lab Informática"),
    ]
    next_ids = {}
    for prefix, _ in prefijos:
        existing = Monitor.objects.filter(asset_id__startswith=prefix).values_list('asset_id', flat=True)
        nums = [int(m.group()) for aid in existing for m in [_re.search(r'\d+$', aid)] if m]
        next_ids[prefix] = f"{prefix}{(max(nums) + 1 if nums else 1):03d}"

    # Modelos de monitores desde catálogo admin
    modelos_monitor = list(ModeloMonitor.objects.values_list('nombre', flat=True))

    # Computadoras para selects
    computadoras = list(Computadora.objects.order_by('asset_id').values('id', 'asset_id', 'modelo', 'serie', 'asignado_a', 'area', 'grado'))

    # Asset IDs de computadoras ya vinculadas a un monitor (para advertencia de duplicado)
    labs_usados    = set(Monitor.objects.exclude(laboratorio='').exclude(laboratorio__isnull=True).values_list('laboratorio', flat=True))
    asignados_usados = set(Monitor.objects.exclude(asignado_a='').exclude(asignado_a__isnull=True).values_list('asignado_a', flat=True))

    return render(request, "inventario/inventario_monitores.html", {
        "form":              form,
        "year":              year,
        "monitores":         Monitor.objects.order_by("-id"),
        "prefijos":          prefijos,
        "next_ids":          next_ids,
        "modelos_monitor":   modelos_monitor,
        "computadoras":      computadoras,
        "labs_usados":       list(labs_usados),
        "asignados_usados":  list(asignados_usados),
    })


# ==============================================================
# INVENTARIO REGISTROS (TABS CONSOLIDADOS)
# ==============================================================

@login_required
def inventario_registros(request):
    year = datetime.datetime.now().year

    prefijos_comp = [
        {"key": "IDANACOMP",   "label": "General",  "count": Computadora.objects.filter(asset_id__startswith="IDANACOMP").count()},
        {"key": "IDANALABBL",  "label": "Lab BL",   "count": Computadora.objects.filter(asset_id__startswith="IDANALABBL").count()},
        {"key": "IDANALABCOL", "label": "Lab COL",  "count": Computadora.objects.filter(asset_id__startswith="IDANALABCOL").count()},
        {"key": "IDCFPLAB",    "label": "Lab CFP",  "count": Computadora.objects.filter(asset_id__startswith="IDCFPLAB").count()},
    ]

    return render(request, "inventario/inventario_registros.html", {
        "computadoras":  Computadora.objects.order_by("id"),
        "impresoras":    Impresora.objects.order_by("id"),
        "televisores":   Televisor.objects.order_by("id"),
        "routers":       Router.objects.order_by("id"),
        "datashows":     DataShow.objects.order_by("id"),
        "monitores":     Monitor.objects.order_by("id"),
        "year":          year,
        "prefijos_comp": prefijos_comp,
        "total_comp":    Computadora.objects.count(),
    })

# ==============================================================
# Asignar / cambiar grupo de computadora(s)
# ==============================================================

@login_required
def asignar_grupo_computadora(request, pk):
    comp = get_object_or_404(Computadora, pk=pk)
    if request.method == 'POST':
        valor = request.POST.get('grupo', '').strip()
        comp.grupo = int(valor) if valor.isdigit() else None
        comp.save(update_fields=['grupo'])
        return JsonResponse({'ok': True, 'grupo': comp.grupo})
    return JsonResponse({'ok': True, 'grupo': comp.grupo})


@login_required
def asignar_grupo_bulk(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    ids   = request.POST.getlist('ids')
    valor = request.POST.get('grupo', '').strip()
    grupo = int(valor) if valor.isdigit() else None
    updated = Computadora.objects.filter(pk__in=ids).update(grupo=grupo)
    return JsonResponse({'ok': True, 'grupo': grupo, 'count': updated})


@login_required
def asignar_categoria_televisor(request, pk):
    from .models import Televisor
    tv = get_object_or_404(Televisor, pk=pk)
    if request.method == 'POST':
        valor = request.POST.get('categoria', '').strip()
        tv.category = valor if valor else None
        tv.save(update_fields=['category'])
        return JsonResponse({'ok': True, 'categoria': tv.category or ''})
    return JsonResponse({'ok': True, 'categoria': tv.category or ''})


@login_required
def asignar_categoria_router(request, pk):
    from .models import Router
    rt = get_object_or_404(Router, pk=pk)
    if request.method == 'POST':
        valor = request.POST.get('categoria', '').strip()
        rt.category = valor if valor else None
        rt.save(update_fields=['category'])
        return JsonResponse({'ok': True, 'categoria': rt.category or ''})
    return JsonResponse({'ok': True, 'categoria': rt.category or ''})


@login_required
def asignar_categoria_impresora(request, pk):
    from .models import Impresora
    obj = get_object_or_404(Impresora, pk=pk)
    if request.method == 'POST':
        valor = request.POST.get('categoria', '').strip()
        obj.category = valor if valor else None
        obj.save(update_fields=['category'])
        return JsonResponse({'ok': True, 'categoria': obj.category or ''})
    return JsonResponse({'ok': True, 'categoria': obj.category or ''})


@login_required
def asignar_categoria_datashow(request, pk):
    from .models import DataShow
    obj = get_object_or_404(DataShow, pk=pk)
    if request.method == 'POST':
        valor = request.POST.get('categoria', '').strip()
        obj.category = valor if valor else None
        obj.save(update_fields=['category'])
        return JsonResponse({'ok': True, 'categoria': obj.category or ''})
    return JsonResponse({'ok': True, 'categoria': obj.category or ''})


@login_required
def asignar_categoria_monitor(request, pk):
    from .models import Monitor
    obj = get_object_or_404(Monitor, pk=pk)
    if request.method == 'POST':
        valor = request.POST.get('categoria', '').strip()
        obj.category = valor if valor else None
        obj.save(update_fields=['category'])
        return JsonResponse({'ok': True, 'categoria': obj.category or ''})
    return JsonResponse({'ok': True, 'categoria': obj.category or ''})


@login_required
def asignar_categoria_bulk(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    tipo = request.POST.get('tipo', '').strip()
    categoria = request.POST.get('categoria', '').strip()
    ids = request.POST.getlist('ids')
    if not ids:
        return JsonResponse({'ok': False, 'error': 'Sin IDs'})
    mapa = {
        'computadora': Computadora,
        'televisor':   Televisor,
        'router':      Router,
        'impresora':   Impresora,
        'datashow':    DataShow,
        'monitor':     Monitor,
    }
    modelo = mapa.get(tipo)
    if not modelo:
        return JsonResponse({'ok': False, 'error': 'Tipo inválido'})
    modelo.objects.filter(pk__in=ids).update(category=categoria or None)
    return JsonResponse({'ok': True, 'categoria': categoria})


# ==============================================================
# GET (Cargar formulario en el modal)
# ==============================================================

@login_required
def get_computadora(request, pk):
    obj = get_object_or_404(Computadora, pk=pk)
    form = ComputadoraForm(instance=obj)
    grado_val = obj.grado or ''
    grado_subtipo = grado_val.split('/', 1)[1] if grado_val.startswith('Otros/') else ''
    return render(request, "inventario/edit_computadora.html", {
        "form": form, "obj": obj,
        "edificios": EdificioComputadora.objects.all(),
        "grado_subtipo": grado_subtipo,
        "subtipos_otros": SubtipoGradoComputadora.objects.all(),
    })

@login_required
def computadora_json(request, pk):
    obj = get_object_or_404(Computadora, pk=pk)
    return JsonResponse({
        'asset_id':   obj.asset_id,
        'modelo':     obj.modelo,
        'serie':      obj.serie,
        'asignado_a': obj.asignado_a,
        'area':       obj.area,
        'grado':      obj.grado,
    })

@login_required
def get_televisor(request, pk):
    obj = get_object_or_404(Televisor, pk=pk)
    form = TelevisorForm(instance=obj)
    return render(request, "inventario/edit_televisor.html", {"form": form, "obj": obj})

@login_required
def get_impresora(request, pk):
    obj = get_object_or_404(Impresora, pk=pk)
    form = ImpresoraForm(instance=obj)
    return render(request, "inventario/edit_impresora.html", {"form": form, "obj": obj})

@login_required
def get_router(request, pk):
    obj = get_object_or_404(Router, pk=pk)
    form = RouterForm(instance=obj)
    return render(request, "inventario/edit_router.html", {"form": form, "obj": obj})

@login_required
def get_datashow(request, pk):
    obj = get_object_or_404(DataShow, pk=pk)
    form = DataShowForm(instance=obj)
    return render(request, "inventario/edit_datashow.html", {"form": form, "obj": obj})

@login_required
def get_monitor(request, pk):
    obj  = get_object_or_404(Monitor, pk=pk)
    form = MonitorForm(instance=obj)
    prefijos = [
        ("ANAMONI",    "General"),
        ("LABBLMONI",  "Lab BL"),
        ("LABCOLMONI", "Lab COL"),
        ("CFPINFO",    "Lab Informática"),
    ]
    computadoras = list(Computadora.objects.order_by('asset_id').values('id', 'asset_id', 'modelo', 'serie', 'asignado_a', 'area', 'grado'))
    labs_usados      = list(Monitor.objects.exclude(pk=pk).exclude(laboratorio='').exclude(laboratorio__isnull=True).values_list('laboratorio', flat=True))
    asignados_usados = list(Monitor.objects.exclude(pk=pk).exclude(asignado_a='').exclude(asignado_a__isnull=True).values_list('asignado_a', flat=True))
    return render(request, "inventario/edit_monitor.html", {
        "form": form, "obj": obj, "prefijos": prefijos,
        "computadoras": computadoras,
        "labs_usados": labs_usados,
        "asignados_usados": asignados_usados,
    })

# ==============================================================
# UPDATE (Guardar cambios vía AJAX)
# ==============================================================

@login_required
def update_computadora(request, pk):
    obj = get_object_or_404(Computadora, pk=pk)
    form = ComputadoraForm(request.POST, instance=obj)
    if form.is_valid():
        comp = form.save(commit=False)
        new_ip = request.POST.get('ip', '').strip()
        if new_ip:
            comp.ip = new_ip
        new_serie = request.POST.get('serie', '').strip()
        if new_serie:
            comp.serie = new_serie
        new_asset_id = request.POST.get('asset_id', '').strip()
        if new_asset_id:
            comp.asset_id = new_asset_id
        comp.edificio = request.POST.get('edificio', '').strip() or None
        grado_subtipo = request.POST.get('grado_subtipo', '').strip()
        if comp.grado == 'Otros' and grado_subtipo:
            comp.grado = f'Otros/{grado_subtipo}'
        comp.save()
        return JsonResponse({"ok": True})
    return JsonResponse({"ok": False, "errors": form.errors})


@login_required
def update_televisor(request, pk):
    obj = get_object_or_404(Televisor, pk=pk)
    form = TelevisorForm(request.POST, instance=obj)
    if form.is_valid():
        form.save()
        return JsonResponse({"ok": True})
    return JsonResponse({"ok": False, "errors": form.errors})


@login_required
def update_impresora(request, pk):
    obj = get_object_or_404(Impresora, pk=pk)
    form = ImpresoraForm(request.POST, instance=obj)
    if form.is_valid():
        form.save()
        return JsonResponse({"ok": True})
    return JsonResponse({"ok": False, "errors": form.errors})


@login_required
def update_router(request, pk):
    obj = get_object_or_404(Router, pk=pk)
    form = RouterForm(request.POST, instance=obj)
    if form.is_valid():
        form.save()
        return JsonResponse({"ok": True})
    return JsonResponse({"ok": False, "errors": form.errors})


@login_required
def update_datashow(request, pk):
    obj = get_object_or_404(DataShow, pk=pk)
    form = DataShowForm(request.POST, instance=obj)
    if form.is_valid():
        form.save()
        return JsonResponse({"ok": True})
    return JsonResponse({"ok": False, "errors": form.errors})


@login_required
def update_monitor(request, pk):
    obj = get_object_or_404(Monitor, pk=pk)
    form = MonitorForm(request.POST, instance=obj)
    if form.is_valid():
        form.save()
        return JsonResponse({"ok": True})
    return JsonResponse({"ok": False, "errors": form.errors})

# ==============================================================
# DELETE (JSON)
# ==============================================================

@login_required
def eliminar_computadora(request, pk):
    get_object_or_404(Computadora, pk=pk).delete()
    return JsonResponse({"ok": True})

@login_required
def eliminar_televisor(request, pk):
    get_object_or_404(Televisor, pk=pk).delete()
    return JsonResponse({"ok": True})

@login_required
def eliminar_impresora(request, pk):
    get_object_or_404(Impresora, pk=pk).delete()
    return JsonResponse({"ok": True})

@login_required
def eliminar_router(request, pk):
    get_object_or_404(Router, pk=pk).delete()
    return JsonResponse({"ok": True})

@login_required
def eliminar_datashow(request, pk):
    get_object_or_404(DataShow, pk=pk).delete()
    return JsonResponse({"ok": True})

@login_required
def eliminar_monitor(request, pk):
    get_object_or_404(Monitor, pk=pk).delete()
    return JsonResponse({"ok": True})


# ==============================================================
# PDF GENERATOR
# ==============================================================

def download_model_pdf(request, tipo, pk):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.utils import ImageReader

    model_map = {
        "computadora": Computadora,
        "televisor":   Televisor,
        "impresora":   Impresora,
        "router":      Router,
        "datashow":    DataShow,
        "monitor":     Monitor,
    }

    fields_map = {
        "computadora": [
            ("ID",              "asset_id"),
            ("Modelo",          "modelo"),
            ("Serie",           "serie"),
            ("IP",              "ip"),
            ("Categoría",       "category"),
            ("Asignado a",      "asignado_a"),
            ("Área",            "area"),
            ("Grado",           "grado"),
            ("Edificio",        "edificio"),
            ("Fecha Instalación","fecha_instalado"),
            ("Observaciones",   "observaciones"),
        ],
        "televisor": [
            ("ID",              "asset_id"),
            ("Modelo",          "modelo"),
            ("Serie",           "serie"),
            ("IP",              "ip"),
            ("Categoría",       "category"),
            ("Grado",           "grado"),
            ("Área",            "area"),
            ("Observaciones",   "observaciones"),
        ],
        "impresora": [
            ("ID",              "asset_id"),
            ("Nombre",          "nombre"),
            ("Modelo",          "modelo"),
            ("Serie",           "serie"),
            ("Categoría",       "category"),
            ("Asignado a",      "asignado_a"),
            ("Nivel Tinta",     "nivel_tinta"),
            ("Últ. Llenado",    "ultima_vez_llenado"),
            ("Impresiones",     "cantidad_impresiones"),
            ("A Color",         "a_color"),
            ("Observaciones",   "observaciones"),
        ],
        "router": [
            ("ID",              "asset_id"),
            ("Modelo",          "modelo"),
            ("Serie",           "serie"),
            ("Categoría",       "category"),
            ("Nombre Router",   "nombre_router"),
            ("Clave Router",    "clave_router"),
            ("IP Asignada",     "ip_asignada"),
            ("IP de Uso",       "ip_uso"),
            ("Ubicado",         "ubicado"),
            ("Observaciones",   "observaciones"),
        ],
        "datashow": [
            ("ID",              "asset_id"),
            ("Nombre",          "nombre"),
            ("Modelo",          "modelo"),
            ("Serie",           "serie"),
            ("Categoría",       "category"),
            ("Estado",          "estado"),
            ("Cable Corriente", "cable_corriente"),
            ("HDMI",            "hdmi"),
            ("VGA",             "vga"),
            ("Extensión",       "extension"),
            ("Observaciones",   "observaciones"),
        ],
        "monitor": [
            ("ID",              "asset_id"),
            ("Modelo",          "modelo"),
            ("Serie",           "serie"),
            ("Tipo Ubicación",  "ubicacion_tipo"),
            ("Laboratorio",     "laboratorio"),
            ("Asignado a",      "asignado_a"),
            ("Categoría",       "category"),
            ("Observaciones",   "observaciones"),
        ],
    }

    # Color de acento por tipo de equipo
    color_map = {
        "computadora": "#1971c2",
        "televisor":   "#0c8599",
        "impresora":   "#d9480f",
        "router":      "#2f9e44",
        "datashow":    "#7048e8",
        "monitor":     "#495057",
    }

    tipo = tipo.lower()
    if tipo not in model_map:
        return HttpResponse("Modelo inválido", status=404)

    Model  = model_map[tipo]
    campos = fields_map[tipo]
    obj    = get_object_or_404(Model, pk=pk)

    accent_hex = color_map.get(tipo, "#0056b3")
    accent     = colors.HexColor(accent_hex)
    grey_line  = colors.HexColor("#dee2e6")
    label_bg   = colors.HexColor("#f1f3f5")
    label_fg   = colors.HexColor("#343a40")
    alt_row    = colors.HexColor("#f8f9fa")

    # QR code apuntando a este mismo PDF
    pdf_path = reverse("inventario:download_model_pdf", args=[tipo, pk])
    pdf_url  = f"https://servicios.ana-hn.org:437{pdf_path}"
    qr_pil   = qrcode.make(pdf_url).convert("RGB")
    qr_buf   = io.BytesIO()
    qr_pil.save(qr_buf, format="PNG")
    qr_buf.seek(0)

    # Canvas (portrait Letter: 612 x 792 pts)
    page_w, page_h = letter
    buffer = io.BytesIO()
    pdf    = canvas.Canvas(buffer, pagesize=letter)
    margin = 1.5 * cm

    today_str = datetime.date.today().strftime("%d/%m/%Y")
    asset_id  = getattr(obj, "asset_id", "")
    titulo    = f"Ficha de {tipo.capitalize()}"

    pdf.setTitle(f"Ficha {tipo.capitalize()} – {asset_id}")
    pdf.setAuthor("TechCare – Instituto ANA Honduras")
    pdf.setSubject(f"Reporte de inventario: {asset_id}")

    # ── CABECERA ──────────────────────────────────────────────
    band_h = 105

    # Banda principal de color
    pdf.setFillColor(accent)
    pdf.rect(0, page_h - band_h, page_w, band_h, fill=1, stroke=0)

    # Banda inferior oscurecida (efecto sombra)
    pdf.setFillColor(colors.HexColor("#00000022"))
    pdf.rect(0, page_h - band_h, page_w, 6, fill=1, stroke=0)

    # Texto blanco
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 8.5)
    pdf.drawString(margin, page_h - 16, "INSTITUTO EDUCATIVO ANA HONDURAS · TechCare")

    pdf.setFont("Helvetica-Bold", 26)
    pdf.drawString(margin, page_h - 50, titulo)

    # Pastilla Asset ID
    badge_x, badge_y, badge_w, badge_h = margin, page_h - 85, 160, 20
    pdf.setFillColor(colors.HexColor("#ffffff33"))
    pdf.roundRect(badge_x, badge_y, badge_w, badge_h, 4, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(badge_x + 8, badge_y + 5, f"Asset ID: {asset_id}")

    # Fecha en cabecera
    pdf.setFont("Helvetica", 8.5)
    pdf.drawRightString(page_w - 100, page_h - 80, today_str)

    # QR code incrustado en la cabecera
    qr_size = 88
    qr_x = page_w - qr_size - 10
    qr_y = page_h - band_h + 8
    pdf.drawImage(ImageReader(qr_buf), qr_x, qr_y,
                  width=qr_size, height=qr_size, preserveAspectRatio=True)

    # Línea separadora debajo de la cabecera
    sep_y = page_h - band_h - 8
    pdf.setStrokeColor(accent)
    pdf.setLineWidth(1.5)
    pdf.line(margin, sep_y, page_w - margin, sep_y)

    # ── TABLA ─────────────────────────────────────────────────
    data = [["Campo", "Valor"]]
    for label, attr in campos:
        val = getattr(obj, attr, None)
        if val is None:
            val = "—"
        elif isinstance(val, bool):
            val = "Sí" if val else "No"
        elif isinstance(val, datetime.date):
            val = val.strftime("%d/%m/%Y")
        else:
            # Usar display label si el campo tiene choices
            display_method = f"get_{attr}_display"
            if hasattr(obj, display_method):
                val = getattr(obj, display_method)() or "—"
            else:
                val = str(val) if str(val) else "—"
        data.append([label, val])

    col1_w = 5.2 * cm
    col2_w = page_w - 2 * margin - col1_w
    n = len(data)

    tstyles = [
        # Fila de encabezado
        ("BACKGROUND",    (0, 0), (-1, 0), accent),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 10),
        ("TOPPADDING",    (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 9),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        # Todas las celdas
        ("FONTSIZE",      (0, 1), (-1, -1), 10),
        ("TOPPADDING",    (0, 1), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        # Columna de etiquetas — fondo gris, negrita
        ("BACKGROUND",    (0, 1), (0, -1), label_bg),
        ("FONTNAME",      (0, 1), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR",     (0, 1), (0, -1), label_fg),
        # Columna de valores
        ("FONTNAME",      (1, 1), (1, -1), "Helvetica"),
        ("TEXTCOLOR",     (1, 1), (1, -1), colors.HexColor("#212529")),
        # Bordes
        ("LINEBELOW",     (0, 0), (-1, -2), 0.4, grey_line),
        ("BOX",           (0, 0), (-1, -1), 0.6, grey_line),
        ("LINEAFTER",     (0, 0), (0, -1),  0.4, grey_line),
    ]

    # Filas alternas en columna de valor
    for i in range(1, n):
        if i % 2 == 0:
            tstyles.append(("BACKGROUND", (1, i), (1, i), alt_row))

    table = Table(data, colWidths=[col1_w, col2_w])
    table.setStyle(TableStyle(tstyles))

    tw, th = table.wrap(col1_w + col2_w, 0)
    table_y = sep_y - 14 - th
    table.drawOn(pdf, margin, table_y)

    current_y = table_y  # rastrea la Y más baja usada

    # ── SECCIÓN MONITOR (solo para computadoras) ───────────────
    if tipo == "computadora":
        monitor = Monitor.objects.filter(laboratorio=obj.asset_id).first()
        if monitor:
            mon_accent = colors.HexColor("#495057")
            gap = 18
            subband_h = 26

            # Sub-cabecera del monitor
            sub_y = current_y - gap
            pdf.setFillColor(mon_accent)
            pdf.rect(margin, sub_y - subband_h, col1_w + col2_w, subband_h, fill=1, stroke=0)
            pdf.setFillColor(colors.white)
            pdf.setFont("Helvetica-Bold", 10)
            pdf.drawString(margin + 8, sub_y - subband_h + 8,
                           f"Monitor Asignado  —  {monitor.asset_id}")

            # Tabla del monitor
            mon_campos = [
                ("Asset ID",       monitor.asset_id or "—"),
                ("Modelo",         monitor.modelo or "—"),
                ("Serie",          monitor.serie or "—"),
                ("Tipo Ubicación", monitor.get_ubicacion_tipo_display() if monitor.ubicacion_tipo else "—"),
                ("Categoría",      monitor.category or "—"),
                ("Observaciones",  monitor.observaciones or "—"),
            ]
            mon_data = [["Campo", "Valor"]] + [[lbl, val] for lbl, val in mon_campos]
            nm = len(mon_data)

            mon_tstyles = [
                ("BACKGROUND",    (0, 0), (-1, 0), mon_accent),
                ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
                ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, 0), 10),
                ("TOPPADDING",    (0, 0), (-1, 0), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 9),
                ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
                ("FONTSIZE",      (0, 1), (-1, -1), 10),
                ("TOPPADDING",    (0, 1), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
                ("LEFTPADDING",   (0, 0), (-1, -1), 10),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND",    (0, 1), (0, -1), label_bg),
                ("FONTNAME",      (0, 1), (0, -1), "Helvetica-Bold"),
                ("TEXTCOLOR",     (0, 1), (0, -1), label_fg),
                ("FONTNAME",      (1, 1), (1, -1), "Helvetica"),
                ("TEXTCOLOR",     (1, 1), (1, -1), colors.HexColor("#212529")),
                ("LINEBELOW",     (0, 0), (-1, -2), 0.4, grey_line),
                ("BOX",           (0, 0), (-1, -1), 0.6, grey_line),
                ("LINEAFTER",     (0, 0), (0, -1),  0.4, grey_line),
            ]
            for i in range(1, nm):
                if i % 2 == 0:
                    mon_tstyles.append(("BACKGROUND", (1, i), (1, i), alt_row))

            mon_table = Table(mon_data, colWidths=[col1_w, col2_w])
            mon_table.setStyle(TableStyle(mon_tstyles))
            _, mon_th = mon_table.wrap(col1_w + col2_w, 0)
            mon_table_y = sub_y - subband_h - mon_th
            mon_table.drawOn(pdf, margin, mon_table_y)
            current_y = mon_table_y

    # ── PIE DE PÁGINA ─────────────────────────────────────────
    pdf.setFillColor(accent)
    pdf.rect(0, 0, page_w, 5, fill=1, stroke=0)

    pdf.setFillColor(colors.HexColor("#6c757d"))
    pdf.setFont("Helvetica", 7.5)
    pdf.drawString(margin, 13, "TechCare — Sistema de Gestión de Inventario · Instituto ANA Honduras")
    pdf.drawRightString(page_w - margin, 13, f"Generado el {today_str}")

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    filename = f"Ficha_{tipo.capitalize()}_{asset_id}.pdf"
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


# ===================== EXPORTAR EXCEL =====================
@login_required
def exportar_excel_inventario(request, categoria):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    def estilizar_cabecera(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        ws.row_dimensions[1].height = 22

    if categoria == 'computadoras':
        ws.title = "Computadoras"
        estilizar_cabecera(ws, ["Asset ID", "Modelo", "Serie", "IP", "Asignado a", "Área", "Grado", "Grupo", "Fecha Instalación", "Observaciones"])
        for c in Computadora.objects.all().order_by('asset_id'):
            ws.append([c.asset_id, c.modelo, c.serie, c.ip, c.asignado_a, c.area, c.grado,
                        c.grupo or '', str(c.fecha_instalado) if c.fecha_instalado else '', c.observaciones or ''])

    elif categoria == 'impresoras':
        ws.title = "Impresoras"
        estilizar_cabecera(ws, ["Asset ID", "Nombre", "Modelo", "Serie", "Asignado a", "Nivel Tinta", "Última vez llenado", "Cant. Impresiones", "A Color", "Observaciones"])
        for i in Impresora.objects.all().order_by('asset_id'):
            ws.append([i.asset_id, i.nombre, i.modelo, i.serie, i.asignado_a, i.nivel_tinta,
                        str(i.ultima_vez_llenado) if i.ultima_vez_llenado else '', i.cantidad_impresiones,
                        'Sí' if i.a_color else 'No', i.observaciones or ''])

    elif categoria == 'televisores':
        ws.title = "Televisores"
        estilizar_cabecera(ws, ["Asset ID", "Modelo", "Serie", "IP", "Grado", "Área", "Observaciones"])
        for t in Televisor.objects.all().order_by('asset_id'):
            ws.append([t.asset_id, t.modelo, t.serie, t.ip, t.grado, t.area, t.observaciones or ''])

    elif categoria == 'routers':
        ws.title = "Routers"
        estilizar_cabecera(ws, ["Asset ID", "Modelo", "Serie", "Nombre Router", "IP Asignada", "IP Uso", "Ubicado", "Observaciones"])
        for r in Router.objects.all().order_by('asset_id'):
            ws.append([r.asset_id, r.modelo, r.serie, r.nombre_router, r.ip_asignada, r.ip_uso, r.ubicado, r.observaciones or ''])

    elif categoria == 'datashows':
        ws.title = "DataShows"
        estilizar_cabecera(ws, ["Asset ID", "Nombre", "Modelo", "Serie", "Estado", "Cable Corriente", "HDMI", "VGA", "Extensión", "Observaciones"])
        for d in DataShow.objects.all().order_by('asset_id'):
            ws.append([d.asset_id, d.nombre, d.modelo, d.serie, d.estado,
                        'Sí' if d.cable_corriente else 'No', 'Sí' if d.hdmi else 'No',
                        'Sí' if d.vga else 'No', 'Sí' if d.extension else 'No', d.observaciones or ''])

    elif categoria == 'monitores':
        ws.title = "Monitores"
        estilizar_cabecera(ws, ["Asset ID", "Modelo", "Serie", "Tipo Ubicación", "Laboratorio", "Asignado a", "Observaciones"])
        for m in Monitor.objects.all().order_by('asset_id'):
            ws.append([m.asset_id, m.modelo, m.serie,
                        m.get_ubicacion_tipo_display() if m.ubicacion_tipo else '',
                        m.laboratorio or '', m.asignado_a or '', m.observaciones or ''])
    else:
        return HttpResponse("Categoría no válida", status=400)

    # Ajustar anchos de columna automáticamente
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inventario_{categoria}.xlsx"'
    return response
