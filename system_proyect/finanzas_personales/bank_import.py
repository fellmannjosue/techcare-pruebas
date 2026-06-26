"""
Lector de estados de cuenta bancarios (BAC y formato tabular genérico).

Soporta:
  - PDF   (pdfplumber)  → estado "Transacciones del mes" de BAC Credomatic.
  - Excel (.xlsx/.xls)  → exportación tabular del banco.
  - CSV                 → exportación tabular del banco.

Devuelve una lista de movimientos normalizados:
  {
    'date'       : 'YYYY-MM-DD',
    'description': str,
    'amount'     : float (positivo),
    'type'       : 'income' | 'expense',
    'ref'        : str,        # referencia del banco (informativa)
    'code'       : str,        # código de movimiento (CP/TF/CR/D4/HT…)
    'balance'    : float|None, # saldo en libros de esa línea (informativo)
  }
"""
import io
import re
import csv
import unicodedata
from datetime import datetime, date


# ── utilidades ───────────────────────────────────────────────────────────────

def _strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFKD', str(s))
                   if not unicodedata.combining(c))


def _norm(s):
    return _strip_accents(s).lower().strip()


def _num(v):
    """Convierte '2,432.26' / 53.57 / '' en float."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '')
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_date(v):
    """Devuelve 'YYYY-MM-DD' o None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _collapse(s):
    return re.sub(r'\s+', ' ', str(s or '')).strip()[:255]


def _mk_mov(d, desc, deb, cred, ref='', code='', balance=None):
    """Arma un movimiento o devuelve None si no aplica."""
    if not d:
        return None
    deb, cred = _num(deb), _num(cred)
    if cred > 0 and deb == 0:
        tipo, amt = 'income', cred
    elif deb > 0:
        tipo, amt = 'expense', deb
    else:
        return None
    return {
        'date': d, 'description': _collapse(desc), 'amount': round(amt, 2),
        'type': tipo, 'ref': _collapse(ref)[:60], 'code': _collapse(code)[:10],
        'balance': balance,
    }


# ── PDF (BAC) ──────────────────────────────────────────────────────────────────
# Línea tipo: "01/06/2026 2560 HT HRE PLUS BASICO 53.57 0.00 2,432.26"
#   fecha  referencia  código(2)  descripción  débito  crédito  balance
# Los montos llevan SIEMPRE 2 decimales; así un número suelto en la descripción
# (p.ej. "...Claro 5") no se confunde con un monto.
_AMT = r'-?[\d,]*\d\.\d{2}'
_PDF_RE = re.compile(
    r'^(\d{2}/\d{2}/\d{4})\s+(\S+)\s+([A-Z0-9]{2})\s+(.+?)\s+'
    r'(' + _AMT + r')\s+(' + _AMT + r')\s+(' + _AMT + r')\s*$'
)


def _parse_pdf(raw):
    import pdfplumber
    movs = []
    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            for line in text.splitlines():
                m = _PDF_RE.match(line.strip())
                if not m:
                    continue
                fecha, ref, code, desc, deb, cred, bal = m.groups()
                mov = _mk_mov(_parse_date(fecha), desc, deb, cred,
                              ref=ref, code=code, balance=_num(bal))
                if mov:
                    movs.append(mov)
    return movs


# ── Tabular (CSV / Excel) ───────────────────────────────────────────────────────

def _parse_table_rows(rows):
    """rows: lista de listas (celdas). Detecta encabezado y extrae movimientos."""
    cols = {}
    hidx = None
    for i, row in enumerate(rows[:60]):
        norm = [_norm(c) for c in row]
        joined = ' '.join(norm)
        if 'fecha' not in joined:
            continue
        if not ('debito' in joined or 'credito' in joined or 'descrip' in joined):
            continue
        tmp = {}
        for j, c in enumerate(norm):
            if c.startswith('fecha') and 'fecha' not in tmp:
                tmp['fecha'] = j
            elif 'descrip' in c:
                tmp['desc'] = j
            elif 'debito' in c or c in ('debe', 'debito'):
                tmp['deb'] = j
            elif 'credito' in c or c == 'haber':
                tmp['cred'] = j
            elif 'referen' in c:
                tmp['ref'] = j
            elif 'codig' in c:
                tmp['code'] = j
            elif c.startswith('balance') or c == 'saldo':
                tmp['bal'] = j
        if 'fecha' in tmp and ('deb' in tmp or 'cred' in tmp):
            cols = tmp
            hidx = i
            break
    if hidx is None:
        return []

    def cell(row, key):
        j = cols.get(key)
        if j is None or j >= len(row):
            return None
        return row[j]

    movs = []
    for row in rows[hidx + 1:]:
        if not any(str(c).strip() for c in row if c is not None):
            continue
        d = _parse_date(cell(row, 'fecha'))
        if not d:  # totales / resumen / filas vacías
            continue
        mov = _mk_mov(d, cell(row, 'desc'), cell(row, 'deb'), cell(row, 'cred'),
                      ref=cell(row, 'ref') or '', code=cell(row, 'code') or '',
                      balance=_num(cell(row, 'bal')))
        if mov:
            movs.append(mov)
    return movs


def _parse_csv(raw):
    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode('latin-1', errors='ignore')
    sample = text[:4096]
    delim = ';' if sample.count(';') > sample.count(',') else ','
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    return _parse_table_rows(rows)


def _rows_xlsx(raw):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _rows_xls_biff(raw):
    """.xls binario clásico (OLE2/BIFF) con xlrd."""
    import xlrd
    book = xlrd.open_workbook(file_contents=raw)
    sh = book.sheet_by_index(0)
    rows = []
    for i in range(sh.nrows):
        row = []
        for j in range(sh.ncols):
            cell = sh.cell(i, j)
            if cell.ctype == 0:        # vacío
                row.append('')
            elif cell.ctype == 3:      # fecha (serial) → datetime
                try:
                    row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                except Exception:
                    row.append(cell.value)
            else:
                row.append(cell.value)
        rows.append(row)
    return rows


def _rows_html(raw):
    """.xls exportado como tabla HTML (muy común en banca en línea)."""
    import lxml.html
    doc = lxml.html.fromstring(raw)
    rows = []
    for tr in doc.xpath('//tr'):
        cells = [(' '.join(c.text_content().split())) for c in tr.xpath('./td | ./th')]
        if any(cells):
            rows.append(cells)
    return rows


def _rows_spreadsheetml(raw):
    """.xls exportado como XML SpreadsheetML (Office 2003)."""
    from lxml import etree
    ns = '{urn:schemas-microsoft-com:office:spreadsheet}'
    root = etree.fromstring(raw)
    rows = []
    for row_el in root.iter(ns + 'Row'):
        cells = []
        for cell_el in row_el.findall(ns + 'Cell'):
            data = cell_el.find(ns + 'Data')
            cells.append((data.text or '').strip() if data is not None else '')
        if any(cells):
            rows.append(cells)
    return rows


def _rows_tabular(raw):
    """Detecta la variante real del archivo 'Excel' y devuelve filas."""
    head = raw[:512].lstrip()
    low = head.lower()
    if raw[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':      # OLE2 → .xls binario
        return _rows_xls_biff(raw)
    if raw[:2] == b'PK':                                    # ZIP → .xlsx
        return _rows_xlsx(raw)
    if low.startswith(b'<?xml') and b'spreadsheet' in raw[:3000].lower():
        return _rows_spreadsheetml(raw)
    if b'<table' in low or b'<html' in low or low.startswith(b'<!doctype html'):
        return _rows_html(raw)
    # último intento: openpyxl
    return _rows_xlsx(raw)


def _parse_xlsx(raw):
    return _parse_table_rows(_rows_tabular(raw))


# ── entrada principal ────────────────────────────────────────────────────────────

def _statement_text(filename, raw):
    """Texto plano del estado de cuenta (para leer los saldos del encabezado)."""
    name = (filename or '').lower().strip()
    try:
        if name.endswith('.pdf') or raw[:4] == b'%PDF':
            import pdfplumber
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                return (pdf.pages[0].extract_text() or '') if pdf.pages else ''
        if name.endswith('.csv') or name.endswith('.txt'):
            for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
                try:
                    return raw.decode(enc)
                except UnicodeDecodeError:
                    continue
            return raw.decode('latin-1', errors='ignore')
        rows = _rows_tabular(raw)
        return '\n'.join(' '.join('' if c is None else str(c) for c in r) for r in rows[:60])
    except Exception:
        return ''


# "LPS 4998.15 3850.02 1148.13"  →  libros, retenido, disponible
_SALDOS_RE = re.compile(
    r'(?:^|\n)\s*(?:[A-Z]{3}\s+)?(-?[\d,]*\d\.\d{2})\s+(-?[\d,]*\d\.\d{2})\s+(-?[\d,]*\d\.\d{2})\s*(?:\n|$)')


def parse_saldos(filename, raw):
    """Devuelve {'libros','retenido','disponible'} del encabezado, o {} si no se hallan."""
    text = _statement_text(filename, raw)
    if 'disponible' not in text.lower():
        return {}
    m = _SALDOS_RE.search(text)
    if not m:
        return {}
    libros, retenido, disponible = (_num(x) for x in m.groups())
    return {'libros': libros, 'retenido': retenido, 'disponible': disponible}


def parse_statement(filename, raw):
    """Despacha por extensión. Lanza ValueError si el formato no se reconoce."""
    name = (filename or '').lower().strip()
    if name.endswith('.pdf'):
        return _parse_pdf(raw)
    if name.endswith('.xlsx') or name.endswith('.xlsm'):
        return _parse_xlsx(raw)
    if name.endswith('.xls'):
        # El ".xls" del banco puede ser BIFF real, HTML o XML SpreadsheetML.
        return _parse_xlsx(raw)
    if name.endswith('.csv') or name.endswith('.txt'):
        return _parse_csv(raw)
    # Sin extensión clara: intentar PDF, luego Excel, luego CSV.
    if raw[:4] == b'%PDF':
        return _parse_pdf(raw)
    if raw[:2] == b'PK':
        return _parse_xlsx(raw)
    return _parse_csv(raw)
