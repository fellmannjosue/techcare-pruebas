import csv
import io
import json
from decimal import Decimal
from datetime import date, datetime
from functools import wraps

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Sum
from django.conf import settings

_FINANZAS_EMAIL = 'cvalle@ana-hn.org'


def finanzas_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect(f"{settings.LOGIN_URL}?next={request.path}")
        if not request.user.is_superuser:
            if request.content_type == 'application/json' or request.headers.get('Accept') == 'application/json':
                return JsonResponse({'error': 'Acceso no autorizado.'}, status=403)
            return HttpResponseForbidden('No tienes permiso para acceder a este módulo.')
        return view_func(request, *args, **kwargs)
    return _wrapped


from .models import (
    Categoria, Cuenta, Transaccion, Pendiente, EntradaRapida, Recurrente,
    Presupuesto, PresupuestoCategoria, MetaAhorro, ConfiguracionUsuario,
)

SEED_CATEGORIAS = [
    {'name': 'Salario',         'type': 'income',  'color': '#2A7A4F', 'icon': 'ti-cash'},
    {'name': 'Ventas',          'type': 'income',  'color': '#C9A84C', 'icon': 'ti-businessplan'},
    {'name': 'Otros ingresos',  'type': 'income',  'color': '#3A8FB7', 'icon': 'ti-plus'},
    {'name': 'Alimentación',    'type': 'expense', 'color': '#B84040', 'icon': 'ti-shopping-cart'},
    {'name': 'Transporte',      'type': 'expense', 'color': '#3A6FA8', 'icon': 'ti-car'},
    {'name': 'Servicios',       'type': 'expense', 'color': '#7B5EA7', 'icon': 'ti-bolt'},
    {'name': 'Vivienda',        'type': 'expense', 'color': '#8A6D3B', 'icon': 'ti-home'},
    {'name': 'Salud',           'type': 'expense', 'color': '#2F9E8F', 'icon': 'ti-heartbeat'},
    {'name': 'Entretenimiento', 'type': 'expense', 'color': '#D4843C', 'icon': 'ti-device-gamepad'},
]

SEED_CUENTAS = [
    {'name': 'Efectivo', 'type': 'cash', 'color': '#2A7A4F', 'icon': 'ti-cash',        'initial': 0},
    {'name': 'Banco',    'type': 'bank', 'color': '#3A6FA8', 'icon': 'ti-building-bank', 'initial': 0},
]


def _body(request):
    try:
        return json.loads(request.body)
    except Exception:
        return {}


def _cat(user, cat_id):
    if not cat_id:
        return None
    return Categoria.objects.filter(pk=cat_id, user=user).first()


def _cuenta(user, acc_id):
    if not acc_id:
        return None
    return Cuenta.objects.filter(pk=acc_id, user=user).first()


def _dec(v, default='0'):
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal(default)


# ─── GENERAR RECURRENTES VENCIDOS ──────────────────────────────────────────────

def _generar_recurrentes(user):
    """Materializa en transacciones los recurrentes vencidos (con catch-up)."""
    hoy = date.today()
    for r in Recurrente.objects.filter(user=user, activo=True):
        guard = 0
        while r.proxima <= hoy and guard < 120:
            Transaccion.objects.create(
                user=user, tipo=r.tipo, monto=r.monto, descripcion=r.nombre,
                categoria=r.categoria, cuenta=r.cuenta, fecha=r.proxima,
            )
            r.avanzar()
            guard += 1
        if guard:
            r.save()


def _balance_cuenta(cuenta_id, txns):
    """Saldo de una cuenta a partir de la lista de transacciones (dicts)."""
    saldo = Decimal('0')
    for t in txns:
        amt = _dec(t['amount'])
        if t['type'] == 'income' and t['accountId'] == cuenta_id:
            saldo += amt
        elif t['type'] == 'expense' and t['accountId'] == cuenta_id:
            saldo -= amt
        elif t['type'] == 'transfer':
            if t['accountId'] == cuenta_id:
                saldo -= amt
            if t['toAccountId'] == cuenta_id:
                saldo += amt
    return saldo


# ─── PÁGINA PRINCIPAL ──────────────────────────────────────────────────────────

@finanzas_required
def index(request):
    return render(request, 'finanzas_personales/index.html')


# ─── API: ESTADO COMPLETO ──────────────────────────────────────────────────────

@finanzas_required
@require_http_methods(['GET'])
def api_data(request):
    user = request.user
    if not Categoria.objects.filter(user=user).exists():
        for c in SEED_CATEGORIAS:
            Categoria.objects.create(user=user, nombre=c['name'], tipo=c['type'],
                                     color=c['color'], icono=c['icon'])
    if not Cuenta.objects.filter(user=user).exists():
        for i, a in enumerate(SEED_CUENTAS):
            Cuenta.objects.create(user=user, nombre=a['name'], tipo=a['type'],
                                  color=a['color'], icono=a['icon'],
                                  saldo_inicial=a['initial'], orden=i)

    _generar_recurrentes(user)

    config, _ = ConfiguracionUsuario.objects.get_or_create(user=user)
    txns = [t.to_dict() for t in Transaccion.objects.filter(user=user).select_related('categoria', 'cuenta')]

    cuentas = []
    for c in Cuenta.objects.filter(user=user):
        d = c.to_dict()
        d['balance'] = float(c.saldo_inicial + _balance_cuenta(str(c.pk), txns))
        cuentas.append(d)

    return JsonResponse({
        'categories':   [c.to_dict() for c in Categoria.objects.filter(user=user)],
        'accounts':     cuentas,
        'transactions': txns,
        'pendings':     [p.to_dict() for p in Pendiente.objects.filter(user=user)],
        'budgets':      [b.to_dict() for b in Presupuesto.objects.filter(user=user).prefetch_related('items')],
        'quickEntries': [e.to_dict() for e in EntradaRapida.objects.filter(user=user)],
        'recurrents':   [r.to_dict() for r in Recurrente.objects.filter(user=user)],
        'goals':        [g.to_dict() for g in MetaAhorro.objects.filter(user=user)],
        'settings':     config.to_dict(),
    })


# ─── API: CATEGORÍAS ──────────────────────────────────────────────────────────

@finanzas_required
@require_http_methods(['POST'])
def api_categoria_crear(request):
    b = _body(request)
    cat = Categoria.objects.create(
        user=request.user,
        nombre=b.get('name', '').strip(),
        tipo=b.get('type', 'expense'),
        color=b.get('color', '#999999'),
        icono=b.get('icon', 'ti-tag'),
    )
    return JsonResponse(cat.to_dict(), status=201)


@finanzas_required
@require_http_methods(['PUT', 'DELETE'])
def api_categoria_detalle(request, pk):
    cat = get_object_or_404(Categoria, pk=pk, user=request.user)
    if request.method == 'DELETE':
        cat.delete()
        return JsonResponse({'ok': True})
    b = _body(request)
    cat.nombre = b.get('name', cat.nombre).strip()
    cat.tipo   = b.get('type', cat.tipo)
    cat.color  = b.get('color', cat.color)
    cat.icono  = b.get('icon', cat.icono)
    cat.save()
    return JsonResponse(cat.to_dict())


# ─── API: CUENTAS ─────────────────────────────────────────────────────────────

@finanzas_required
@require_http_methods(['POST'])
def api_cuenta_crear(request):
    b = _body(request)
    n = Cuenta.objects.filter(user=request.user).count()
    cta = Cuenta.objects.create(
        user=request.user,
        nombre=b.get('name', '').strip(),
        tipo=b.get('type', 'cash'),
        moneda=b.get('currency', 'L'),
        saldo_inicial=_dec(b.get('initial', 0)),
        color=b.get('color', '#2A7A4F'),
        icono=b.get('icon', 'ti-wallet'),
        orden=n,
    )
    return JsonResponse(cta.to_dict(), status=201)


@finanzas_required
@require_http_methods(['PUT', 'DELETE'])
def api_cuenta_detalle(request, pk):
    cta = get_object_or_404(Cuenta, pk=pk, user=request.user)
    if request.method == 'DELETE':
        cta.delete()
        return JsonResponse({'ok': True})
    b = _body(request)
    cta.nombre        = b.get('name', cta.nombre).strip()
    cta.tipo          = b.get('type', cta.tipo)
    cta.moneda        = b.get('currency', cta.moneda)
    cta.saldo_inicial = _dec(b.get('initial', cta.saldo_inicial))
    cta.color         = b.get('color', cta.color)
    cta.icono         = b.get('icon', cta.icono)
    cta.archivada     = bool(b.get('archived', cta.archivada))
    cta.save()
    return JsonResponse(cta.to_dict())


# ─── API: TRANSACCIONES ───────────────────────────────────────────────────────

@finanzas_required
@require_http_methods(['POST'])
def api_transaccion_crear(request):
    b = _body(request)
    txn = Transaccion.objects.create(
        user=request.user,
        tipo=b.get('type', 'expense'),
        monto=_dec(b.get('amount', 0)),
        descripcion=b.get('description', '').strip(),
        categoria=_cat(request.user, b.get('categoryId')),
        cuenta=_cuenta(request.user, b.get('accountId')),
        cuenta_destino=_cuenta(request.user, b.get('toAccountId')),
        fecha=b.get('date', str(date.today())),
    )
    return JsonResponse(txn.to_dict(), status=201)


@finanzas_required
@require_http_methods(['POST'])
def api_transaccion_bulk(request):
    """Acciones masivas: asignar categoría o eliminar varias transacciones."""
    b = _body(request)
    ids = b.get('ids', [])
    qs = Transaccion.objects.filter(user=request.user, pk__in=ids)
    accion = b.get('action')
    if accion == 'delete':
        n = qs.count()
        qs.delete()
        return JsonResponse({'ok': True, 'afectados': n})
    if accion == 'categoria':
        cat = _cat(request.user, b.get('categoryId'))
        n = qs.update(categoria=cat)
        return JsonResponse({'ok': True, 'afectados': n})
    return JsonResponse({'error': 'Acción inválida.'}, status=400)


@finanzas_required
@require_http_methods(['PUT', 'DELETE'])
def api_transaccion_detalle(request, pk):
    txn = get_object_or_404(Transaccion, pk=pk, user=request.user)
    if request.method == 'DELETE':
        txn.delete()
        return JsonResponse({'ok': True})
    b = _body(request)
    txn.tipo           = b.get('type', txn.tipo)
    txn.monto          = _dec(b.get('amount', txn.monto))
    txn.descripcion    = b.get('description', txn.descripcion).strip()
    txn.categoria      = _cat(request.user, b.get('categoryId'))
    txn.cuenta         = _cuenta(request.user, b.get('accountId'))
    txn.cuenta_destino = _cuenta(request.user, b.get('toAccountId'))
    txn.fecha          = b.get('date', str(txn.fecha))
    txn.save()
    return JsonResponse(txn.to_dict())


# ─── API: PENDIENTES ──────────────────────────────────────────────────────────

@finanzas_required
@require_http_methods(['POST'])
def api_pendiente_crear(request):
    b = _body(request)
    p = Pendiente.objects.create(
        user=request.user,
        tipo=b.get('type', 'income'),
        nombre=b.get('name', '').strip(),
        monto=_dec(b.get('amount', 0)),
        fecha=b.get('date', str(date.today())),
    )
    return JsonResponse(p.to_dict(), status=201)


@finanzas_required
@require_http_methods(['DELETE'])
def api_pendiente_eliminar(request, pk):
    get_object_or_404(Pendiente, pk=pk, user=request.user).delete()
    return JsonResponse({'ok': True})


@finanzas_required
@require_http_methods(['POST'])
def api_pendiente_confirmar(request, pk):
    p = get_object_or_404(Pendiente, pk=pk, user=request.user)
    b = _body(request)
    cat = Categoria.objects.filter(user=request.user, tipo=p.tipo).first()
    txn = Transaccion.objects.create(
        user=request.user, tipo=p.tipo, monto=p.monto,
        descripcion=p.nombre, categoria=cat,
        cuenta=_cuenta(request.user, b.get('accountId')),
        fecha=date.today(),
    )
    p.delete()
    return JsonResponse(txn.to_dict(), status=201)


# ─── API: ENTRADAS RÁPIDAS ────────────────────────────────────────────────────

@finanzas_required
@require_http_methods(['POST'])
def api_qe_crear(request):
    b = _body(request)
    qe = EntradaRapida.objects.create(
        user=request.user,
        nombre=b.get('name', '').strip(),
        monto=_dec(b.get('amount', 0)),
        tipo=b.get('type', 'expense'),
        categoria=_cat(request.user, b.get('categoryId')),
        cuenta=_cuenta(request.user, b.get('accountId')),
        color=b.get('color', '#2A7A4F'),
        icono=b.get('icon', 'ti-bolt'),
    )
    return JsonResponse(qe.to_dict(), status=201)


@finanzas_required
@require_http_methods(['DELETE'])
def api_qe_eliminar(request, pk):
    get_object_or_404(EntradaRapida, pk=pk, user=request.user).delete()
    return JsonResponse({'ok': True})


@finanzas_required
@require_http_methods(['POST'])
def api_qe_ejecutar(request, pk):
    qe = get_object_or_404(EntradaRapida, pk=pk, user=request.user)
    txn = Transaccion.objects.create(
        user=request.user, tipo=qe.tipo, monto=qe.monto,
        descripcion=qe.nombre, categoria=qe.categoria, cuenta=qe.cuenta, fecha=date.today(),
    )
    return JsonResponse(txn.to_dict(), status=201)


# ─── API: RECURRENTES ─────────────────────────────────────────────────────────

@finanzas_required
@require_http_methods(['POST'])
def api_recurrente_crear(request):
    b = _body(request)
    r = Recurrente.objects.create(
        user=request.user,
        tipo=b.get('type', 'expense'),
        nombre=b.get('name', '').strip(),
        monto=_dec(b.get('amount', 0)),
        categoria=_cat(request.user, b.get('categoryId')),
        cuenta=_cuenta(request.user, b.get('accountId')),
        frecuencia=b.get('freq', 'monthly'),
        proxima=b.get('next', str(date.today())),
        activo=bool(b.get('active', True)),
    )
    return JsonResponse(r.to_dict(), status=201)


@finanzas_required
@require_http_methods(['PUT', 'DELETE'])
def api_recurrente_detalle(request, pk):
    r = get_object_or_404(Recurrente, pk=pk, user=request.user)
    if request.method == 'DELETE':
        r.delete()
        return JsonResponse({'ok': True})
    b = _body(request)
    r.tipo       = b.get('type', r.tipo)
    r.nombre     = b.get('name', r.nombre).strip()
    r.monto      = _dec(b.get('amount', r.monto))
    r.categoria  = _cat(request.user, b.get('categoryId'))
    r.cuenta     = _cuenta(request.user, b.get('accountId'))
    r.frecuencia = b.get('freq', r.frecuencia)
    r.proxima    = b.get('next', str(r.proxima))
    if 'active' in b:
        r.activo = bool(b['active'])
    r.save()
    return JsonResponse(r.to_dict())


# ─── API: PRESUPUESTOS ────────────────────────────────────────────────────────

@finanzas_required
@require_http_methods(['POST'])
def api_presupuesto_crear(request):
    b = _body(request)
    pres = Presupuesto.objects.create(
        user=request.user,
        nombre=b.get('name', '').strip(),
        limite=_dec(b.get('limit', 0)),
    )
    for item in b.get('items', []):
        cat = _cat(request.user, item.get('catId'))
        if cat:
            PresupuestoCategoria.objects.create(presupuesto=pres, categoria=cat)
    return JsonResponse(pres.to_dict(), status=201)


@finanzas_required
@require_http_methods(['DELETE'])
def api_presupuesto_eliminar(request, pk):
    get_object_or_404(Presupuesto, pk=pk, user=request.user).delete()
    return JsonResponse({'ok': True})


# ─── API: METAS DE AHORRO ────────────────────────────────────────────────────

@finanzas_required
@require_http_methods(['POST'])
def api_meta_crear(request):
    b = _body(request)
    meta = MetaAhorro.objects.create(
        user=request.user,
        nombre=b.get('name', '').strip(),
        objetivo=_dec(b.get('target', 0)),
        fecha_limite=b.get('deadline') or None,
        emoji=b.get('emoji', '🎯'),
    )
    return JsonResponse(meta.to_dict(), status=201)


@finanzas_required
@require_http_methods(['PUT', 'DELETE'])
def api_meta_detalle(request, pk):
    meta = get_object_or_404(MetaAhorro, pk=pk, user=request.user)
    if request.method == 'DELETE':
        meta.delete()
        return JsonResponse({'ok': True})
    b = _body(request)
    meta.nombre       = b.get('name', meta.nombre).strip()
    meta.objetivo     = _dec(b.get('target', meta.objetivo))
    meta.fecha_limite = b.get('deadline') or None
    meta.emoji        = b.get('emoji', meta.emoji)
    meta.save()
    return JsonResponse(meta.to_dict())


@finanzas_required
@require_http_methods(['POST'])
def api_meta_ahorrar(request, pk):
    meta = get_object_or_404(MetaAhorro, pk=pk, user=request.user)
    b = _body(request)
    amount = _dec(b.get('amount', 0))
    if amount <= 0:
        return JsonResponse({'error': 'Monto inválido'}, status=400)
    meta.ahorrado = min(meta.ahorrado + amount, meta.objetivo)
    meta.save()
    cat = Categoria.objects.filter(user=request.user, tipo='expense').first()
    txn = Transaccion.objects.create(
        user=request.user, tipo='expense', monto=amount,
        descripcion=f'Ahorro: {meta.nombre}', categoria=cat,
        cuenta=_cuenta(request.user, b.get('accountId')), fecha=date.today(),
    )
    return JsonResponse({'meta': meta.to_dict(), 'txn': txn.to_dict()})


# ─── API: CONFIGURACIÓN ───────────────────────────────────────────────────────

@finanzas_required
@require_http_methods(['POST'])
def api_configuracion(request):
    b = _body(request)
    config, _ = ConfiguracionUsuario.objects.get_or_create(user=request.user)
    if 'currency' in b:
        config.moneda = b['currency']
    if 'theme' in b:
        config.tema = b['theme']
    config.save()
    return JsonResponse(config.to_dict())


# ─── REPORTES / EXPORTACIÓN ───────────────────────────────────────────────────

def _rango_txns(user, desde, hasta):
    qs = Transaccion.objects.filter(user=user).select_related('categoria', 'cuenta', 'cuenta_destino')
    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)
    return qs.order_by('fecha', 'creado')


@finanzas_required
@require_http_methods(['GET'])
def api_reporte(request):
    user = request.user
    desde = request.GET.get('from') or None
    hasta = request.GET.get('to') or None
    fmt = request.GET.get('format', 'csv')
    config, _ = ConfiguracionUsuario.objects.get_or_create(user=user)
    sym = config.moneda
    txns = _rango_txns(user, desde, hasta)

    tipo_lbl = {'income': 'Ingreso', 'expense': 'Gasto', 'transfer': 'Transferencia'}
    rows = []
    tot_in = tot_out = Decimal('0')
    for t in txns:
        if t.tipo == 'income':
            tot_in += t.monto
        elif t.tipo == 'expense':
            tot_out += t.monto
        rows.append([
            str(t.fecha), tipo_lbl.get(t.tipo, t.tipo),
            t.categoria.nombre if t.categoria else '',
            t.cuenta.nombre if t.cuenta else '',
            t.descripcion or '', float(t.monto),
        ])

    periodo = f"{desde or '—'} a {hasta or '—'}"
    headers = ['Fecha', 'Tipo', 'Categoría', 'Cuenta', 'Descripción', 'Monto']

    if fmt == 'csv':
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(headers)
        for r in rows:
            w.writerow(r)
        w.writerow([])
        w.writerow(['', '', '', '', 'Total ingresos', float(tot_in)])
        w.writerow(['', '', '', '', 'Total gastos', float(tot_out)])
        w.writerow(['', '', '', '', 'Balance', float(tot_in - tot_out)])
        resp = HttpResponse(out.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = 'attachment; filename="reporte_finanzas.csv"'
        return resp

    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = 'Movimientos'
        hdr_fill = PatternFill('solid', fgColor='2A7A4F')
        hdr_font = Font(bold=True, color='FFFFFF')
        ws.append([f'Reporte de Finanzas — {periodo}'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(headers)
        for c in ws[3]:
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal='center')
        for r in rows:
            ws.append(r)
        ws.append([])
        ws.append(['', '', '', '', 'Total ingresos', float(tot_in)])
        ws.append(['', '', '', '', 'Total gastos', float(tot_out)])
        ws.append(['', '', '', '', 'Balance', float(tot_in - tot_out)])
        for col, width in zip('ABCDEF', (12, 14, 18, 16, 40, 14)):
            ws.column_dimensions[col].width = width
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="reporte_finanzas.xlsx"'
        return resp

    # PDF (WeasyPrint)
    from weasyprint import HTML
    # Resumen por categoría (gastos)
    por_cat = {}
    for t in txns:
        if t.tipo == 'expense':
            k = t.categoria.nombre if t.categoria else 'Sin categoría'
            por_cat[k] = por_cat.get(k, Decimal('0')) + t.monto
    por_cat = sorted(por_cat.items(), key=lambda x: x[1], reverse=True)
    html = render(request, 'finanzas_personales/reporte_pdf.html', {
        'periodo': periodo, 'rows': rows, 'headers': headers,
        'tot_in': tot_in, 'tot_out': tot_out, 'balance': tot_in - tot_out,
        'sym': sym, 'por_cat': por_cat, 'generado': datetime.now(),
    }).content.decode('utf-8')
    pdf = HTML(string=html).write_pdf()
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = 'inline; filename="reporte_finanzas.pdf"'
    return resp


# ─── API: IMPORTAR / LIMPIAR ──────────────────────────────────────────────────

def _wipe(user):
    Transaccion.objects.filter(user=user).delete()
    Recurrente.objects.filter(user=user).delete()
    EntradaRapida.objects.filter(user=user).delete()
    Presupuesto.objects.filter(user=user).delete()
    MetaAhorro.objects.filter(user=user).delete()
    Pendiente.objects.filter(user=user).delete()
    Cuenta.objects.filter(user=user).delete()
    Categoria.objects.filter(user=user).delete()


@finanzas_required
@require_http_methods(['POST'])
def api_importar(request):
    b = _body(request)
    user = request.user
    _wipe(user)

    cat_map = {}
    for c in b.get('categories', []):
        obj = Categoria.objects.create(user=user, nombre=c['name'], tipo=c['type'],
                                       color=c.get('color', '#999999'), icono=c.get('icon', 'ti-tag'))
        cat_map[str(c['id'])] = obj

    acc_map = {}
    for i, a in enumerate(b.get('accounts', [])):
        obj = Cuenta.objects.create(user=user, nombre=a['name'], tipo=a.get('type', 'cash'),
                                    saldo_inicial=_dec(a.get('initial', 0)), color=a.get('color', '#2A7A4F'),
                                    icono=a.get('icon', 'ti-wallet'), archivada=a.get('archived', False), orden=i)
        acc_map[str(a['id'])] = obj

    for t in b.get('transactions', []):
        Transaccion.objects.create(
            user=user, tipo=t['type'], monto=_dec(t['amount']),
            descripcion=t.get('description', ''), fecha=t.get('date', str(date.today())),
            categoria=cat_map.get(str(t.get('categoryId', ''))),
            cuenta=acc_map.get(str(t.get('accountId', ''))),
            cuenta_destino=acc_map.get(str(t.get('toAccountId', ''))),
        )

    for p in b.get('pendings', []):
        Pendiente.objects.create(user=user, tipo=p['type'], nombre=p['name'],
                                 monto=_dec(p['amount']), fecha=p.get('date', str(date.today())))

    for qe in b.get('quickEntries', []):
        EntradaRapida.objects.create(user=user, nombre=qe['name'], monto=_dec(qe['amount']), tipo=qe['type'],
                                     categoria=cat_map.get(str(qe.get('categoryId', ''))),
                                     cuenta=acc_map.get(str(qe.get('accountId', ''))),
                                     color=qe.get('color', '#2A7A4F'), icono=qe.get('icon', 'ti-bolt'))

    for r in b.get('recurrents', []):
        Recurrente.objects.create(user=user, tipo=r['type'], nombre=r['name'], monto=_dec(r['amount']),
                                  categoria=cat_map.get(str(r.get('categoryId', ''))),
                                  cuenta=acc_map.get(str(r.get('accountId', ''))),
                                  frecuencia=r.get('freq', 'monthly'), proxima=r.get('next', str(date.today())),
                                  activo=r.get('active', True))

    for bud in b.get('budgets', []):
        pres = Presupuesto.objects.create(user=user, nombre=bud['name'], limite=_dec(bud['limit']))
        for item in bud.get('items', []):
            cat = cat_map.get(str(item.get('catId', '')))
            if cat:
                PresupuestoCategoria.objects.create(presupuesto=pres, categoria=cat)

    for g in b.get('goals', []):
        MetaAhorro.objects.create(user=user, nombre=g['name'], objetivo=_dec(g['target']),
                                  ahorrado=_dec(g.get('saved', 0)), fecha_limite=g.get('deadline') or None,
                                  emoji=g.get('emoji', '🎯'))

    if 'settings' in b:
        config, _ = ConfiguracionUsuario.objects.get_or_create(user=user)
        config.moneda = b['settings'].get('currency', config.moneda)
        config.tema   = b['settings'].get('theme', config.tema)
        config.save()

    return JsonResponse({'ok': True})


# ─── API: IMPORTAR ESTADO DE CUENTA (PDF / EXCEL / CSV) ───────────────────────

def _transfer_cat(user, tipo):
    """Categoría 'Transferencias' (excluida de totales) para el tipo dado."""
    cat = Categoria.objects.filter(user=user, nombre='Transferencias', tipo=tipo).first()
    if not cat:
        cat = Categoria.objects.create(
            user=user, nombre='Transferencias', tipo=tipo,
            color='#3a8fb7', icono='ti-arrows-exchange', excluir_totales=True)
    return cat


def _es_transferencia(mv):
    """Detecta movimientos que son transferencias entre cuentas (no ingreso/gasto real)."""
    return (mv.get('code') or '').upper() == 'TF'


def _ajustar_saldo(user, cuenta, objetivo):
    """Fija saldo_inicial para que el saldo mostrado = objetivo."""
    qs = Transaccion.objects.filter(user=user)
    inc   = qs.filter(tipo='income',  cuenta=cuenta).aggregate(s=Sum('monto'))['s'] or 0
    exp   = qs.filter(tipo='expense', cuenta=cuenta).aggregate(s=Sum('monto'))['s'] or 0
    t_out = qs.filter(tipo='transfer', cuenta=cuenta).aggregate(s=Sum('monto'))['s'] or 0
    t_in  = qs.filter(tipo='transfer', cuenta_destino=cuenta).aggregate(s=Sum('monto'))['s'] or 0
    neto = Decimal(str(inc)) - Decimal(str(exp)) - Decimal(str(t_out)) + Decimal(str(t_in))
    cuenta.saldo_inicial = (Decimal(str(objetivo)) - neto).quantize(Decimal('0.01'))
    cuenta.save(update_fields=['saldo_inicial'])


@finanzas_required
@require_http_methods(['POST'])
def api_importar_estado(request):
    from collections import defaultdict
    from .bank_import import parse_statement, parse_saldos

    archivo = request.FILES.get('archivo')
    cuenta  = _cuenta(request.user, request.POST.get('cuenta'))
    commit  = request.POST.get('commit') == '1'

    if not archivo:
        return JsonResponse({'error': 'Selecciona un archivo.'}, status=400)
    if not cuenta:
        return JsonResponse({'error': 'Selecciona una cuenta destino.'}, status=400)

    raw = archivo.read()
    try:
        movs = parse_statement(archivo.name, raw)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    except Exception:
        return JsonResponse({'error': 'No se pudo leer el archivo. Verifica que sea un estado '
                                      'de cuenta válido (PDF, Excel o CSV).'}, status=400)
    if not movs:
        return JsonResponse({'error': 'No se encontraron movimientos en el archivo.'}, status=400)

    saldos = parse_saldos(archivo.name, raw) or {}

    # Rango de fechas y saldo de cierre (en libros), tolerante al orden del archivo
    fechas = [m['date'] for m in movs]
    dmin, dmax = min(fechas), max(fechas)
    ascendente = movs[0]['date'] <= movs[-1]['date']
    cierre = None
    for m in (movs if ascendente else reversed(movs)):
        if m.get('balance') is not None:
            cierre = round(float(m['balance']), 2)
    disponible = saldos.get('disponible')
    sugerido = disponible if disponible is not None else cierre

    # Marcar transferencias y calcular totales reales (excluyendo transferencias)
    for mv in movs:
        mv['transfer'] = _es_transferencia(mv)
    ingresos = round(sum(m['amount'] for m in movs if m['type'] == 'income'  and not m['transfer']), 2)
    gastos   = round(sum(m['amount'] for m in movs if m['type'] == 'expense' and not m['transfer']), 2)
    transferencias = round(sum(m['amount'] for m in movs if m['transfer']), 2)

    # Existentes importados en el rango (para informar)
    en_rango = Transaccion.objects.filter(
        user=request.user, cuenta=cuenta, origen='import',
        fecha__gte=dmin, fecha__lte=dmax).count()

    resumen = {
        'total': len(movs), 'ingresos': ingresos, 'gastos': gastos,
        'transferencias': transferencias, 'cuenta': cuenta.nombre,
        'cierre': cierre, 'disponible': disponible, 'sugerido': sugerido,
        'desde': str(dmin), 'hasta': str(dmax), 'en_rango': en_rango,
    }

    if not commit:
        return JsonResponse({'rows': movs, 'summary': resumen})

    reemplazar = request.POST.get('reemplazar', '1') == '1'
    try:
        objetivo = float(request.POST.get('saldo_objetivo'))
    except (TypeError, ValueError):
        objetivo = sugerido

    if reemplazar:
        # Borrar lo importado antes en este rango (no toca lo manual) y reinsertar
        Transaccion.objects.filter(
            user=request.user, cuenta=cuenta, origen='import',
            fecha__gte=dmin, fecha__lte=dmax).delete()
        a_crear = movs
    else:
        # Anti-duplicados robusto por (fecha, monto, saldo corrido) — no depende del texto
        existentes = defaultdict(int)
        for f, mo, sr in Transaccion.objects.filter(
                user=request.user, cuenta=cuenta, origen='import',
                saldo_ref__isnull=False).values_list('fecha', 'monto', 'saldo_ref'):
            existentes[(str(f), f"{Decimal(str(mo)):.2f}", f"{Decimal(str(sr)):.2f}")] += 1
        vistos = defaultdict(int)
        a_crear = []
        for mv in movs:
            bal = mv.get('balance')
            key = (mv['date'], f"{Decimal(str(mv['amount'])):.2f}",
                   f"{Decimal(str(bal)):.2f}" if bal is not None else '')
            vistos[key] += 1
            if vistos[key] > existentes[key]:
                a_crear.append(mv)

    creados = 0
    cat_cache = {}
    for mv in a_crear:
        categoria = None
        if mv['transfer']:
            categoria = cat_cache.get(mv['type']) or _transfer_cat(request.user, mv['type'])
            cat_cache[mv['type']] = categoria
        Transaccion.objects.create(
            user=request.user, tipo=mv['type'], monto=_dec(mv['amount']),
            descripcion=mv['description'], cuenta=cuenta, fecha=mv['date'],
            categoria=categoria, origen='import',
            saldo_ref=_dec(mv['balance']) if mv.get('balance') is not None else None,
        )
        creados += 1

    ajustado = None
    if request.POST.get('ajustar', '1') == '1' and objetivo is not None:
        _ajustar_saldo(request.user, cuenta, objetivo)
        ajustado = round(float(objetivo), 2)

    return JsonResponse({'ok': True, 'creados': creados, 'cuenta': cuenta.nombre,
                         'ajustado': ajustado})


@finanzas_required
@require_http_methods(['POST'])
def api_limpiar(request):
    user = request.user
    _wipe(user)
    for c in SEED_CATEGORIAS:
        Categoria.objects.create(user=user, nombre=c['name'], tipo=c['type'], color=c['color'], icono=c['icon'])
    for i, a in enumerate(SEED_CUENTAS):
        Cuenta.objects.create(user=user, nombre=a['name'], tipo=a['type'], color=a['color'],
                              icono=a['icon'], saldo_inicial=a['initial'], orden=i)
    return JsonResponse({'ok': True})
