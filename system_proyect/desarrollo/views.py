# <--- hecho por claude code: vistas del módulo Gestión de Desarrollo (FASE 1 + FASE 2).
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import (
    RequerimientoDesarrollo, ProyectoDesarrollo, HistorialRequerimiento,
    ComentarioRequerimiento, AdjuntoRequerimiento, SolicitanteCatalogo, OpcionCatalogo,
    EstadoReq, Prioridad, Impacto, Urgencia, TipoReq, Clasificacion,
    EstadoProyecto, ESTADOS_CERRADOS,
)
from .forms import (RequerimientoNuevoForm, RequerimientoEditForm, ComentarioForm,
                    AdjuntoForm, ProyectoForm, ConvertirProyectoForm)
from .utils import (es_admin_desarrollo, es_dev_desarrollo, puede_ver_requerimiento,
                    es_solo_solicitante)
from . import services


# ────────────────────────────── Dashboard ──────────────────────────────
_ESTADO_HEX = {
    'recibido': '#6c757d', 'evaluacion': '#f59f00', 'pendiente_aprob': '#f76707',
    'aprobado': '#74b816', 'rechazado': '#d63939', 'planificado': '#17a2b8',
    'desarrollo': '#4263eb', 'pruebas': '#ae3ec9', 'listo_prod': '#0ca678',
    'produccion': '#2fb344', 'pausado': '#adb5bd', 'cancelado': '#343a40',
}
_PRIORIDAD_HEX = {'critica': '#d63939', 'alta': '#f76707', 'media': '#4263eb', 'baja': '#adb5bd'}
_SEMAFORO_HEX = {'verde': '#2fb344', 'amarillo': '#f59f00', 'rojo': '#d63939', 'gris': '#adb5bd'}


@login_required
def dashboard(request):
    """Dashboard (FASE 4): KPIs, gráficos (estado/tipo/área/prioridad),
    semáforos, próximos a vencer y actividad reciente."""
    if es_solo_solicitante(request.user):            # <--- staff: solo crear + su historial
        return redirect('desarrollo:mis_requerimientos')
    from collections import Counter, OrderedDict
    es_admin = es_admin_desarrollo(request.user)
    qs = RequerimientoDesarrollo.objects.all()
    if not es_admin:
        qs = qs.filter(Q(solicitante=request.user) | Q(responsable=request.user))

    kpis = {
        'total':      qs.count(),
        'recibidos':  qs.filter(estado=EstadoReq.RECIBIDO).count(),
        'evaluacion': qs.filter(estado=EstadoReq.EVALUACION).count(),
        'aprobados':  qs.filter(estado=EstadoReq.APROBADO).count(),
        'desarrollo': qs.filter(estado=EstadoReq.DESARROLLO).count(),
        'pruebas':    qs.filter(estado=EstadoReq.PRUEBAS).count(),
        'produccion': qs.filter(estado=EstadoReq.PRODUCCION).count(),
        'proyectos':  ProyectoDesarrollo.objects.count(),
    }

    # ── Conteos para gráficos ──
    def _grafico(campo, choices, colores):
        cont = dict(qs.values_list(campo).annotate(n=Count('id')))
        labels, data, bg = [], [], []
        for val, etq in choices:
            n = cont.get(val, 0)
            if n:
                labels.append(str(etq)); data.append(n); bg.append(colores.get(val, '#868e96'))
        return {'labels': labels, 'data': data, 'colors': bg}

    g_estado    = _grafico('estado', EstadoReq.choices, _ESTADO_HEX)
    g_prioridad = _grafico('prioridad', Prioridad.choices, _PRIORIDAD_HEX)
    g_tipo      = _grafico('tipo', TipoReq.choices, {})

    # Por área (dinámico, top 8)
    area_cont = (qs.exclude(area='').values('area').annotate(n=Count('id')).order_by('-n')[:8])
    g_area = {'labels': [a['area'] for a in area_cont], 'data': [a['n'] for a in area_cont],
              'colors': ['#206bc4'] * len(area_cont)}

    # ── Semáforos (propiedad Python; solo requerimientos abiertos) ──
    abiertos = [r for r in qs if not r.cerrado]
    sem = Counter(r.semaforo for r in abiertos)
    semaforos = OrderedDict([
        ('rojo',     sem.get('rojo', 0)),
        ('amarillo', sem.get('amarillo', 0)),
        ('verde',    sem.get('verde', 0)),
        ('gris',     sem.get('gris', 0)),
    ])
    g_semaforo = {'labels': ['Atrasado', 'Próximo', 'En tiempo', 'Sin fecha'],
                  'data': list(semaforos.values()),
                  'colors': [_SEMAFORO_HEX['rojo'], _SEMAFORO_HEX['amarillo'],
                             _SEMAFORO_HEX['verde'], _SEMAFORO_HEX['gris']]}

    # Próximos a vencer / atrasados (rojo + amarillo), ordenados por fecha
    por_vencer = sorted([r for r in abiertos if r.semaforo in ('rojo', 'amarillo')],
                        key=lambda r: r.fecha_estimada)[:8]

    # ── Actividad reciente global ──
    hist_qs = HistorialRequerimiento.objects.select_related('requerimiento', 'usuario')
    if not es_admin:
        hist_qs = hist_qs.filter(Q(requerimiento__solicitante=request.user) |
                                 Q(requerimiento__responsable=request.user))
    actividad = hist_qs.order_by('-fecha', '-id')[:12]

    return render(request, 'desarrollo/dashboard.html', {
        'kpis': kpis, 'es_admin': es_admin, 'nav_home_url': '/',
        'g_estado': g_estado, 'g_prioridad': g_prioridad, 'g_tipo': g_tipo,
        'g_area': g_area, 'g_semaforo': g_semaforo,
        'semaforos': semaforos, 'por_vencer': por_vencer, 'actividad': actividad,
    })


# ────────────────────────────── Nuevo ──────────────────────────────
@login_required
def requerimiento_nuevo(request):
    if request.method == 'POST':
        form = RequerimientoNuevoForm(request.POST)
        if form.is_valid():
            req = form.save(commit=False)
            req.solicitante = request.user           # el usuario obtiene desde request.user
            req.estado = EstadoReq.RECIBIDO
            req.fecha_solicitud = timezone.now().date()   # USE_TZ=False → now() es naive local
            req.creado_por = request.user
            req.modificado_por = request.user
            req.save()                               # genera REQ-AAAA-XXXX
            HistorialRequerimiento.objects.create(
                requerimiento=req, usuario=request.user,
                estado_anterior='', estado_nuevo=EstadoReq.RECIBIDO,
                comentario='Requerimiento creado')
            services.notificar_evento(req, 'nuevo')
            messages.success(request, f'Requerimiento {req.codigo} registrado.')
            return redirect('desarrollo:req_detail', codigo=req.codigo)
    else:
        form = RequerimientoNuevoForm()
    return render(request, 'desarrollo/req_form.html', {
        'form': form, 'solo_solicitante': es_solo_solicitante(request.user), 'nav_home_url': '/'})


@login_required
@require_POST
def solicitante_agregar(request):
    """Agrega un solicitante al catálogo desde el formulario (AJAX). Devuelve JSON."""
    nombre = (request.POST.get('nombre') or '').strip()
    if not nombre:
        return JsonResponse({'ok': False, 'error': 'Escribe un nombre.'}, status=400)
    if len(nombre) > 120:
        return JsonResponse({'ok': False, 'error': 'Nombre demasiado largo.'}, status=400)
    obj, creado = SolicitanteCatalogo.objects.get_or_create(
        nombre__iexact=nombre, defaults={'nombre': nombre})
    if not obj.activo:
        obj.activo = True
        obj.save(update_fields=['activo'])
    return JsonResponse({'ok': True, 'value': obj.pk, 'label': obj.nombre, 'creado': creado})


@login_required
@require_POST
def catalogo_agregar(request):
    """Agrega una opción de Área/Módulo al catálogo (AJAX). Guarda y devuelve texto."""
    grupo = (request.POST.get('grupo') or '').strip()
    nombre = (request.POST.get('nombre') or '').strip()
    if grupo not in ('area', 'modulo'):
        return JsonResponse({'ok': False, 'error': 'Grupo inválido.'}, status=400)
    if not nombre:
        return JsonResponse({'ok': False, 'error': 'Escribe un nombre.'}, status=400)
    if len(nombre) > 120:
        return JsonResponse({'ok': False, 'error': 'Nombre demasiado largo.'}, status=400)
    obj, creado = OpcionCatalogo.objects.get_or_create(
        grupo=grupo, valor__iexact=nombre, defaults={'valor': nombre})
    if not obj.activo:
        obj.activo = True
        obj.save(update_fields=['activo'])
    return JsonResponse({'ok': True, 'value': obj.valor, 'label': obj.valor, 'creado': creado})


# ────────────────────────────── Listado + filtros ──────────────────────────────
def _requerimientos_filtrados(request):
    """QuerySet de requerimientos aplicando permisos + filtros GET (reusado por listado y export)."""
    qs = (RequerimientoDesarrollo.objects
          .select_related('proyecto', 'solicitante', 'responsable'))
    if not es_admin_desarrollo(request.user):
        qs = qs.filter(Q(solicitante=request.user) | Q(responsable=request.user))
    g = request.GET
    buscar = (g.get('q') or '').strip()
    if buscar:
        qs = qs.filter(
            Q(codigo__icontains=buscar) | Q(titulo__icontains=buscar) |
            Q(descripcion__icontains=buscar) | Q(proyecto__nombre__icontains=buscar) |
            Q(solicitante__first_name__icontains=buscar) | Q(solicitante__username__icontains=buscar) |
            Q(responsable__first_name__icontains=buscar) | Q(responsable__username__icontains=buscar))
    for campo in ('estado', 'area', 'tipo', 'prioridad'):
        val = (g.get(campo) or '').strip()
        if val:
            qs = qs.filter(**{campo: val})
    if g.get('proyecto'):
        qs = qs.filter(proyecto_id=g['proyecto'])
    if g.get('responsable'):
        qs = qs.filter(responsable_id=g['responsable'])
    return qs


@login_required
def requerimiento_list(request):
    if es_solo_solicitante(request.user):            # <--- staff: su historial es Mis Requerimientos
        return redirect('desarrollo:mis_requerimientos')
    qs = _requerimientos_filtrados(request)
    g = request.GET
    paginator = Paginator(qs, 25)
    page = paginator.get_page(g.get('page'))

    ctx = {
        'page': page, 'total': paginator.count, 'filtros': g,
        'estados': EstadoReq.choices, 'tipos': TipoReq.choices, 'prioridades': Prioridad.choices,
        'proyectos': ProyectoDesarrollo.objects.order_by('nombre'),
        'areas': (RequerimientoDesarrollo.objects.exclude(area='')
                  .values_list('area', flat=True).distinct().order_by('area')),
        'es_admin': es_admin_desarrollo(request.user),
        'nav_home_url': '/',
    }
    return render(request, 'desarrollo/req_list.html', ctx)


# ────────────────────────────── Detalle ──────────────────────────────
@login_required
def requerimiento_detail(request, codigo):
    req = get_object_or_404(
        RequerimientoDesarrollo.objects.select_related('proyecto', 'solicitante', 'responsable'),
        codigo=codigo)
    if not puede_ver_requerimiento(request.user, req):
        raise PermissionDenied('No tienes permiso para ver este requerimiento.')
    return render(request, 'desarrollo/req_detail.html', {
        'req': req,
        'historial': req.historial.select_related('usuario', 'responsable_anterior', 'responsable_nuevo').all(),
        'comentarios': req.comentarios.select_related('usuario').all(),
        'adjuntos': req.adjuntos.select_related('usuario').all(),
        'comentario_form': ComentarioForm(),
        'adjunto_form': AdjuntoForm(),
        'puede_editar': es_dev_desarrollo(request.user) or req.responsable_id == request.user.id,
        'puede_proyecto': es_admin_desarrollo(request.user) or es_dev_desarrollo(request.user),
        'es_admin': es_admin_desarrollo(request.user),
        'nav_home_url': '/',
    })


# ────────────────────────────── Edición / seguimiento ──────────────────────────────
@login_required
def requerimiento_editar(request, codigo):
    req = get_object_or_404(RequerimientoDesarrollo, codigo=codigo)
    if not (es_dev_desarrollo(request.user) or req.responsable_id == request.user.id):
        raise PermissionDenied('No puedes editar este requerimiento.')

    if request.method == 'POST':
        form = RequerimientoEditForm(request.POST, instance=req)
        if form.is_valid():
            cd = form.cleaned_data
            # Campos NO rastreados en historial: se asignan directo.
            req.prioridad = cd['prioridad']; req.impacto = cd['impacto']; req.urgencia = cd['urgencia']
            req.proyecto = cd['proyecto']
            req.fecha_inicio = cd['fecha_inicio']; req.fecha_estimada = cd['fecha_estimada']
            req.fecha_finalizacion = cd['fecha_finalizacion']
            req.version_implementada = cd['version_implementada']; req.observaciones = cd['observaciones']
            # Estado/responsable/avance → registrar_cambio crea el historial y guarda.
            services.registrar_cambio(
                req, request.user,
                nuevo_estado=cd['estado'],
                nuevo_responsable=cd['responsable'],
                nuevo_porcentaje=cd['porcentaje_avance'])
            # Notificaciones de eventos importantes
            ev = {EstadoReq.APROBADO: 'aprobado', EstadoReq.PRUEBAS: 'pruebas',
                  EstadoReq.PRODUCCION: 'produccion'}.get(cd['estado'])
            if ev:
                services.notificar_evento(req, ev)
            messages.success(request, f'{req.codigo} actualizado.')
            return redirect('desarrollo:req_detail', codigo=req.codigo)
    else:
        form = RequerimientoEditForm(instance=req)
    return render(request, 'desarrollo/req_edit.html', {'form': form, 'req': req, 'nav_home_url': '/'})


# ────────────────────────────── Bitácora / adjuntos ──────────────────────────────
@login_required
@require_POST
def comentario_agregar(request, codigo):
    req = get_object_or_404(RequerimientoDesarrollo, codigo=codigo)
    if not puede_ver_requerimiento(request.user, req):
        raise PermissionDenied()
    form = ComentarioForm(request.POST)
    if form.is_valid() and form.cleaned_data['comentario'].strip():
        c = form.save(commit=False)
        c.requerimiento = req; c.usuario = request.user
        c.save()
        messages.success(request, 'Comentario agregado.')
    return redirect('desarrollo:req_detail', codigo=req.codigo)


@login_required
@require_POST
def adjunto_subir(request, codigo):
    req = get_object_or_404(RequerimientoDesarrollo, codigo=codigo)
    if not puede_ver_requerimiento(request.user, req):
        raise PermissionDenied()
    form = AdjuntoForm(request.POST, request.FILES)
    if form.is_valid():
        a = form.save(commit=False)
        a.requerimiento = req; a.usuario = request.user
        a.nombre_original = request.FILES['archivo'].name[:200]
        a.save()
        messages.success(request, 'Archivo adjuntado.')
    else:
        messages.error(request, form.errors.get('archivo', ['No se pudo subir el archivo.'])[0])
    return redirect('desarrollo:req_detail', codigo=req.codigo)


@login_required
@require_POST
def adjunto_eliminar(request, pk):
    a = get_object_or_404(AdjuntoRequerimiento.objects.select_related('requerimiento'), pk=pk)
    req = a.requerimiento
    if not (es_dev_desarrollo(request.user) or a.usuario_id == request.user.id):
        raise PermissionDenied()
    a.archivo.delete(save=False)
    a.delete()
    messages.success(request, 'Adjunto eliminado.')
    return redirect('desarrollo:req_detail', codigo=req.codigo)


# ══════════════════════════════ FASE 3 · Proyectos ══════════════════════════════
@login_required
def proyecto_list(request):
    """Listado de proyectos con conteo de requerimientos y avance promedio."""
    if es_solo_solicitante(request.user):
        return redirect('desarrollo:mis_requerimientos')
    qs = (ProyectoDesarrollo.objects
          .select_related('responsable')
          .annotate(n_req=Count('requerimientos', distinct=True),
                    avance=Avg('requerimientos__porcentaje_avance')))
    g = request.GET
    buscar = (g.get('q') or '').strip()
    if buscar:
        qs = qs.filter(Q(codigo__icontains=buscar) | Q(nombre__icontains=buscar) |
                       Q(modulo__icontains=buscar) | Q(area__icontains=buscar))
    estado = (g.get('estado') or '').strip()
    if estado:
        qs = qs.filter(estado=estado)
    qs = qs.order_by('-fecha_creado')
    return render(request, 'desarrollo/proy_list.html', {
        'proyectos': qs, 'total': qs.count(), 'filtros': g,
        'estados': EstadoProyecto.choices,
        'es_admin': es_admin_desarrollo(request.user), 'nav_home_url': '/',
    })


def _generar_prompt_claude(proy, reqs):
    """Arma un prompt listo para pegar en Claude Code con lo que hay que construir."""
    L = []
    L.append("Eres desarrollador del sistema TechCare (Django 6 / Python 3.13, Apache + mod_wsgi).")
    L.append("Construye e implementa el siguiente proyecto siguiendo las convenciones del repo")
    L.append("(JS y CSS en archivos aparte, comentarios de una sola línea, reiniciar Apache tras cambios,")
    L.append("y validar con manage.py check + prueba de humo antes de dar por hecho).")
    L.append("")
    L.append(f"# PROYECTO {proy.codigo} — {proy.nombre}")
    ficha = []
    if proy.modulo:     ficha.append(f"Módulo: {proy.modulo}")
    if proy.area:       ficha.append(f"Área: {proy.area}")
    ficha.append(f"Stack: {proy.tecnologia_principal or 'Python'} / Django")
    if proy.base_datos: ficha.append(f"Base de datos: {proy.base_datos}")
    if proy.app_django: ficha.append(f"App Django: {proy.app_django}")
    if proy.ruta:       ficha.append(f"Ruta: {proy.ruta}")
    if proy.url:        ficha.append(f"URL: {proy.url}")
    if proy.usuarios_beneficiados: ficha.append(f"Usuarios beneficiados: {proy.usuarios_beneficiados}")
    L.append(" · ".join(ficha))
    if proy.descripcion:
        L.append("")
        L.append(f"Descripción: {proy.descripcion.strip()}")
    if proy.problema_que_resuelve:
        L.append(f"Problema que resuelve: {proy.problema_que_resuelve.strip()}")
    L.append("")
    lista = list(reqs)
    L.append(f"# REQUERIMIENTOS A IMPLEMENTAR ({len(lista)})")
    if not lista:
        L.append("(Aún sin requerimientos asociados: define y agrega los requerimientos del proyecto.)")
    for r in lista:
        L.append("")
        L.append(f"## [{r.codigo}] {r.titulo}  (prioridad {r.get_prioridad_display()} · {r.get_estado_display()})")
        if r.descripcion:
            L.append(f"Descripción: {r.descripcion.strip()}")
        if r.problema_actual:
            L.append(f"Problema actual: {r.problema_actual.strip()}")
        if r.resultado_esperado:
            L.append(f"Resultado esperado: {r.resultado_esperado.strip()}")
    L.append("")
    L.append("Entrega el módulo funcionando y validado, e indica los archivos creados/modificados.")
    return "\n".join(L)


@login_required
def proyecto_detail(request, codigo):
    """Detalle del proyecto: ficha técnica + requerimientos + timeline agregado."""
    if es_solo_solicitante(request.user):
        return redirect('desarrollo:mis_requerimientos')
    proy = get_object_or_404(ProyectoDesarrollo.objects.select_related('responsable', 'solicitado_por'),
                             codigo=codigo)
    reqs = (proy.requerimientos
            .select_related('responsable', 'solicitante').order_by('-fecha_creado'))
    total = reqs.count()
    cerrados = sum(1 for r in reqs if r.estado in ESTADOS_CERRADOS)
    avance = round(sum(r.porcentaje_avance for r in reqs) / total) if total else 0
    # Timeline: últimos movimientos de historial de todos los requerimientos del proyecto.
    timeline = (HistorialRequerimiento.objects
                .filter(requerimiento__proyecto=proy)
                .select_related('requerimiento', 'usuario')
                .order_by('-fecha', '-id')[:30])
    return render(request, 'desarrollo/proy_detail.html', {
        'proy': proy, 'reqs': reqs, 'timeline': timeline,
        'stats': {'total': total, 'cerrados': cerrados,
                  'abiertos': total - cerrados, 'avance': avance},
        'prompt_claude': _generar_prompt_claude(proy, reqs),
        'puede_editar': es_admin_desarrollo(request.user) or es_dev_desarrollo(request.user),
        'es_admin': es_admin_desarrollo(request.user), 'nav_home_url': '/',
    })


@login_required
def proyecto_nuevo(request):
    if not (es_admin_desarrollo(request.user) or es_dev_desarrollo(request.user)):
        raise PermissionDenied('No tienes permiso para crear proyectos.')
    if request.method == 'POST':
        form = ProyectoForm(request.POST)
        if form.is_valid():
            proy = form.save(commit=False)
            proy.creado_por = request.user
            proy.modificado_por = request.user
            if not proy.solicitado_por_id:            # <--- hecho por claude code: registrar quién solicitó
                proy.solicitado_por = request.user
            proy.fecha_solicitud = timezone.now().date()
            proy.save()                              # genera TC-PRY-XXXX
            messages.success(request, f'Proyecto {proy.codigo} creado.')
            return redirect('desarrollo:proy_detail', codigo=proy.codigo)
    else:
        form = ProyectoForm()
    return render(request, 'desarrollo/proy_form.html',
                  {'form': form, 'modo': 'nuevo', 'nav_home_url': '/'})


@login_required
def proyecto_editar(request, codigo):
    proy = get_object_or_404(ProyectoDesarrollo, codigo=codigo)
    if not (es_admin_desarrollo(request.user) or es_dev_desarrollo(request.user)):
        raise PermissionDenied('No puedes editar proyectos.')
    if request.method == 'POST':
        form = ProyectoForm(request.POST, instance=proy)
        if form.is_valid():
            proy = form.save(commit=False)
            proy.modificado_por = request.user
            proy.save()
            messages.success(request, f'{proy.codigo} actualizado.')
            return redirect('desarrollo:proy_detail', codigo=proy.codigo)
    else:
        form = ProyectoForm(instance=proy)
    return render(request, 'desarrollo/proy_form.html',
                  {'form': form, 'modo': 'editar', 'proy': proy, 'nav_home_url': '/'})


@login_required
def requerimiento_convertir(request, codigo):
    """Convierte un requerimiento en proyecto (services.convertir_en_proyecto)."""
    req = get_object_or_404(RequerimientoDesarrollo, codigo=codigo)
    if not (es_admin_desarrollo(request.user) or es_dev_desarrollo(request.user)):
        raise PermissionDenied('No tienes permiso para convertir requerimientos en proyectos.')
    if req.proyecto_id:
        messages.info(request, f'Este requerimiento ya pertenece al proyecto {req.proyecto.codigo}.')
        return redirect('desarrollo:proy_detail', codigo=req.proyecto.codigo)

    if request.method == 'POST':
        form = ConvertirProyectoForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            proy, creado = services.convertir_en_proyecto(
                req, request.user,
                nombre=cd['nombre'], modulo=cd['modulo'],
                area=cd['area'], descripcion=cd['descripcion'])
            messages.success(request, f'Requerimiento convertido en proyecto {proy.codigo}.')
            return redirect('desarrollo:proy_detail', codigo=proy.codigo)
    else:
        form = ConvertirProyectoForm(initial={
            'nombre': req.titulo, 'modulo': req.modulo,
            'area': req.area, 'descripcion': req.descripcion})
    return render(request, 'desarrollo/req_convertir.html',
                  {'form': form, 'req': req, 'nav_home_url': '/'})


# ══════════════════════════════ FASE 5 · Roadmap / Mis requerimientos ══════════════════════════════
@login_required
def roadmap(request):
    """Tablero Kanban: requerimientos agrupados por estado (arrastrar para mover)."""
    if es_solo_solicitante(request.user):
        return redirect('desarrollo:mis_requerimientos')
    es_admin = es_admin_desarrollo(request.user)
    qs = (RequerimientoDesarrollo.objects
          .select_related('responsable', 'proyecto').order_by('-fecha_creado'))
    if not es_admin:
        qs = qs.filter(Q(solicitante=request.user) | Q(responsable=request.user))
    proyecto_id = (request.GET.get('proyecto') or '').strip()
    if proyecto_id:
        qs = qs.filter(proyecto_id=proyecto_id)

    por_estado = {clave: [] for clave, _ in EstadoReq.choices}
    for r in qs:
        por_estado[r.estado].append(r)
    columnas = [{'clave': clave, 'label': label, 'items': por_estado[clave]}
                for clave, label in EstadoReq.choices]

    return render(request, 'desarrollo/roadmap.html', {
        'columnas': columnas,
        'puede_mover': es_admin or es_dev_desarrollo(request.user),
        'proyectos': ProyectoDesarrollo.objects.order_by('nombre'),
        'filtros': request.GET,
        'es_admin': es_admin, 'nav_home_url': '/',
    })


@login_required
@require_POST
def requerimiento_mover(request, codigo):
    """Cambia el estado desde el Kanban (drag & drop). Devuelve JSON."""
    req = get_object_or_404(RequerimientoDesarrollo, codigo=codigo)
    if not (es_dev_desarrollo(request.user) or req.responsable_id == request.user.id):
        return JsonResponse({'ok': False, 'error': 'Sin permiso para mover este requerimiento.'}, status=403)
    nuevo = (request.POST.get('estado') or '').strip()
    if nuevo not in {c for c, _ in EstadoReq.choices}:
        return JsonResponse({'ok': False, 'error': 'Estado inválido.'}, status=400)
    if nuevo != req.estado:
        services.registrar_cambio(req, request.user, nuevo_estado=nuevo,
                                  comentario='Movido en el roadmap')
        ev = {EstadoReq.APROBADO: 'aprobado', EstadoReq.PRUEBAS: 'pruebas',
              EstadoReq.PRODUCCION: 'produccion'}.get(nuevo)
        if ev:
            services.notificar_evento(req, ev)
    return JsonResponse({'ok': True, 'estado': req.estado,
                         'display': req.get_estado_display(), 'semaforo': req.semaforo})


@login_required
def mis_requerimientos(request):
    """Los requerimientos donde el usuario es solicitante o responsable."""
    base = RequerimientoDesarrollo.objects.select_related('proyecto', 'solicitante', 'responsable')
    solicitados = base.filter(solicitante=request.user).order_by('-fecha_creado')
    asignados   = base.filter(responsable=request.user).order_by('-fecha_creado')
    return render(request, 'desarrollo/mis_requerimientos.html', {
        'solicitados': solicitados, 'asignados': asignados,
        'es_admin': es_admin_desarrollo(request.user),
        'solo_solicitante': es_solo_solicitante(request.user), 'nav_home_url': '/',
    })


# ══════════════════════════════ FASE 6 · Exportación a Excel ══════════════════════════════
@login_required
def requerimiento_exportar(request):
    """Exporta a .xlsx los requerimientos filtrados (respeta permisos y filtros GET)."""
    if es_solo_solicitante(request.user):
        return redirect('desarrollo:mis_requerimientos')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    qs = _requerimientos_filtrados(request).order_by('codigo')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Requerimientos'
    encabezados = ['Código', 'Título', 'Proyecto', 'Área', 'Módulo', 'Tipo', 'Estado',
                   'Prioridad', 'Impacto', 'Urgencia', 'Solicitante', 'Responsable',
                   'Avance %', 'Semáforo', 'F. solicitud', 'F. estimada', 'F. finalización',
                   'Versión', 'Clasificación']
    ws.append(encabezados)
    cab_fill = PatternFill('solid', fgColor='206BC4')
    cab_font = Font(color='FFFFFF', bold=True)
    for col, _ in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = cab_fill; cell.font = cab_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'

    sem_txt = {'verde': 'En tiempo', 'amarillo': 'Próximo a vencer',
               'rojo': 'Atrasado', 'gris': 'Sin fecha'}

    def _nombre(u):
        if not u:
            return ''
        return u.get_full_name() or u.username

    for r in qs:
        ws.append([
            r.codigo, r.titulo,
            r.proyecto.codigo if r.proyecto else '',
            r.area, r.modulo, r.get_tipo_display(), r.get_estado_display(),
            r.get_prioridad_display(), r.get_impacto_display(), r.get_urgencia_display(),
            _nombre(r.solicitante), _nombre(r.responsable),
            r.porcentaje_avance, sem_txt.get(r.semaforo, ''),
            r.fecha_solicitud, r.fecha_estimada, r.fecha_finalizacion,
            r.version_implementada, r.get_clasificacion_display(),
        ])

    anchos = [15, 40, 12, 14, 14, 16, 18, 10, 8, 12, 20, 20, 9, 16, 12, 12, 14, 12, 22]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and len(cell.value) > 40:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    fecha = timezone.now().strftime('%Y%m%d_%H%M')
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="requerimientos_{fecha}.xlsx"'
    wb.save(resp)
    return resp
